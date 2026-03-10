import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import difflib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_base_dir() -> Path:
    """
    获取程序所在目录。兼容脚本运行和Jupyter运行。
    """
    try:
        # 普通脚本运行
        return Path(__file__).resolve().parent
    except NameError:
        # Jupyter环境
        return Path(".").resolve()


def find_files_by_prefix(base_dir: Path, prefixes: List[str]) -> List[Path]:
    """
    在 base_dir 下查找以指定前缀开头的 .xlsx 文件（不区分大小写）。
    文件名可能正好是“信选”或“清元”，也包括“信选*.xlsx”“清元*.xlsx”。
    """
    targets = []
    for p in base_dir.glob("*.xlsx"):
        fname = p.stem.lower()
        if any(fname.startswith(prefix.lower()) for prefix in prefixes):
            targets.append(p)
    return targets


def make_modified_copy(src: Path) -> Path:
    """
    为 src 生成一个“-修改后.xlsx”副本路径。
    """
    suffix = src.suffix  # .xlsx
    base = src.with_suffix("")  # 去掉后缀
    new_path = base.with_name(f"{base.name}-修改后{suffix}")
    # 复制文件字节内容
    new_path.write_bytes(src.read_bytes())
    print(f"[INFO] 已为 {src.name} 生成副本：{new_path.name}")
    return new_path


def load_mapping_pandian_spec(base_dir: Path) -> Dict[str, str]:
    """
    读取“盘点-规格代码匹配.xlsx”的sheet1的A、B列，返回 {货品编码 -> 规格代码} 映射。
    - 键进行规范化：若“看起来是数字”，整数转为无前导零的整数字符串，小数转为规范浮点字符串。
    - 值保持原文本，写入时再做数值化。
    """
    mapping_file = base_dir / "盘点-规格代码匹配.xlsx"
    if not mapping_file.exists():
        print("[WARN] 未找到 盘点-规格代码匹配.xlsx，规格代码列将无法自动填充。")
        return {}
    def normalize_key(s: str) -> str:
        s = s.strip()
        if re.fullmatch(r"\d+(\.\d+)?", s):
            if "." in s:
                try:
                    return str(float(s))
                except Exception:
                    return s
            try:
                return str(int(float(s)))
            except Exception:
                return s
        return s
    try:
        df = pd.read_excel(mapping_file, sheet_name=0, dtype=str, header=None)
        mapping = {}
        for _, row in df.iterrows():
            raw_key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            val = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if raw_key:
                mapping[normalize_key(raw_key)] = val
        print(f"[INFO] 已加载规格代码映射，共 {len(mapping)} 条。")
        return mapping
    except Exception as e:
        print(f"[ERROR] 读取 盘点-规格代码匹配.xlsx 失败：{e}")
        return {}


def looks_like_number(text: str) -> bool:
    """
    判断文本是否“看起来是数字”（纯数字或小数）。
    """
    if text is None:
        return False
    s = str(text).strip()
    return re.fullmatch(r"\d+(\.\d+)?", s) is not None


def to_number_or_keep(text):
    """
    尝试将文本转换为数值，失败则返回原文本。
    整数返回 int，小数返回 float。
    """
    if text is None:
        return None
    s = str(text).strip()
    if looks_like_number(s):
        try:
            if "." in s:
                return float(s)
            return int(s)
        except Exception:
            return text
    return text


def find_sheet_by_name(wb, name: str):
    """
    在工作簿中查找指定名称的Sheet（完全匹配）。
    """
    if name in wb.sheetnames:
        return wb[name]
    return None


def insert_column_right_of(ws, header_name: str, new_header: str, fill_hex: str = "FFFF00") -> Optional[int]:
    """
    在指定标题列右侧插入新列，设置新列标题为 new_header，且将新列第一个单元格填充颜色。
    返回新列的列索引（1-based）。若未找到标题列返回 None。
    """
    # 假设第一行是标题行
    titles = [cell.value for cell in ws[1]]
    if header_name not in titles:
        print(f"[WARN] 工作表 {ws.title} 未找到列标题：{header_name}")
        return None
    idx = titles.index(header_name) + 1  # 1-based
    insert_at = idx + 1
    ws.insert_cols(insert_at, amount=1)
    ws.cell(row=1, column=insert_at).value = new_header
    ws.cell(row=1, column=insert_at).fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    return insert_at


def convert_column_numbers(ws, header_name: str):
    """
    将指定标题列中“看起来是数字”的文本转换为数值。
    同时设置显示格式避免科学计数法：
    - 整数：0
    - 小数：0.##########
    """
    titles = [cell.value for cell in ws[1]]
    if header_name not in titles:
        print(f"[WARN] 工作表 {ws.title} 未找到列标题：{header_name}")
        return
    col_idx = titles.index(header_name) + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        new_val = to_number_or_keep(cell.value)
        cell.value = new_val
        if isinstance(new_val, int):
            cell.number_format = "0"
        elif isinstance(new_val, float):
            cell.number_format = "0.##########"


def build_group_summary_from_asheet(ws_asheet) -> pd.DataFrame:
    """
    读取Asheet（盘点表）为DataFrame，按“规格代码”分组，汇总“在库数量”“盘点数”，并按盘点数降序。
    """
    # 读取整表为 DataFrame（openpyxl->pandas）
    rows = ws_asheet.iter_rows(values_only=True)
    data = list(rows)
    if not data:
        return pd.DataFrame(columns=["规格代码", "在库数量(求和)", "盘点数(求和)"])
    header = [str(h) if h is not None else "" for h in data[0]]
    df = pd.DataFrame(data[1:], columns=header)

    required_cols = ["规格代码", "在库数量", "盘点数"]
    for c in required_cols:
        if c not in df.columns:
            print(f"[WARN] Asheet 缺失列：{c}，汇总表将可能为空或不完整。")

    # 尝试数值化
    for c in ["在库数量", "盘点数"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if "规格代码" not in df.columns:
        return pd.DataFrame(columns=["规格代码", "在库数量(求和)", "盘点数(求和)"])

    grouped = df.groupby("规格代码", dropna=False).agg({
        "在库数量": "sum" if "在库数量" in df.columns else (lambda x: 0),
        "盘点数": "sum" if "盘点数" in df.columns else (lambda x: 0),
    }).reset_index()

    # 重命名
    grouped = grouped.rename(columns={
        "规格代码": "规格代码",
        "在库数量": "在库数量(求和)",
        "盘点数": "盘点数(求和)",
    })

    # 排序
    if "盘点数(求和)" in grouped.columns:
        grouped = grouped.sort_values(by="盘点数(求和)", ascending=False)

    return grouped


def write_summary_to_asheet(ws_asheet, df_summary: pd.DataFrame, start_cell: str = "M1"):
    """
    将汇总表写入Asheet，从指定起始单元格（默认M1）开始。
    写出时对所有数值统一设置显示格式，避免科学计数法与尾随点：
    - 整数值：0
    - 非整数值：0.##########
    """
    from openpyxl.utils import column_index_from_string, get_column_letter

    start_col = column_index_from_string(re.sub(r"\d", "", start_cell))
    start_row = int(re.sub(r"\D", "", start_cell))

    headers = list(df_summary.columns)
    # 写表头
    for j, h in enumerate(headers):
        ws_asheet.cell(row=start_row, column=start_col + j).value = h

    # 写数据并设置格式
    for i, (_, row) in enumerate(df_summary.iterrows(), start=1):
        for j, h in enumerate(headers):
            val = row[h]
            # 规格代码也做数值化
            if str(h) == "规格代码":
                val = to_number_or_keep(val)
            cell = ws_asheet.cell(row=start_row + i, column=start_col + j)
            cell.value = val
            # 统一格式：整数0，非整数0.##########
            try:
                v2 = to_number_or_keep(val)
                if isinstance(v2, int):
                    cell.number_format = "0"
                elif isinstance(v2, float):
                    if v2.is_integer():
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.##########"
            except Exception:
                # 保底：用非整数格式
                cell.number_format = "0.##########"

    # 统一设置汇总区域列宽为12，避免异常变宽
    for j in range(len(headers)):
        col_letter = get_column_letter(start_col + j)
        ws_asheet.column_dimensions[col_letter].width = 12


def read_bsheet_best_match(base_dir: Path, current_filename_stem: str) -> Optional[Tuple[Path, str, pd.DataFrame]]:
    """
    在 base_dir 下搜索“商品库存导出*.xlsx”，在其所有工作表中找到与当前文件名（不含扩展）相似度最高的sheet。
    返回：(文件路径, sheet名, pandas数据帧)；若找不到任何文件或sheet返回None。
    """
    candidates = list(base_dir.glob("商品库存导出*.xlsx"))
    if not candidates:
        print("[WARN] 未找到商品库存导出*.xlsx 文件。")
        return None

    best_score = -1.0
    best_item = None

    for p in candidates:
        try:
            xls = pd.ExcelFile(p)
            for sheet in xls.sheet_names:
                score = difflib.SequenceMatcher(None, current_filename_stem, sheet).ratio()
                if score > best_score:
                    # 读取该sheet到DataFrame（只关心A、B列）
                    df = pd.read_excel(p, sheet_name=sheet, dtype=str, header=None)
                    best_score = score
                    best_item = (p, sheet, df)
        except Exception as e:
            print(f"[WARN] 读取 {p.name} 失败：{e}")
            continue

    if best_item is None:
        print("[WARN] 未能在商品库存导出*.xlsx 中选择到匹配的Sheet。")
        return None

    print(f"[INFO] Bsheet已选择：文件={best_item[0].name}，Sheet={best_item[1]}，相似度={best_score:.3f}")
    return best_item


def df_ab_to_mapping(df_ab: pd.DataFrame) -> Dict[str, float]:
    """
    将Bsheet的A、B列（A=规格代码，B=可配数(求和)）转换为映射 {规范化规格代码字符串 -> 可配数(float)}。
    规格代码规范化：若看起来是数字，整数用无前导零的整数字符串，小数用规范浮点字符串。
    """
    def normalize_key(s: str) -> str:
        s = s.strip()
        if re.fullmatch(r"\d+(\.\d+)?", s):
            if "." in s:
                try:
                    return str(float(s))
                except Exception:
                    return s
            try:
                return str(int(float(s)))
            except Exception:
                return s
        return s
    mapping = {}
    for _, row in df_ab.iterrows():
        a = row.iloc[0]
        b = row.iloc[1] if len(row) > 1 else None
        key_raw = str(a).strip() if pd.notna(a) else ""
        val = None
        if pd.notna(b):
            try:
                val = float(str(b).replace(",", "").strip())
            except Exception:
                val = None
        if key_raw:
            mapping[normalize_key(key_raw)] = val if val is not None else 0.0
    return mapping


def fill_spec_code_from_mapping(ws_asheet, spec_col_idx: int, code_mapping: Dict[str, str], item_code_col_name: str = "货品编码"):
    """
    用映射填充Asheet中的“规格代码”列（spec_col_idx）。
    - 匹配键规范化处理，避免数字/字符串不一致。
    - 填充后的规格代码值若看起来是数字，转为数值。
    """
    def normalize_key(s: str) -> str:
        s = s.strip()
        if re.fullmatch(r"\d+(\.\d+)?", s):
            if "." in s:
                try:
                    return str(float(s))
                except Exception:
                    return s
            try:
                return str(int(float(s)))
            except Exception:
                return s
        return s

    titles = [cell.value for cell in ws_asheet[1]]
    if item_code_col_name not in titles:
        print(f"[WARN] Asheet 缺失列：{item_code_col_name}，无法填充规格代码。")
        return
    item_idx = titles.index(item_code_col_name) + 1

    for row in range(2, ws_asheet.max_row + 1):
        val = ws_asheet.cell(row=row, column=item_idx).value
        key_norm = ""
        if val is not None:
            s = str(val).strip()
            key_norm = normalize_key(s)
        if key_norm in code_mapping:
            mapped = code_mapping[key_norm]
            new_val = to_number_or_keep(mapped)
            cell = ws_asheet.cell(row=row, column=spec_col_idx)
            cell.value = new_val
            if isinstance(new_val, int):
                cell.number_format = "0"
            elif isinstance(new_val, float):
                cell.number_format = "0.##########"


def extend_summary_with_bsheet(df_summary: pd.DataFrame, b_mapping: Dict[str, float]) -> pd.DataFrame:
    """
    在汇总表上增加“管易可配数”“差异”两列。
    管易可配数通过规格代码 -> 可配数(求和) 映射填充；差异=盘点数(求和)-管易可配数。
    规范化键：数字整数值转为整数字符串，非整数值用 {:g} 去除尾随.0 等。
    """
    df_summary = df_summary.copy()

    def normalize_key_str(s: str) -> str:
        s = s.strip()
        if re.fullmatch(r"\d+(\.\d+)?", s):
            try:
                f = float(s)
                if f.is_integer():
                    return str(int(f))
                return "{:g}".format(f)
            except Exception:
                return s
        return s

    df_summary["管易可配数"] = df_summary["规格代码"].map(
        lambda k: b_mapping.get(normalize_key_str(str(k)), 0.0)
    )
    df_summary["差异"] = df_summary["盘点数(求和)"].fillna(0) - df_summary["管易可配数"].fillna(0)
    return df_summary


def write_extended_summary(ws_asheet, df_extended: pd.DataFrame, start_cell: str = "M1"):
    """
    重写扩展后的汇总表（包含管易可配数与差异）。
    """
    # 直接覆盖写
    write_summary_to_asheet(ws_asheet, df_extended, start_cell=start_cell)


def update_bsheet_with_real_count(b_file: Path, b_sheet_name: str, df_summary: pd.DataFrame):
    """
    在Bsheet中右侧新增“实盘数”“差异”：
    - 实盘数：用Bsheet的A列（规格代码，规范化匹配）匹配Asheet汇总中的“盘点数(求和)”
    - 差异：实盘数 - 可配数(求和)
    直接保存修改到商品库存导出工作簿。
    """
    def normalize_key(s: str) -> str:
        s = s.strip()
        if re.fullmatch(r"\d+(\.\d+)?", s):
            if "." in s:
                try:
                    return str(float(s))
                except Exception:
                    return s
            try:
                return str(int(float(s)))
            except Exception:
                return s
        return s

    try:
        wb_b = load_workbook(b_file)
        if b_sheet_name not in wb_b.sheetnames:
            print(f"[WARN] 工作簿 {b_file.name} 未找到Sheet：{b_sheet_name}")
            return
        ws_b = wb_b[b_sheet_name]

        rows = ws_b.iter_rows(values_only=True)
        data = list(rows)
        if not data:
            print(f"[WARN] Bsheet {b_sheet_name} 为空，跳过更新。")
            return

        # 构建规范化的映射：规格代码 -> 盘点数(求和)
        spec_to_pandian = {}
        if "规格代码" in df_summary.columns and "盘点数(求和)" in df_summary.columns:
            def normalize_key_str(s: str) -> str:
                s = s.strip()
                if re.fullmatch(r"\d+(\.\d+)?", s):
                    try:
                        f = float(s)
                        if f.is_integer():
                            return str(int(f))
                        return "{:g}".format(f)
                    except Exception:
                        return s
                return s
            for _, row in df_summary.iterrows():
                raw = str(row["规格代码"]) if pd.notna(row["规格代码"]) else ""
                key = normalize_key_str(raw)
                spec_to_pandian[key] = float(row["盘点数(求和)"]) if pd.notna(row["盘点数(求和)"]) else 0.0

        max_col = ws_b.max_column
        # 标题
        ws_b.cell(row=1, column=max_col + 1).value = "实盘数"
        ws_b.cell(row=1, column=max_col + 2).value = "差异"

        # 写入数据并设置显示格式：整数0，非整数0.##########
        for r in range(2, ws_b.max_row + 1):
            spec_code = ws_b.cell(row=r, column=1).value  # A列
            kepaisu = ws_b.cell(row=r, column=2).value    # B列 可配数(求和)
            try:
                kepaisu_num = float(str(kepaisu).replace(",", "").strip()) if kepaisu is not None else 0.0
            except Exception:
                kepaisu_num = 0.0

            # 使用更严格的规范化，优先 normalize_key_str（存在时），否则回退 normalize_key
            key = (normalize_key_str(str(spec_code)) if 'normalize_key_str' in locals() else normalize_key(str(spec_code))) if spec_code is not None else ""
            real_count = spec_to_pandian.get(key, 0.0)

            cell_real = ws_b.cell(row=r, column=max_col + 1)
            cell_diff = ws_b.cell(row=r, column=max_col + 2)
            # 赋值
            cell_real.value = real_count
            diff_val = real_count - kepaisu_num
            cell_diff.value = diff_val

            # 实盘数格式
            try:
                rc = float(real_count)
                if rc.is_integer():
                    cell_real.number_format = "0"
                else:
                    cell_real.number_format = "0.##########"
            except Exception:
                cell_real.number_format = "0.##########"

            # 差异格式
            try:
                dv = float(diff_val)
                if dv.is_integer():
                    cell_diff.number_format = "0"
                else:
                    cell_diff.number_format = "0.##########"
            except Exception:
                cell_diff.number_format = "0.##########"

        # 统一设置新增列列宽为12
        from openpyxl.utils import get_column_letter
        ws_b.column_dimensions[get_column_letter(max_col + 1)].width = 12
        ws_b.column_dimensions[get_column_letter(max_col + 2)].width = 12

        wb_b.save(b_file)
        print(f"[INFO] 已更新并保存 {b_file.name} 的Sheet：{b_sheet_name}")
    except Exception as e:
        print(f"[ERROR] 更新Bsheet失败：{e}")


def process_one_file(modified_path: Path, base_dir: Path, code_mapping: Dict[str, str]):
    """
    处理一个“-修改后.xlsx”文件：
    - 保护旧“差异”列的值（用data_only缓存值覆盖公式）
    - 转换“货品编码”列数值
    - 插入“规格代码”列并填充（填充值数值化）
    - 将“规格代码”列整体数值化
    - 生成汇总并写入M1（写出的规格代码数值化）
    - 匹配Bsheet，扩展汇总并写回
    - 更新Bsheet（商品库存导出）新增“实盘数”“差异”（规范化匹配）
    """
    try:
        wb = load_workbook(modified_path)
        wb_values = load_workbook(modified_path, data_only=True)
    except Exception as e:
        print(f"[ERROR] 打开 {modified_path.name} 失败：{e}")
        return

    ws_asheet = find_sheet_by_name(wb, "盘点表")
    ws_vals = find_sheet_by_name(wb_values, "盘点表")
    if ws_asheet is None or ws_vals is None:
        print(f"[WARN] 文件 {modified_path.name} 未找到“盘点表”Sheet，跳过。")
        return

    # 0) 预采集旧“差异”列缓存值（插入列前采集）
    def capture_old_diff_values(ws_vals_local):
        titles_vals = [cell.value for cell in ws_vals_local[1]]
        if "差异" not in titles_vals:
            return None
        diff_idx_vals = titles_vals.index("差异") + 1
        cache = []
        for r in range(2, ws_vals_local.max_row + 1):
            cache.append(ws_vals_local.cell(row=r, column=diff_idx_vals).value)
        return cache

    old_diff_cache = capture_old_diff_values(ws_vals)

    # 1) 先去除“货品编码”左右空白，再转为数值
    titles_now = [cell.value for cell in ws_asheet[1]]
    if "货品编码" in titles_now:
        col_idx_code = titles_now.index("货品编码") + 1
        for r in range(2, ws_asheet.max_row + 1):
            c = ws_asheet.cell(row=r, column=col_idx_code)
            if c.value is not None:
                c.value = str(c.value).strip()
    convert_column_numbers(ws_asheet, "货品编码")

    # 2) 在右侧插入“规格代码”列并填充颜色与数据
    spec_col_idx = insert_column_right_of(ws_asheet, "货品编码", "规格代码", fill_hex="FFFF00")
    if spec_col_idx:
        from openpyxl.utils import get_column_letter
        spec_letter = get_column_letter(spec_col_idx)
        ws_asheet.column_dimensions[spec_letter].width = 12
        fill_spec_code_from_mapping(ws_asheet, spec_col_idx, code_mapping, item_code_col_name="货品编码")
        # 同步将规格代码列整体进行数值化
        convert_column_numbers(ws_asheet, "规格代码")

    # 2.5) 用缓存值覆盖旧“差异”列（插入后按标题重新定位）
    if old_diff_cache is not None:
        titles_now = [cell.value for cell in ws_asheet[1]]
        if "差异" in titles_now:
            diff_idx_now = titles_now.index("差异") + 1
            for i, r in enumerate(range(2, ws_asheet.max_row + 1)):
                if i < len(old_diff_cache):
                    ws_asheet.cell(row=r, column=diff_idx_now).value = to_number_or_keep(old_diff_cache[i])
        else:
            print("[WARN] 插入后未找到旧“差异”列标题，无法覆盖缓存值。")
    else:
        print("[WARN] 未获取到旧“差异”列的缓存值（可能Excel未保存公式结果），保留原列。")

    # 3) 构建分组汇总并写入M1
    df_summary = build_group_summary_from_asheet(ws_asheet)
    write_summary_to_asheet(ws_asheet, df_summary, start_cell="M1")

    # 4) 查找Bsheet（商品库存导出*.xlsx）
    stem = modified_path.stem.replace("-修改后", "")
    b_info = read_bsheet_best_match(base_dir, stem)
    if b_info is None:
        wb.save(modified_path)
        print(f"[INFO] 已保存 {modified_path.name}（无Bsheet，跳过后续扩展）")
        return

    b_file, b_sheet_name, df_b = b_info
    b_mapping = df_ab_to_mapping(df_b)

    # 5) 在Asheet扩展汇总：增加“管易可配数”“差异”
    df_extended = extend_summary_with_bsheet(df_summary, b_mapping)
    write_extended_summary(ws_asheet, df_extended, start_cell="M1")

    wb.save(modified_path)
    print(f"[INFO] 已保存 {modified_path.name} 的Asheet扩展汇总。")

    # 6) 更新Bsheet：新增“实盘数”“差异”
    update_bsheet_with_real_count(b_file, b_sheet_name, df_extended)


def main():
    base_dir = get_base_dir()
    print(f"[INFO] 程序目录：{base_dir}")

    # 加载规格代码映射
    code_mapping = load_mapping_pandian_spec(base_dir)

    # 搜索目标文件
    targets = find_files_by_prefix(base_dir, prefixes=["信选", "清元"])
    if not targets:
        print("[WARN] 未在当前目录找到以“信选”或“清元”开头的xlsx文件。")
        return

    # 为每个目标文件创建副本并处理
    for src in targets:
        try:
            modified = make_modified_copy(src)
            process_one_file(modified, base_dir, code_mapping)
        except Exception as e:
            print(f"[ERROR] 处理文件 {src.name} 失败：{e}")

    print("[INFO] 全部任务完成。")


if __name__ == "__main__":
    # 允许在Jupyter中使用：%run process_inventory.py
    main()