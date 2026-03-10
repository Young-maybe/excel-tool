"""
盘点的初步处理 - 管易基础表处理工具1

说明：
- 你给的“源文件文件夹”后续会删除，所以这里把原脚本逻辑直接迁入 modules 内部，不再依赖外部源文件。
- 仅做 I/O 适配：输入由 UI 选取文件夹；输出写到程序统一输出目录下的独立子目录。
- 核心处理逻辑保持与原脚本一致。
"""

from __future__ import annotations

import sys
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def read_csv_smart(csv_path: Path) -> pd.DataFrame:
    last_err: Optional[Exception] = None
    df: Optional[pd.DataFrame] = None
    for enc in ["utf-8-sig", "gb18030"]:
        try:
            df = pd.read_csv(csv_path, encoding=enc, dtype=str)
            break
        except Exception as e:
            last_err = e
            df = None
    if df is None:
        raise RuntimeError(f"无法读取CSV文件: {csv_path.name}, 错误: {last_err}")

    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    obj_cols = df.select_dtypes(include=["object"]).columns
    for c in obj_cols:
        df[c] = df[c].str.strip()
    return df


def load_mapping(mapping_xlsx: Path) -> Dict[str, str]:
    if not mapping_xlsx.exists():
        print(f"[警告] 未找到映射文件: {mapping_xlsx.name}，将跳过“仓库主体”填充。")
        return {}

    try:
        mdf = pd.read_excel(mapping_xlsx, sheet_name="Sheet1", dtype=str)
    except Exception:
        mdf = pd.read_excel(mapping_xlsx, dtype=str)

    mdf.columns = [c.strip() if isinstance(c, str) else c for c in mdf.columns]
    if mdf.shape[1] < 2:
        print(f"[警告] 映射文件列数不足（需要至少两列A、B），实际列数: {mdf.shape[1]}。")
        return {}

    colA = mdf.columns[0]
    colB = mdf.columns[1]
    mdf[colA] = mdf[colA].fillna("").astype(str).map(lambda x: x.strip())
    mdf[colB] = mdf[colB].fillna("").astype(str).map(lambda x: x.strip())

    mapping: Dict[str, str] = {}
    for a, b in zip(mdf[colA], mdf[colB]):
        if a != "":
            mapping[a] = b
    return mapping


def insert_after(df: pd.DataFrame, target_col: str, new_col: str, values=None) -> pd.DataFrame:
    if target_col in df.columns:
        idx = list(df.columns).index(target_col) + 1
    else:
        idx = len(df.columns)
    df.insert(idx, new_col, values)
    return df


def swap_columns(df: pd.DataFrame, col_a: str, col_b: str) -> pd.DataFrame:
    cols = list(df.columns)
    if col_a in cols and col_b in cols:
        i, j = cols.index(col_a), cols.index(col_b)
        cols[i], cols[j] = cols[j], cols[i]
        df = df[cols]
    else:
        if col_a not in cols:
            print(f"[提示] 未找到列: {col_a}，跳过交换。")
        if col_b not in cols:
            print(f"[提示] 未找到列: {col_b}，跳过交换。")
    return df


def enforce_two_barcodes_pandas(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    ku_idx = cols.index("库位") if "库位" in cols else None
    bar_idxs = [i for i, c in enumerate(cols) if c == "商品条码"]

    if bar_idxs:
        src_idx = None
        if ku_idx is not None:
            for i in bar_idxs:
                if i > ku_idx:
                    src_idx = i
                    break
        if src_idx is None:
            src_idx = bar_idxs[0]
        src_series = df.iloc[:, src_idx].astype(str)
    else:
        src_series = pd.Series([""] * len(df), index=df.index, dtype=object)

    base_pos = [i for i, c in enumerate(cols) if c != "商品条码"]
    df_base = df.iloc[:, base_pos].copy()
    base_names = list(df_base.columns)

    parts: List[pd.Series] = []
    names: List[str] = []

    start_mid = 0
    if len(df_base.columns) >= 1:
        parts.append(df_base.iloc[:, 0])
        names.append(df_base.columns[0])
        start_mid = 1

    parts.append(src_series)
    names.append("商品条码")

    if "库位" in base_names:
        ku_pos = base_names.index("库位")
        mid_end = ku_pos + 1
    else:
        mid_end = len(df_base.columns)
    for j in range(start_mid, mid_end):
        parts.append(df_base.iloc[:, j])
        names.append(df_base.columns[j])

    has_ku = "库位" in base_names
    if has_ku:
        parts.append(src_series)
        names.append("商品条码")

    for j in range(mid_end, len(df_base.columns)):
        parts.append(df_base.iloc[:, j])
        names.append(df_base.columns[j])

    if not has_ku:
        parts.append(src_series)
        names.append("商品条码")

    df_final = pd.concat(parts, axis=1)
    df_final.columns = names

    bar_pos = [i for i, n in enumerate(names) if n == "商品条码"]
    if len(bar_pos) > 2:
        keep = {1}
        for i in bar_pos:
            if i != 1:
                keep.add(i)
                break
        keep_list = sorted(keep)
        rebuilt_parts = []
        rebuilt_names = []
        for i, n in enumerate(names):
            if n == "商品条码" and i not in keep_list:
                continue
            rebuilt_parts.append(df_final.iloc[:, i])
            rebuilt_names.append(n)
        df_final = pd.concat(rebuilt_parts, axis=1)
        df_final.columns = rebuilt_names

    return df_final


def drop_barcode_2_and_4(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    barcode_positions = [i for i, c in enumerate(cols) if c == "商品条码"]
    to_drop_idx: List[int] = []
    if len(barcode_positions) >= 2:
        to_drop_idx.append(barcode_positions[1])
    if len(barcode_positions) >= 4:
        to_drop_idx.append(barcode_positions[3])
    if not to_drop_idx:
        return df
    drop_set = set(to_drop_idx)
    keep_indices = [i for i in range(len(cols)) if i not in drop_set]
    return df.iloc[:, keep_indices].copy()


def style_and_hide_columns(xlsx_path: Path, header_fill_targets: List[str], hide_cols: List[str]) -> None:
    wb = load_workbook(xlsx_path)
    ws: Worksheet = cast(Worksheet, wb.active)

    headers: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        key = val.strip() if isinstance(val, str) else val
        if key is not None and key != "":
            headers.setdefault(key, col_idx)

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for name in header_fill_targets:
        if name == "商品条码":
            v2 = ws.cell(row=1, column=2).value
            v2s = v2.strip() if isinstance(v2, str) else v2
            if v2s == "商品条码":
                ws.cell(row=1, column=2).fill = fill
                continue
        if name in headers:
            ws.cell(row=1, column=headers[name]).fill = fill
        else:
            print(f"[提示] 未找到需标色列: {name}")

    for name in hide_cols:
        if name == "商品条码":
            print("[提示] 已忽略隐藏请求：商品条码 不应被隐藏")
            continue
        if name in headers:
            col_letter = get_column_letter(headers[name])
            ws.column_dimensions[col_letter].hidden = True
        else:
            print(f"[提示] 未找到需隐藏列: {name}")

    wb.save(xlsx_path)


def tweak_formats_post(xlsx_path: Path) -> None:
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"[提示] 打开目标工作簿失败: {e}")
        return

    ws_main: Worksheet = cast(Worksheet, wb.active)

    def clear_header_border(ws: Worksheet) -> None:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).border = Border()

    clear_header_border(ws_main)
    from openpyxl.utils import get_column_letter as _gcl

    for col_idx, width in [(2, 16.5), (3, 16.5)]:
        if col_idx <= ws_main.max_column:
            ws_main.column_dimensions[_gcl(col_idx)].width = width

    def to_number_if_numeric(s: Any) -> Any:
        if isinstance(s, str):
            t = s.strip()
            if t == "":
                return s
            try:
                if t.isdigit() or (t.startswith("-") and t[1:].isdigit()):
                    return int(t)
                return float(t)
            except Exception:
                return s
        return s

    for col_idx in (2, 3, 17):
        if ws_main.max_row >= 2 and col_idx <= ws_main.max_column:
            for r in range(2, ws_main.max_row + 1):
                cell = ws_main.cell(row=r, column=col_idx)
                cell.value = to_number_if_numeric(cell.value)

    for name in ["信选", "清元", "信选次品", "清元次品"]:
        if name not in wb.sheetnames:
            continue
        ws = cast(Worksheet, wb[name])
        clear_header_border(ws)
        ws.column_dimensions[_gcl(1)].width = 18.5
        ws.column_dimensions[_gcl(2)].width = 14

        if ws.max_row >= 2:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=1)
                cell.value = to_number_if_numeric(cell.value)

    try:
        wb.save(xlsx_path)
    except Exception as e:
        print(f"[提示] 保存美化后的工作簿失败: {e}")


def create_pivots(xlsx_path: Path, df: pd.DataFrame) -> None:
    targets = ["信选", "清元", "信选次品", "清元次品"]
    need_cols = {"仓库主体", "商品规格代码", "可配数"}
    if not need_cols.issubset(set(df.columns)):
        print("[提示] 生成透视表所需列缺失（需要：仓库主体、商品规格代码、可配数），已跳过透视表生成。")
        return

    work = df.copy()
    work["可配数"] = cast(pd.Series, pd.to_numeric(work["可配数"], errors="coerce")).fillna(0.0).astype(float)

    outputs: Dict[str, pd.DataFrame] = {}
    for name in targets:
        sub = work[work["仓库主体"] == name]
        if sub.empty:
            outputs[name] = pd.DataFrame(columns=["商品规格代码", "可配数(求和)"])
            continue
        agg = sub.groupby("商品规格代码", dropna=False, sort=False)["可配数"].sum().reset_index()
        agg.rename(columns={"可配数": "可配数(求和)"}, inplace=True)
        agg = agg.sort_values(by="可配数(求和)", ascending=False, kind="mergesort", ignore_index=True)
        outputs[name] = agg

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            for sheet_name, df_out in outputs.items():
                df_out.to_excel(writer, sheet_name=sheet_name, index=False)
        print("[透视] 已生成工作表: " + "、".join(targets))
    except Exception as e:
        print(f"[提示] 写入透视表失败: {e}")


def save_df_to_excel_return_path(df: pd.DataFrame, target_path: Path) -> Path:
    for i in range(100):
        cand = target_path if i == 0 else target_path.with_name(f"{target_path.stem}_{i}{target_path.suffix}")
        try:
            try:
                df.to_excel(cand, index=False)
            except Exception:
                df.to_excel(cand, index=False, engine="openpyxl")
            if i > 0:
                print(f"[提示] 目标文件被占用，已改名保存为: {cand.name}")
            return cand
        except PermissionError:
            continue
        except OSError as e:
            if getattr(e, "errno", None) == 13:
                continue
            raise
    raise RuntimeError(f"无法保存Excel文件（目标被占用或无权限）：{target_path.name}")


def process_csv(csv_path: Path, mapping: Dict[str, str]) -> Optional[Path]:
    print(f"[处理] {csv_path.name}")
    df = read_csv_smart(csv_path)

    df = enforce_two_barcodes_pandas(df)
    df = swap_columns(df, "商品规格代码", "商品名称")

    if "商品标签" in df.columns:
        df = df.drop(columns=["商品标签"])

    df = insert_after(df, "仓库", "仓库主体", values="")

    if "仓库" in df.columns and "仓库主体" in df.columns and mapping:
        df["仓库主体"] = df["仓库"].map(lambda x: mapping.get(x, "") if isinstance(x, str) else "")
    elif not mapping:
        print("[提示] 未加载到有效映射，跳过“仓库主体”填充。")
    else:
        print("[提示] 缺少“仓库”或“仓库主体”列，跳过填充。")

    if "仓库主体" in df.columns:
        before = len(df)
        df["仓库主体"] = df["仓库主体"].fillna("").map(lambda x: x.strip() if isinstance(x, str) else x)
        df = df[df["仓库主体"] != ""].copy()
        after = len(df)
        print(f"[过滤] 移除“仓库主体”为空的行: {before - after} 条")
    else:
        print("[提示] 未找到“仓库主体”列，无法按空值过滤行。")

    df = drop_barcode_2_and_4(cast(pd.DataFrame, df))

    desired_path = csv_path.with_suffix(".xlsx")
    xlsx_path = save_df_to_excel_return_path(df, desired_path)

    header_fill_targets = ["商品条码", "商品规格代码", "仓库主体", "可销数", "可配数"]
    hide_cols = ["未付款数", "锁定数", "库位"]
    try:
        style_and_hide_columns(xlsx_path, header_fill_targets, hide_cols)
    except Exception as e:
        print(f"[提示] 样式或隐藏列处理失败: {e}")

    try:
        create_pivots(xlsx_path, df)
    except Exception as e:
        print(f"[提示] 透视表生成失败: {e}")

    try:
        tweak_formats_post(xlsx_path)
    except Exception as e:
        print(f"[提示] 表格美化失败: {e}")

    print(f"[完成] 生成: {xlsx_path.name}")
    return xlsx_path


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb=None,
) -> List[str]:
    def report(p: int, s: str) -> None:
        if progress_cb:
            try:
                progress_cb(int(p), str(s))
            except Exception:
                pass

    in_dir = Path(input_dir)
    out_dir = Path(output_dir)
    if not in_dir.exists():
        raise FileNotFoundError(f"输入目录不存在：{input_dir}")

    work_dir = out_dir / "盘点的初步处理" / "管易基础表处理工具1"
    work_dir.mkdir(parents=True, exist_ok=True)

    pattern = "商品库存导出*.csv"
    report(5, "扫描输入 CSV…")
    csv_files = sorted(in_dir.glob(pattern))
    if not csv_files:
        raise FileNotFoundError(f"未在输入文件夹中找到匹配文件：{pattern}")

    mapping_file = in_dir / "盘点-仓库货主匹配.xlsx"
    mapping = load_mapping(mapping_file)

    outputs: List[str] = []
    total = len(csv_files)
    for i, csv_path in enumerate(csv_files, start=1):
        base = 10 + int((i - 1) / max(total, 1) * 80)
        report(base, f"处理 {i}/{total}：{csv_path.name}")
        dst = work_dir / csv_path.name
        report(min(base + 3, 90), "复制 CSV 到工作目录…")
        try:
            shutil.copy2(csv_path, dst)
        except PermissionError as exc:
            raise PermissionError(f"文件正在被占用，请先关闭Excel/WPS后重试：{csv_path}") from exc
        report(min(base + 8, 92), "读取 CSV 并清洗字段…")
        report(min(base + 13, 95), "计算并生成 Excel（样式/透视/美化）…")
        out_xlsx = process_csv(dst, mapping)
        if out_xlsx is not None:
            outputs.append(str(out_xlsx))

    report(100, "完成")
    return outputs


