import sys
from pathlib import Path
from typing import Dict, Optional, Tuple, Any
from datetime import datetime
import warnings
from copy import copy

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import MergedCell
    import pandas as pd
except ImportError:
    print("缺少依赖，请先安装：pip install openpyxl pandas")
    raise


def get_base_dir() -> Path:
    """
    获取脚本运行的基准目录：
    - 脚本环境：使用脚本所在目录
    - Jupyter/交互环境：使用当前工作目录
    """
    if hasattr(sys, "frozen") and getattr(sys, "frozen"):
        # 如果是打包后可执行
        return Path(sys.executable).parent
    try:
        base = Path(__file__).parent.resolve()
    except NameError:
        base = Path.cwd().resolve()
    return base


def find_latest_inventory_file(base_dir: Path) -> Optional[Path]:
    """
    在 base_dir 下查找匹配“库存快照明细*.xlsx”的最新修改文件
    """
    candidates = list(base_dir.glob("库存快照明细*.xlsx"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def is_number_like(text: str) -> bool:
    """
    判断文本是否“看起来是数字”：
    - 去除首尾空格
    - 去除千分位逗号
    - 允许正负号与小数点
    - 不包含其他字母或特殊字符
    """
    s = text.strip().replace(",", "")
    if s == "":
        return False
    # 允许单个小数点和正负号
    # 快速排除包含字母的情况
    if any(c.isalpha() for c in s):
        return False
    # 允许形如 +123, -123.45, .5, 0. 等
    try:
        float(s)
        return True
    except ValueError:
        return False


def to_number(text: str):
    """
    将文本转为数字：
    - 优先转为 int（如果与 int 等价）
    - 否则为 float
    """
    s = text.strip().replace(",", "")
    # 处理仅为"."的情况
    if s == "." or s == "+." or s == "-.":
        # 这类不应视为有效数字
        return text
    try:
        f = float(s)
        # 判断是否整数（避免 1.0000001 这种浮点误差）
        if abs(f - round(f)) < 1e-12:
            return int(round(f))
        return f
    except ValueError:
        return text


def load_mapping_from_match_file(base_dir: Path) -> Dict[str, str]:
    """
    保留原函数，仅加载“文本键”映射，便于兼容。
    """
    match_path = base_dir / "盘点-规格代码匹配.xlsx"
    if not match_path.exists():
        print(f"未找到映射文件：{match_path}")
        return {}
    wb = load_workbook(match_path, data_only=True)
    if "sheet1" in wb.sheetnames:
        mapping_ws = wb["sheet1"]
    else:
        mapping_ws = wb.active
    if not isinstance(mapping_ws, Worksheet):
        print("映射文件的活动表类型异常，跳过映射加载。")
        return {}
    mapping: Dict[str, str] = {}
    for idx, row in enumerate(mapping_ws.iter_rows(min_row=1, values_only=True), start=1):
        code = row[0]
        spec = row[1] if len(row) > 1 else None
        if code is None or spec is None:
            continue
        if isinstance(code, str) and code.strip() in {"货品编码", "编码", "item_code", "ItemCode"}:
            continue
        key = str(code).strip()
        val = str(spec).strip()
        if key:
            mapping[key] = val
    return mapping

def load_mappings(base_dir: Path) -> Tuple[Dict[Any, str], Dict[str, str]]:
    """
    新增：同时加载“数字键映射”和“文本键映射”
    - 数字键映射：A列值若为数字或看起来是数字，则转换为 int/float 作为键
    - 文本键映射：A列值统一为去空格的字符串
    """
    match_path = base_dir / "盘点-规格代码匹配.xlsx"
    if not match_path.exists():
        print(f"未找到映射文件：{match_path}")
        return {}, {}
    wb = load_workbook(match_path, data_only=True)
    mapping_ws = wb["sheet1"] if "sheet1" in wb.sheetnames else wb.active
    if not isinstance(mapping_ws, Worksheet):
        print("映射文件的活动表类型异常，跳过映射加载。")
        return {}, {}
    map_num: Dict[Any, str] = {}
    map_text: Dict[str, str] = {}
    for idx, row in enumerate(mapping_ws.iter_rows(min_row=1, values_only=True), start=1):
        code = row[0]
        spec = row[1] if len(row) > 1 else None
        if code is None or spec is None:
            continue
        if isinstance(code, str) and code.strip() in {"货品编码", "编码", "item_code", "ItemCode"}:
            continue
        spec_str = str(spec).strip()
        # 文本键
        text_key = str(code).strip()
        if text_key:
            map_text[text_key] = spec_str
        # 数字键
        num_key: Any = None
        if isinstance(code, (int, float)):
            num_key = code
        elif isinstance(code, str) and is_number_like(code):
            num_key = to_number(code)
        if num_key is not None:
            map_num[num_key] = spec_str
    return map_num, map_text


def find_header_column(ws: "Worksheet", header_name: str) -> Optional[int]:
    """
    在第一行查找指定表头名，返回列索引（1-based）
    """
    for cell in ws[1]:
        if cell.value is None:
            continue
        if str(cell.value).strip() == header_name:
            return cell.column  # 1-based
    return None


def insert_column_right_of(ws: "Worksheet", col_idx: int, header: str) -> int:
    """
    在指定列右侧插入新列并写入表头，返回新列的列索引（1-based）
    """
    new_col_idx = col_idx + 1
    ws.insert_cols(new_col_idx, amount=1)
    ws.cell(row=1, column=new_col_idx, value=header)
    return new_col_idx


def hide_columns_range(ws: "Worksheet", start_letter: str, end_letter: str):
    """
    隐藏从 start_letter 到 end_letter 的列（包含两端）
    """
    def col_letter_range(start: str, end: str):
        import string
        letters = list(string.ascii_uppercase)
        # 支持到两位列字母（例如 AA, AB），这里简化处理到单字母即可满足 I-M
        start_idx = letters.index(start)
        end_idx = letters.index(end)
        for i in range(start_idx, end_idx + 1):
            yield letters[i]

    for letter in col_letter_range(start_letter, end_letter):
        ws.column_dimensions[letter].hidden = True


def apply_header_fill(ws: "Worksheet", col_idx: int, hex_color: str = "FFFF00"):
    """
    为指定列的表头单元格填充颜色
    """
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    ws.cell(row=1, column=col_idx).fill = fill

def format_number_cell(cell):
    """
    将单元格设为非科学计数法显示（与主表C/D列一致，不使用货币格式）：
    - int: "0"
    - float: "0.####################"
    """
    v = cell.value
    if isinstance(v, int):
        cell.number_format = "0"
    elif isinstance(v, float):
        cell.number_format = "0.####################"

def set_column_widths(ws: "Worksheet", widths: Dict[str, float]):
    """
    批量设置列宽：传入如 {"C":16, "D":16}
    """
    for letter, width in widths.items():
        try:
            ws.column_dimensions[letter].width = width
        except Exception:
            # 忽略异常，确保不影响主流程
            pass

def copy_cell_style(src, dst, exclude: Tuple[str, ...] = ("fill",)):
    """
    复制单元格样式（字体、对齐、边框、数字格式等），默认排除填充以便后续上色。
    使用浅拷贝避免样式对象共享导致的未生效问题。
    """
    try:
        # 字体
        if "font" not in exclude and getattr(src, "font", None):
            try:
                dst.font = copy(src.font)
            except Exception:
                dst.font = src.font
        # 对齐
        if "alignment" not in exclude and getattr(src, "alignment", None):
            try:
                dst.alignment = copy(src.alignment)
            except Exception:
                dst.alignment = src.alignment
        # 边框
        if "border" not in exclude and getattr(src, "border", None):
            try:
                dst.border = copy(src.border)
            except Exception:
                dst.border = src.border
        # 数字格式
        if "number_format" not in exclude and getattr(src, "number_format", None):
            dst.number_format = src.number_format
        # 填充（通常排除，之后再统一上色）
        if "fill" not in exclude and getattr(src, "fill", None):
            try:
                dst.fill = copy(src.fill)
            except Exception:
                dst.fill = src.fill
    except Exception:
        # 某些工作簿缺省样式可能引发警告或异常，忽略以保证流程继续
        pass

def get_c_column_number_format(ws: "Worksheet") -> Optional[str]:
    """
    获取主表 C 列首个有效数据单元格的 number_format，若不可用返回 None
    """
    try:
        max_row = ws.max_row or 1
        for r in range(2, max_row + 1):
            cell = ws.cell(row=r, column=3)
            fmt = getattr(cell, "number_format", None)
            if cell.value is not None and isinstance(fmt, str) and fmt.strip() and fmt.strip().lower() != "general":
                return fmt
        return None
    except Exception:
        return None

def load_product_name_mapping(base_dir: Path) -> Tuple[Dict[Any, str], Dict[str, str]]:
    """
    查找最新“商品库存导出*.xlsx”，读取活动表 C/D 列，构建：
    - 数值键映射：C列数值键 -> D列商品名称
    - 文本键映射：C列文本键(去空格) -> D列商品名称
    """
    candidates = list(base_dir.glob("商品库存导出*.xlsx"))
    if not candidates:
        return {}, {}
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    path = candidates[0]
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        if not isinstance(ws, Worksheet):
            return {}, {}
        map_num: Dict[Any, str] = {}
        map_text: Dict[str, str] = {}
        max_row = ws.max_row or 1
        for r in range(2, max_row + 1):
            key_val = ws.cell(row=r, column=3).value  # C
            name_val = ws.cell(row=r, column=4).value  # D
            if key_val is None or name_val is None:
                continue
            name = str(name_val).strip()
            text_key = str(key_val).strip()
            if text_key:
                map_text[text_key] = name
            num_key: Any = None
            if isinstance(key_val, (int, float)):
                num_key = key_val
            elif isinstance(key_val, str) and is_number_like(key_val):
                num_key = to_number(key_val)
            if num_key is not None:
                map_num[num_key] = name
        return map_num, map_text
    except Exception:
        return {}, {}

def build_pivot_sheet_enhanced(wb, src_ws: "Worksheet", base_dir: Optional[Path] = None):
    """
    在新建 sheet '透视表' 中，从 B 列开始生成数据透视表：
    - 行：规格代码；列：货主；值：在库数量（求和）
    - 最右侧添加“总计”列，并按该列降序排序
    - A 列“商品名称”基于“商品库存导出*.xlsx”的 C/D 映射按规格代码填充
    - 数值列格式继承主表 C 列的 number_format，若不可用则退回为非科学计数法格式
    """
    # 定位字段
    col_spec = find_header_column(src_ws, "规格代码")
    col_owner = find_header_column(src_ws, "货主")
    col_qty = find_header_column(src_ws, "在库数量")
    if not all([col_spec, col_owner, col_qty]):
        print("生成透视表失败：未找到'规格代码'、'货主'或'在库数量'列。")
        return

    # 主表 C 列数字格式
    c_fmt = get_c_column_number_format(src_ws)
    # 商品名称映射
    # - 按原始脚本：默认使用 get_base_dir()
    # - 在本项目“复制到 work_dir 再处理”的模式下：传入 base_dir=work_dir，确保能找到同目录的 商品库存导出*.xlsx
    if base_dir is None:
        base_dir = get_base_dir()
    name_map_num, name_map_text = load_product_name_mapping(base_dir)

    # 收集数据
    max_row = src_ws.max_row or 1
    records = []
    for r in range(2, max_row + 1):
        spec = src_ws.cell(row=r, column=col_spec).value
        owner = src_ws.cell(row=r, column=col_owner).value
        qty = src_ws.cell(row=r, column=col_qty).value
        if spec is None or owner is None:
            continue
        if isinstance(qty, (int, float)):
            v = float(qty)
        elif isinstance(qty, str) and is_number_like(qty):
            v = float(to_number(qty))
        else:
            v = 0.0
        records.append({"规格代码": str(spec).strip(), "货主": str(owner).strip(), "在库数量": v})
    if not records:
        print("生成透视表失败：无有效数据。")
        return

    import pandas as pd  # 再次导入以保险
    df = pd.DataFrame.from_records(records)
    pv = pd.pivot_table(df, index="规格代码", columns="货主", values="在库数量", aggfunc="sum", fill_value=0)
    pv["__行总计__"] = pv.sum(axis=1)
    pv = pv.sort_values(by="__行总计__", ascending=False)
    totals_series = pv["__行总计__"].copy()
    pv_out = pv.drop(columns="__行总计__")

    # 写入
    ws_pivot = wb.create_sheet("透视表")
    ws_pivot.cell(row=1, column=1, value="商品名称")  # A1
    ws_pivot.cell(row=1, column=2, value="规格代码")  # B1

    # 货主列标题 + 总计
    col_names = [str(c) for c in pv_out.columns.tolist()]
    for j, name in enumerate(col_names, start=0):
        ws_pivot.cell(row=1, column=3 + j, value=name)
    total_col_index = 3 + len(col_names)
    ws_pivot.cell(row=1, column=total_col_index, value="总计")

    # 数据行
    for i, (spec_code, row_series) in enumerate(pv_out.iterrows(), start=2):
        # 规格代码处理：数字转数值，文本保留
        if isinstance(spec_code, str) and is_number_like(spec_code):
            spec_out = to_number(spec_code)
        else:
            spec_out = spec_code
        spec_cell = ws_pivot.cell(row=i, column=2, value=spec_out)
        format_number_cell(spec_cell)

        # 商品名称填充（优先数值键，再文本键）
        name_val = None
        if isinstance(spec_out, (int, float)):
            name_val = name_map_num.get(spec_out)
        if name_val is None:
            name_val = name_map_text.get(str(spec_code).strip())
        if name_val is None and isinstance(spec_code, str) and is_number_like(spec_code):
            name_val = name_map_num.get(to_number(spec_code))
        ws_pivot.cell(row=i, column=1, value=name_val)

        # 各货主值
        for j, name in enumerate(col_names, start=0):
            val = float(row_series.get(name, 0.0))
            cell = ws_pivot.cell(row=i, column=3 + j, value=val)
            if c_fmt:
                cell.number_format = c_fmt
            else:
                format_number_cell(cell)

        # 总计
        total_val = float(totals_series.get(spec_code, 0.0))
        total_cell = ws_pivot.cell(row=i, column=total_col_index, value=total_val)
        if c_fmt:
            total_cell.number_format = c_fmt
        else:
            format_number_cell(total_cell)

    set_column_widths(ws_pivot, {"A": 18, "B": 16})

def strip_sheet_strings(ws: "Worksheet"):
    """
    去除主表数据区（从第2行开始）中所有字符串单元格两端的空格。
    跳过合并单元格以避免赋值异常。
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(cell, MergedCell):
                continue
            if isinstance(v, str):
                cell.value = v.strip()

def remove_rows_with_item_code_prefix(ws: "Worksheet", item_code_col: int, prefix: str = "XX"):
    """
    删除“货品编码”列以指定前缀开头的行（不区分大小写，剔除两端空格后判断）。
    先收集行号，再自底向上删除以保持行号稳定。
    """
    rows_to_delete = []
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=item_code_col)
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if s.upper().startswith(prefix.upper()):
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

def build_pivot_sheet(wb, src_ws: "Worksheet"):
    """
    在新建 sheet '透视表' 中，从 B 列开始生成数据透视表：
    - 行：规格代码
    - 列：货主
    - 值：在库数量（求和）
    - 在最右侧添加“总计”列，并按该列降序排序（可见的排序依据）
    A1 标题为“商品名称”（暂留空数据，后续填充）
    """
    # 定位所需字段列
    col_spec = find_header_column(src_ws, "规格代码")
    col_owner = find_header_column(src_ws, "货主")
    col_qty = find_header_column(src_ws, "在库数量")
    if not all([col_spec, col_owner, col_qty]):
        print("生成透视表失败：未找到'规格代码'、'货主'或'在库数量'列。")
        return

    max_row = src_ws.max_row or 1
    records = []
    for r in range(2, max_row + 1):
        spec = src_ws.cell(row=r, column=col_spec).value
        owner = src_ws.cell(row=r, column=col_owner).value
        qty = src_ws.cell(row=r, column=col_qty).value
        if spec is None or owner is None:
            continue
        # 数量转为数值
        if isinstance(qty, (int, float)):
            v = float(qty)
        elif isinstance(qty, str) and is_number_like(qty):
            v = float(to_number(qty))  # to_number 可能返回 int/float
        else:
            v = 0.0
        records.append({"规格代码": str(spec).strip(), "货主": str(owner).strip(), "在库数量": v})

    if not records:
        print("生成透视表失败：无有效数据。")
        return

    df = pd.DataFrame.from_records(records)
    pv = pd.pivot_table(df, index="规格代码", columns="货主", values="在库数量", aggfunc="sum", fill_value=0)
    # 计算总计并排序
    pv["__行总计__"] = pv.sum(axis=1)
    pv = pv.sort_values(by="__行总计__", ascending=False)
    totals_series = pv["__行总计__"].copy()
    pv_out = pv.drop(columns="__行总计__")

    # 创建新 sheet 并写入，从 B1 开始
    ws_pivot = wb.create_sheet("透视表")
    ws_pivot.cell(row=1, column=1, value="商品名称")  # A1
    ws_pivot.cell(row=1, column=2, value="规格代码")  # B1

    # 写列标题（货主列表）从 C1 开始，最后一列为“总计”
    col_names = [str(c) for c in pv_out.columns.tolist()]
    for j, name in enumerate(col_names, start=0):
        ws_pivot.cell(row=1, column=3 + j, value=name)
    total_col_index = 3 + len(col_names)
    ws_pivot.cell(row=1, column=total_col_index, value="总计")

    # 写数据行：B 列为规格代码；C 起为各货主汇总；最右为总计
    for i, (spec_code, row_series) in enumerate(pv_out.iterrows(), start=2):
        # 规格代码数字转数值，文本保留
        if isinstance(spec_code, str) and is_number_like(spec_code):
            spec_out = to_number(spec_code)
        else:
            spec_out = spec_code
        spec_cell = ws_pivot.cell(row=i, column=2, value=spec_out)  # B列规格代码
        # 给规格代码设置合适的格式（如果是数字）
        format_number_cell(spec_cell)

        # 写各货主汇总
        for j, name in enumerate(col_names, start=0):
            val = float(row_series.get(name, 0.0))
            cell = ws_pivot.cell(row=i, column=3 + j, value=val)
            format_number_cell(cell)

        # 写总计
        total_val = float(totals_series.get(spec_code, 0.0))
        total_cell = ws_pivot.cell(row=i, column=total_col_index, value=total_val)
        format_number_cell(total_cell)

    # 设置列宽：A(商品名称)、B(规格代码)与“总计”列
    set_column_widths(ws_pivot, {"A": 18, "B": 16})


def process_inventory_file(inv_path: Path, base_dir: Path) -> Tuple[bool, str]:
    """
    执行处理逻辑（在“修改后”副本上进行，不触碰原文件）：
    - 将 C 列“看起来是数字”的文本转为数值
    - 右侧插入“规格代码”列并填充（先数值匹配，再文本匹配；对照出的码若看起来是数字则转为数值）
    - 隐藏 I-M 列
    - 着色“规格代码”表头
    - 保存到新文件：原文件名-修改后.xlsx
    """
    if not inv_path.exists():
        return False, f"文件不存在：{inv_path}"

    wb = load_workbook(inv_path, data_only=True)
    ws = wb.active
    if not isinstance(ws, Worksheet):
        # 若类型异常，则尝试使用第一个工作表；若不存在则返回失败
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            return False, "工作簿中没有可用的工作表"

    # 1) 处理 C 列数字文本转换
    C_COL = 3
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=C_COL)
        val = cell.value
        if isinstance(val, str) and is_number_like(val) and not isinstance(cell, MergedCell):
            cell.value = to_number(val)
        # 设置显示格式，避免出现科学计数法
        format_number_cell(cell)

    # 2) 插入“规格代码”列，并基于“货品编码”做映射填充（两次对照）
    new_col_idx = insert_column_right_of(ws, C_COL, "规格代码")
    # 设置列宽：C/D/E/O -> 16
    set_column_widths(ws, {"C": 16, "D": 16, "E": 16, "O": 16})
    # D1 先刷 C1 的格式（排除填充），再上色
    try:
        copy_cell_style(ws["C1"], ws.cell(row=1, column=new_col_idx), exclude=("fill",))
    except Exception:
        pass
    apply_header_fill(ws, new_col_idx, "FFFF00")

    # 加载双映射
    map_num, map_text = load_mappings(base_dir)
    if not map_num and not map_text:
        print("规格代码映射为空或未找到，将创建空列。")

    # 查找主表中的“货品编码”列索引
    item_code_col = find_header_column(ws, "货品编码")
    if item_code_col is None:
        print("未找到主表的“货品编码”列，无法进行映射填充。")
    else:
        # 预处理：去除字符串空格并删除以“XX”开头的货品编码行
        strip_sheet_strings(ws)
        remove_rows_with_item_code_prefix(ws, item_code_col, prefix="XX")
        # 删除行后刷新最大行数
        max_row = ws.max_row or 1

        # 第一次：数值匹配
        for r in range(2, max_row + 1):
            key_cell = ws.cell(row=r, column=item_code_col)
            val = key_cell.value
            target_cell = ws.cell(row=r, column=new_col_idx)
            if isinstance(target_cell, MergedCell) or target_cell.value not in (None, ""):
                continue
            num_key: Any = None
            if isinstance(val, (int, float)):
                num_key = val
            elif isinstance(val, str) and is_number_like(val):
                num_key = to_number(val)
            if num_key is not None and num_key in map_num:
                spec = map_num[num_key]
                # 对照出的码若为数字文本则转为数值
                if isinstance(spec, str) and is_number_like(spec):
                    spec_out = to_number(spec)
                else:
                    spec_out = spec
                target_cell.value = spec_out
                # 设置显示格式，避免出现科学计数法
                format_number_cell(target_cell)

        # 第二次：文本匹配（填充剩余空白行）
        for r in range(2, max_row + 1):
            target_cell = ws.cell(row=r, column=new_col_idx)
            if isinstance(target_cell, MergedCell) or target_cell.value not in (None, ""):
                continue
            key_cell = ws.cell(row=r, column=item_code_col)
            val = key_cell.value
            key_text = str(val).strip() if val is not None else ""
            if key_text:
                spec = map_text.get(key_text)
                if spec is not None:
                    if isinstance(spec, str) and is_number_like(spec):
                        spec_out = to_number(spec)
                    else:
                        spec_out = spec
                    target_cell.value = spec_out
                    # 设置显示格式，避免出现科学计数法
                    format_number_cell(target_cell)

    # 3) 隐藏 I-M 列
    hide_columns_range(ws, "I", "M")

    # 4) 在新 sheet 生成透视表（增强版）
    try:
        build_pivot_sheet_enhanced(wb, ws, base_dir=base_dir)
    except Exception as e:
        print(f"生成透视表时出现问题：{e}")

    # 5) 保存到“修改后”副本
    modified_path = inv_path.with_name(f"{inv_path.stem}-修改后.xlsx")
    wb.save(modified_path)
    return True, f"已处理并另存为：{modified_path.name}"


def main():
    # 抑制 openpyxl 不影响结果的样式警告
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

    base_dir = get_base_dir()
    print(f"工作目录：{base_dir}")

    inv_path = find_latest_inventory_file(base_dir)
    if inv_path is None:
        print("未找到符合条件的文件：库存快照明细*.xlsx")
        return

    ok, msg = process_inventory_file(inv_path, base_dir)
    print(msg)

    # 同步生成/更新需求文档
    write_requirements_doc(base_dir)


def write_requirements_doc(base_dir: Path):
    """
    自动生成/更新需求文档，记录当前处理规则与时间戳
    """
    doc_path = base_dir / "需求文档.md"
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    content = f"""# 库存快照处理需求文档

更新时间：{ts}

处理流程：
- 在程序所在目录（或Jupyter当前目录）搜索最新的“库存快照明细*.xlsx”文件。
- 复制原文件并在副本上处理（不修改原文件），副本命名为“原文件名-修改后.xlsx”：
  1) 将 C 列中“看起来是数字”的文本转为数值（支持去空格和千分位逗号；无法解析保留原文本）。
  2) 在 C 列右侧插入新列“规格代码”，读取同目录下“盘点-规格代码匹配.xlsx”的 sheet1 的 A/B 列：
     - A 列为“货品编码”，B 列为“规格代码”；
     - 分两次对照：先按“数值键”（int/float 或形如数字的文本）匹配，后按“文本键”匹配；
     - 对照得到的“规格代码”若看起来是数字则转换为数值，文本保持不变；
     - 若未找到映射或主表中无“货品编码”列，则保留空列。
  3) 隐藏 I 至 M 列。
  4) 插入新列后设置 C/D/E/O 列宽为 16；将 D1 先复制 C1 的格式（不含填充），再将标题单元格填充为 FFFF00。
  5) 对数值单元格设置合理的显示格式（整数为 "0"，小数为 "0.####################"），避免科学计数法显示。
  6) 在新建工作表“透视表”中，从 B 列开始生成数据透视表：
     - 行：规格代码；列：货主；值：在库数量（求和）
     - 以在库数量总计降序排序
     - A1 标题为“商品名称”，内容后续再填充
- 将处理结果保存到“原文件名-修改后.xlsx”。

其他：
- 运行时抑制不影响结果的 openpyxl 样式警告，避免干扰判断。

注意事项：
- 若同目录内存在多个“库存快照明细*.xlsx”，选择最新修改的一个进行处理。
- 若“盘点-规格代码匹配.xlsx”未找到或 sheet1 不存在，则跳过映射。
- 映射时将货品编码统一按字符串对齐，避免数字文本差异导致匹配失败。
"""
    try:
        doc_path.write_text(content, encoding="utf-8")
        print(f"需求文档已更新：{doc_path}")
    except Exception as e:
        print(f"需求文档写入失败：{e}")


if __name__ == "__main__":
    main()