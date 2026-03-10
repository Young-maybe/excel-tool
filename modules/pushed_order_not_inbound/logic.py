"""
已推单未入库表处理

来源：已推单未入库表处理/已推单未入库处理.ipynb（Cell 0）
规则：
- 只做 I/O 适配：输入由 UI 选取文件夹；输出写入统一输出目录下的子目录
- 不使用动态加载外部源脚本
- 核心处理逻辑保持一致
"""

from __future__ import annotations

import glob
import os
import re
import shutil
from pathlib import Path
from typing import Callable, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class 退货商品处理器:
    def __init__(self, base_dir: Path):
        self.预定义表头 = [
            "店铺名称",
            "分销商名称",
            "单据编号",
            "订单编号",
            "平台单号",
            "退货原因",
            "物流公司",
            "商品代码",
            "商品名称",
            "规格代码",
            "退货数量",
            "入库数量",
            "仓库名称",
            "制单时间",
            "审核时间",
            "入库时间",
            "退回物流单号",
            "售后类型",
        ]
        self.当前目录 = str(base_dir)
        self.progress_cb: Optional[Callable[[int, str], None]] = None

    def _report(self, p: int, s: str) -> None:
        if self.progress_cb:
            self.progress_cb(p, s)

    def 读取表头配置(self):
        self._report(26, f"读取内化表头配置（{len(self.预定义表头)}列）…")
        print(f"[OK] 使用内化的表头配置，共{len(self.预定义表头)}个字段")

    def 搜索CSV文件(self) -> Optional[str]:
        self._report(30, "搜索 退货商品明细汇总*.csv …")
        搜索模式 = os.path.join(self.当前目录, "退货商品明细汇总*.csv")
        找到的文件 = glob.glob(搜索模式)

        if not 找到的文件:
            print("[ERROR] 未找到匹配的CSV文件")
            print(f"搜索模式：{搜索模式}")
            return None
        return 找到的文件[0]

    def 读取自营配置(self) -> Optional[pd.DataFrame]:
        try:
            self._report(34, "读取 店铺匹配仓库配置.xlsx（自营仓）…")
            配置文件路径 = os.path.join(self.当前目录, "店铺匹配仓库配置.xlsx")
            if os.path.exists(配置文件路径):
                df = pd.read_excel(配置文件路径, sheet_name="自营仓")
                print(f"[OK] 成功读取自营配置，共{len(df)}条记录")
                return df
            print("[WARN] 未找到店铺匹配仓库配置.xlsx文件")
            return None
        except Exception as e:
            print(f"[ERROR] 读取自营配置时出错：{str(e)}")
            return None

    def 读取CSV文件(self, 文件路径: str) -> Optional[pd.DataFrame]:
        try:
            self._report(40, "读取 CSV（自动尝试编码）…")
            编码列表 = ["utf-8", "gbk", "gb2312", "utf-8-sig"]
            for 编码 in 编码列表:
                try:
                    df = pd.read_csv(文件路径, encoding=编码)
                    print(f"[OK] 成功读取CSV文件，使用编码：{编码}")
                    return df
                except UnicodeDecodeError:
                    continue
            raise Exception("无法使用常见编码格式读取文件")
        except Exception as e:
            print(f"[ERROR] 读取CSV文件时出错：{str(e)}")
            return None

    def 筛选列(self, df: Optional[pd.DataFrame]):
        if df is None:
            return None

        self._report(52, "匹配列名并筛选需要的列…")
        原始列名 = list(df.columns)
        匹配的列 = []
        未匹配的列 = []

        for 预定义列 in self.预定义表头:
            找到匹配 = False
            for 原始列 in 原始列名:
                if 预定义列 in 原始列 or 原始列 in 预定义列:
                    匹配的列.append(原始列)
                    找到匹配 = True
                    break
            if not 找到匹配:
                print(f"[WARN] 未找到匹配的列：{预定义列}")

        for 原始列 in 原始列名:
            if 原始列 not in 匹配的列:
                未匹配的列.append(原始列)

        return df, 匹配的列, 未匹配的列

    def 添加新列(self, df: pd.DataFrame) -> pd.DataFrame:
        self._report(62, "新增计算列（未入库数/自营否/单据编号2）…")
        df_副本 = df.copy()
        列名列表 = list(df_副本.columns)

        if "入库数量" in 列名列表:
            入库数量索引 = 列名列表.index("入库数量")
            df_副本.insert(入库数量索引 + 1, "未入库数", "")

        if "仓库名称" in 列名列表:
            列名列表 = list(df_副本.columns)
            仓库名称索引 = 列名列表.index("仓库名称")
            df_副本.insert(仓库名称索引 + 1, "自营否", "")

        if "售后类型" in 列名列表 and "单据编号" in 列名列表:
            列名列表 = list(df_副本.columns)
            售后类型索引 = 列名列表.index("售后类型")
            单据编号数据 = df_副本["单据编号"].copy()
            df_副本.insert(售后类型索引 + 1, "单据编号2", 单据编号数据)

        if {"未入库数", "退货数量", "入库数量"}.issubset(df_副本.columns):
            退货数量_数值 = pd.to_numeric(df_副本["退货数量"], errors="coerce").fillna(0)
            入库数量_数值 = pd.to_numeric(df_副本["入库数量"], errors="coerce").fillna(0)
            df_副本["未入库数"] = 退货数量_数值 - 入库数量_数值

        if {"自营否", "仓库名称"}.issubset(df_副本.columns):
            自营配置 = self.读取自营配置()
            if 自营配置 is not None:
                仓库映射 = dict(zip(自营配置["管易仓库名称"], 自营配置["仓库盘点货主"]))
                df_副本["自营否"] = df_副本["仓库名称"].map(仓库映射)
                原始行数 = len(df_副本)
                df_副本 = df_副本.dropna(subset=["自营否"])
                删除行数 = 原始行数 - len(df_副本)
                print(f"[OK] 删除了{删除行数}行匹配不到的数据，剩余{len(df_副本)}行")
            else:
                df_副本["自营否"] = "配置文件读取失败"

        return df_副本

    def 获取可见列列表(self, df: pd.DataFrame):
        匹配的列 = self.预定义表头
        新增列名 = ["未入库数", "自营否", "单据编号2"]
        可见列 = []
        for 列名 in df.columns:
            if 列名 in 匹配的列 or 列名 in 新增列名:
                可见列.append(列名)
        return 可见列

    def 应用隐藏列设置(self, ws, df: pd.DataFrame):
        try:
            所有列 = list(df.columns)
            匹配的列 = self.预定义表头
            新增列名 = ["未入库数", "自营否", "单据编号2"]

            需要隐藏的列 = []
            for 列名 in 所有列:
                if 列名 not in 匹配的列 and 列名 not in 新增列名:
                    需要隐藏的列.append(列名)

            for 列名 in 需要隐藏的列:
                if 列名 in 所有列:
                    col_idx = 所有列.index(列名) + 1
                    from openpyxl.utils import get_column_letter

                    列字母 = get_column_letter(col_idx)
                    ws.column_dimensions[列字母].hidden = True
        except Exception as e:
            print(f"[ERROR] 应用隐藏列设置时出错：{str(e)}")

    def 添加未审核工作表(self, wb, df: pd.DataFrame, 原文件路径: str):
        try:
            原文件名 = os.path.splitext(os.path.basename(原文件路径))[0]
            if "退货商品明细汇总" in 原文件名:
                文件标识 = 原文件名.replace("退货商品明细汇总", "")
            else:
                文件标识 = "未知"
            新工作表名 = f"已制单未审核-退货单，至{文件标识}（客服）"

            筛选条件 = (df["审核时间"].isna() | df["审核时间"].eq("")) | (
                df["退回物流单号"].isna() | df["退回物流单号"].eq("")
            )
            未审核数据 = df[筛选条件].copy()
            if len(未审核数据) <= 0:
                return

            ws_新 = wb.create_sheet(title=新工作表名)

            表头字体 = Font(bold=True)
            居中对齐 = Alignment(horizontal="center", vertical="center")
            黄色填充 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            无边框 = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style=None),
                bottom=Side(style=None),
            )

            新增列名 = ["未入库数", "自营否", "单据编号2"]
            for col_idx, 列名 in enumerate(未审核数据.columns, 1):
                cell = ws_新.cell(row=1, column=col_idx, value=列名)
                cell.font = 表头字体
                cell.alignment = 居中对齐
                cell.border = 无边框
                if 列名 in 新增列名:
                    cell.fill = 黄色填充

            for row_idx, (_, row_data) in enumerate(未审核数据.iterrows(), 2):
                for col_idx, 值 in enumerate(row_data, 1):
                    cell = ws_新.cell(row=row_idx, column=col_idx, value=值)
                    cell.border = 无边框

            self.应用隐藏列设置(ws_新, df)

            可见列 = self.获取可见列列表(df)
            for 列名 in 可见列:
                if 列名 in 未审核数据.columns:
                    col_idx = list(未审核数据.columns).index(列名) + 1
                    from openpyxl.utils import get_column_letter

                    列字母 = get_column_letter(col_idx)
                    最大宽度 = max(len(str(列名)), 10)
                    for 值 in 未审核数据[列名]:
                        if pd.notna(值):
                            最大宽度 = max(最大宽度, len(str(值)))
                    ws_新.column_dimensions[列字母].width = min(最大宽度 + 2, 50)
        except Exception as e:
            print(f"[ERROR] 创建未审核工作表时出错：{str(e)}")

    def 添加仓库回复工作表(self, wb, df: pd.DataFrame):
        try:
            新工作表名 = "仓库回复登记源表"
            筛选条件 = (
                (pd.to_numeric(df["未入库数"], errors="coerce") != 0)
                & (pd.to_numeric(df["未入库数"], errors="coerce").notna())
                & (df["审核时间"].notna() & df["审核时间"].ne(""))
                & (df["退回物流单号"].notna() & df["退回物流单号"].ne(""))
            )
            仓库回复数据 = df[筛选条件].copy()
            if len(仓库回复数据) <= 0:
                return

            可见列 = self.获取可见列列表(df)
            仓库回复数据_筛选 = 仓库回复数据[可见列].copy()
            仓库回复数据_筛选["未入库原因（仓库回复）"] = ""

            ws_新 = wb.create_sheet(title=新工作表名)

            表头字体 = Font(bold=True)
            居中对齐 = Alignment(horizontal="center", vertical="center")
            黄色填充 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            绿色填充 = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            无边框 = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style=None),
                bottom=Side(style=None),
            )

            新增列名 = ["未入库数", "自营否", "单据编号2"]
            for col_idx, 列名 in enumerate(仓库回复数据_筛选.columns, 1):
                cell = ws_新.cell(row=1, column=col_idx, value=列名)
                cell.font = 表头字体
                cell.alignment = 居中对齐
                cell.border = 无边框
                if 列名 in 新增列名:
                    cell.fill = 黄色填充
                elif 列名 == "未入库原因（仓库回复）":
                    cell.fill = 绿色填充

            for row_idx, (_, row_data) in enumerate(仓库回复数据_筛选.iterrows(), 2):
                for col_idx, 值 in enumerate(row_data, 1):
                    cell = ws_新.cell(row=row_idx, column=col_idx, value=值)
                    cell.border = 无边框

            for col_idx, 列名 in enumerate(仓库回复数据_筛选.columns, 1):
                from openpyxl.utils import get_column_letter

                列字母 = get_column_letter(col_idx)
                最大宽度 = max(len(str(列名)), 10)
                for 值 in 仓库回复数据_筛选[列名]:
                    if pd.notna(值):
                        最大宽度 = max(最大宽度, len(str(值)))
                ws_新.column_dimensions[列字母].width = min(最大宽度 + 2, 50)
        except Exception as e:
            print(f"[ERROR] 创建仓库回复工作表时出错：{str(e)}")

    def 转换为Excel(self, df: pd.DataFrame, 匹配的列, 未匹配的列, 原文件路径: str) -> Optional[str]:
        try:
            原文件名 = os.path.splitext(os.path.basename(原文件路径))[0]
            日期匹配 = re.search(r"退货商品明细汇总(.+)", 原文件名)
            if 日期匹配:
                日期标识 = 日期匹配.group(1)
                输出文件名 = f"退货未入库-截至{日期标识}.xlsx"
            else:
                输出文件名 = f"退货未入库-截至{原文件名}.xlsx"
            输出路径 = os.path.join(self.当前目录, 输出文件名)

            df_处理后 = self.添加新列(df)

            self._report(78, "写入主表（退货明细）…")
            with pd.ExcelWriter(输出路径, engine="openpyxl") as writer:
                df_处理后.to_excel(writer, sheet_name="退货明细", index=False)

            wb = load_workbook(输出路径)
            ws = wb["退货明细"]

            表头字体 = Font(bold=True)
            居中对齐 = Alignment(horizontal="center", vertical="center")
            黄色填充 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            无边框 = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style=None),
                bottom=Side(style=None),
            )

            新增列名 = ["未入库数", "自营否", "单据编号2"]
            for cell in ws[1]:
                cell.font = 表头字体
                cell.alignment = 居中对齐
                cell.border = 无边框
                if cell.value in 新增列名:
                    cell.fill = 黄色填充

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = 无边框

            self._report(86, "隐藏无关列 & 调整列宽…")
            列索引映射 = {列名: idx + 1 for idx, 列名 in enumerate(df_处理后.columns)}
            for 列名 in 未匹配的列:
                if 列名 in 列索引映射:
                    列索引 = 列索引映射[列名]
                    from openpyxl.utils import get_column_letter

                    列字母 = get_column_letter(列索引)
                    ws.column_dimensions[列字母].hidden = True

            可见列 = 匹配的列 + 新增列名
            for 列名 in 可见列:
                if 列名 in 列索引映射:
                    列索引 = 列索引映射[列名]
                    from openpyxl.utils import get_column_letter

                    列字母 = get_column_letter(列索引)
                    最大宽度 = max(len(str(列名)), 10)
                    for row in ws.iter_rows(min_col=列索引, max_col=列索引, min_row=2):
                        for cell in row:
                            if cell.value:
                                最大宽度 = max(最大宽度, len(str(cell.value)))
                    ws.column_dimensions[列字母].width = min(最大宽度 + 2, 50)

            self._report(90, "生成附加工作表（未审核/仓库回复）…")
            self.添加未审核工作表(wb, df_处理后, 原文件路径)
            self.添加仓库回复工作表(wb, df_处理后)

            self._report(95, "保存最终 Excel…")
            wb.save(输出路径)
            return 输出路径
        except Exception as e:
            print(f"[ERROR] 转换为Excel时出错：{str(e)}")
            return None

    def 处理文件(self) -> bool:
        self.读取表头配置()
        csv文件路径 = self.搜索CSV文件()
        if not csv文件路径:
            return False
        df = self.读取CSV文件(csv文件路径)
        if df is None:
            return False
        结果 = self.筛选列(df)
        if 结果 is None:
            return False
        df, 匹配的列, 未匹配的列 = 结果
        输出路径 = self.转换为Excel(df, 匹配的列, 未匹配的列, csv文件路径)
        return bool(输出路径)


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> List[str]:
    def report(p: int, s: str) -> None:
        if progress_cb:
            progress_cb(p, s)

    in_dir = Path(input_dir)
    out_dir = Path(output_dir)
    if not in_dir.exists():
        raise FileNotFoundError(f"输入目录不存在：{input_dir}")

    work_dir = out_dir / "已推单未入库表处理"
    work_dir.mkdir(parents=True, exist_ok=True)

    # 复制输入文件到工作目录，确保输出统一落在 work_dir（避免锁文件/覆盖）
    report(5, "检查输入文件…")
    csv_files = sorted(in_dir.glob("退货商品明细汇总*.csv"))
    if not csv_files:
        raise FileNotFoundError("未找到输入文件：退货商品明细汇总*.csv")

    cfg = in_dir / "店铺匹配仓库配置.xlsx"
    if not cfg.exists():
        raise FileNotFoundError("缺少必需文件：店铺匹配仓库配置.xlsx")

    report(15, "复制输入文件到工作目录…")
    for p in csv_files:
        shutil.copy2(p, work_dir / p.name)
    shutil.copy2(cfg, work_dir / cfg.name)

    processor = 退货商品处理器(work_dir)
    processor.progress_cb = progress_cb
    report(25, "读取 CSV 并匹配表头…")
    ok = processor.处理文件()
    if not ok:
        raise RuntimeError("处理失败：请检查输入文件格式/列名是否符合要求。")

    report(90, "收集输出文件…")
    outputs = sorted(work_dir.glob("退货未入库-截至*.xlsx"))
    if not outputs:
        # 兜底：输出命名可能退化为“退货未入库-截至{原文件名}.xlsx”
        outputs = sorted(work_dir.glob("退货未入库-截至*.xlsx"))
    if not outputs:
        raise FileNotFoundError("未找到输出文件：退货未入库-截至*.xlsx")
    report(100, "完成")
    return [str(outputs[-1])]


