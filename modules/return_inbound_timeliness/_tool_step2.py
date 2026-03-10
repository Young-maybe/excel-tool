import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import threading
from datetime import datetime
import re

try:
    from PIL import Image, ImageTk
    pil_available = True
except ImportError:
    pil_available = False
    print("PIL库未安装，将使用备用方案")

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    pandas_available = True
except ImportError:
    pandas_available = False
    print("pandas/openpyxl库未安装，数据处理功能将受限")

class ReturnDataProcessor:
    def __init__(self):
        self.input_file = None
        self.output_file = None
        self.month = None
        self.df = None
        
    def find_input_file(self, folder_path):
        """查找退货入库时效分析文件"""
        for file in os.listdir(folder_path):
            if file.startswith("退货入库时效分析") and file.endswith(".xlsx"):
                # 提取月份
                match = re.search(r'(\d+)月', file)
                if match:
                    self.month = match.group(1)
                    self.input_file = os.path.join(folder_path, file)
                    return True
        return False
    
    def process(self, input_folder, output_folder):
        """完整的处理流程"""
        try:
            if not pandas_available:
                return False, "缺少pandas或openpyxl库，请先安装: pip install pandas openpyxl"
            
            print("[INFO] 开始数据处理...")
            
            # 1. 查找输入文件
            if not self.find_input_file(input_folder):
                return False, "未找到退货入库时效分析文件"
            
            print(f"[INFO] 找到输入文件: {os.path.basename(self.input_file)}")
            print(f"[INFO] 提取月份: {self.month}月")
            
            # 2. 加载数据
            try:
                self.df = pd.read_excel(self.input_file)
                print(f"[OK] 成功加载数据，共 {len(self.df)} 行")
            except Exception as e:
                return False, f"数据加载失败: {e}"
            
            # 3. 检查必要的列
            required_cols = ['自营仓主体', '48h入库时效是否满足', '退回物流单号']
            missing_cols = [col for col in required_cols if col not in self.df.columns]
            if missing_cols:
                return False, f"缺少必要列: {missing_cols}"
            
            # 4. 清理数据
            self.df['自营仓主体'] = self.df['自营仓主体'].replace({
                '信选次品': '信选',
                '北分': '清元',
                '清元次品': '清元'
            })
            print("[OK] 自营仓主体数据清理完成")
            
            # 5. 提取需要的三列
            analysis_data = self.df[required_cols].copy()
            
            # 6. 去除重复的物流单号
            original_count = len(analysis_data)
            analysis_data = analysis_data.drop_duplicates(subset=['退回物流单号'])
            dedup_count = len(analysis_data)
            print(f"[OK] 去重完成: {original_count} -> {dedup_count} 行")
            
            # 7. 生成输出文件路径
            output_filename = f"退货商品明细汇总-{self.month}月推单.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            # 8. 创建Excel文件
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "时效分析"
            
            # 写入数据
            for r in dataframe_to_rows(analysis_data, index=False, header=True):
                worksheet.append(r)
            
            # 9. 创建汇总表
            summary = analysis_data.groupby(['自营仓主体', '48h入库时效是否满足']).size().unstack(fill_value=0)
            
            # 写入汇总表到I3位置
            start_row = 3
            start_col = 9  # I列
            
            # 写入标题
            worksheet.cell(row=start_row, column=start_col, value="自营仓主体")
            for i, col in enumerate(summary.columns):
                worksheet.cell(row=start_row, column=start_col + i + 1, value=col)
            
            # 写入数据
            for i, (index, row) in enumerate(summary.iterrows()):
                worksheet.cell(row=start_row + i + 1, column=start_col, value=index)
                for j, value in enumerate(row):
                    worksheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)
            
            # 10. 计算满足率
            total_count = len(analysis_data)
            satisfied_count = len(analysis_data[analysis_data['48h入库时效是否满足'] == '满足'])
            satisfaction_rate = (satisfied_count / total_count * 100) if total_count > 0 else 0
            
            # 在I13单元格填入结果
            result_text = f"{self.month}月整体退货满足率{satisfaction_rate:.2f}%，达标96%"
            worksheet.cell(row=13, column=9, value=result_text)
            
            print(f"[OK] 满足率计算完成: {satisfaction_rate:.2f}%")
            
            # 11. 保存文件
            workbook.save(output_path)
            print(f"[OK] 分析文件创建完成: {output_path}")
            
            print(f"[OK] 处理完成，输出文件: {output_filename}")
            return True, f"处理成功！生成文件: {output_filename}\n满足率: {satisfaction_rate:.2f}%"
            
        except Exception as e:
            error_msg = f"处理过程中出现错误: {str(e)}"
            print(f"[ERROR] {error_msg}")
            return False, error_msg

class GenshinSimpleApp:
    def __init__(self):
        # 初始化属性
        self.selected_folder = os.getcwd()
        self.output_folder = os.getcwd()
        self.files_data = []
        self.bg_image = None
        self.original_image = None
        self.path_var = None
        self.output_path_var = None
        self.file_tree = None
        self.process_btn = None
        self.status_var = None
        self.processor = ReturnDataProcessor()
        
        # 创建主窗口
        self.root = tk.Tk()
        self.root.title("退货数据处理工具")
        
        # 设置默认窗口大小 - 保留75%的尺寸（减少25%）
        self.default_width = 900  # 1200 * 0.75
        self.default_height = 600  # 800 * 0.75
        self.window_width = self.default_width
        self.window_height = self.default_height
        self.root.geometry(f"{self.window_width}x{self.window_height}")
        
        # 加载原图并设置窗口大小
        self.load_original_image()
        
        # 居中显示
        self.center_window()
        
        # 设置样式
        self.setup_styles()
        
        # 创建主界面
        self.create_main_interface()
        
    def load_original_image(self):
        """加载用户提供的原图"""
        # 尝试加载原图 - 优先使用用户提供的神里凌华图片
        image_files = ['神里凌华_01.jpg', 'genshin_original.jpg', 'genshin_original.png', 
                      'genshin_bg.jpg', 'genshin_bg.png']
        
        if pil_available:
            for img_file in image_files:
                if os.path.exists(img_file):
                    try:
                        # 加载原图
                        self.original_image = Image.open(img_file)
                        
                        # 获取图片尺寸
                        img_width, img_height = self.original_image.size
                        
                        # 计算合适的显示尺寸（保持比例，但不超过屏幕）
                        screen_width = self.root.winfo_screenwidth()
                        screen_height = self.root.winfo_screenheight()
                        
                        # 设置合适的显示尺寸 - 保留75%大小
                        max_width = int(screen_width * 0.75)  # 屏幕宽度的75%
                        max_height = int(screen_height * 0.75)  # 屏幕高度的75%
                        
                        # 计算缩放比例，保持图片比例
                        scale_w = max_width / img_width
                        scale_h = max_height / img_height
                        scale = min(scale_w, scale_h, 1.0)  # 不放大，只缩小
                        
                        # 计算图片缩放后的尺寸，保持原始比例
                        scaled_img_width = int(img_width * scale)
                        scaled_img_height = int(img_height * scale)
                        
                        # 一级窗口保持图片比例，不变形
                        self.window_width = scaled_img_width
                        self.window_height = scaled_img_height
                        
                        # 调整图片大小，保持原始比例
                        resized_image = self.original_image.resize(
                            (self.window_width, self.window_height), 
                            Image.Resampling.LANCZOS
                        )
                        
                        # 转换为tkinter可用的格式
                        self.bg_image = ImageTk.PhotoImage(resized_image)
                        
                        # 设置窗口大小
                        self.root.geometry(f"{self.window_width}x{self.window_height}")
                        
                        print(f"成功加载原图: {img_file}")
                        print(f"原图尺寸: {img_width}x{img_height}")
                        print(f"窗口尺寸: {self.window_width}x{self.window_height}")
                        return
                        
                    except Exception as e:
                        print(f"加载图片 {img_file} 失败: {e}")
        
        print("未找到原图，使用默认窗口尺寸")
        
    def center_window(self):
        """窗口居中"""
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (self.window_width // 2)
        y = (self.root.winfo_screenheight() // 2) - (self.window_height // 2)
        self.root.geometry(f"{self.window_width}x{self.window_height}+{x}+{y}")
        
    def setup_styles(self):
        """设置ttk样式"""
        style = ttk.Style()
        
        # 主按钮样式
        style.configure('MainGenshin.TButton',
                       background='#4169E1',
                       foreground='white',
                       font=('Microsoft YaHei UI', 12, 'bold'),
                       padding=(20, 10),
                       relief='raised')
        
        style.map('MainGenshin.TButton',
                 background=[('active', '#FFD700'),
                           ('pressed', '#FF69B4')])
        
        # 处理按钮样式
        style.configure('Process.TButton',
                       background='#FF69B4',
                       foreground='white',
                       font=('Microsoft YaHei UI', 12, 'bold'),
                       padding=(20, 10),
                       relief='raised')
        
        # 返回按钮样式
        style.configure('Back.TButton',
                       background='#87CEEB',
                       foreground='white',
                       font=('Microsoft YaHei UI', 10, 'bold'),
                       padding=(15, 8),
                       relief='raised')
        
        # 表格样式
        style.configure('Genshin.Treeview',
                       background='#FFFAF0',
                       foreground='#2F4F4F',
                       font=('Microsoft YaHei UI', 9),
                       rowheight=25)
        
        style.configure('Genshin.Treeview.Heading',
                       background='#4169E1',
                       foreground='white',
                       font=('Microsoft YaHei UI', 10, 'bold'))
        
    def create_main_interface(self):
        """创建主界面"""
        # 如果有背景图片，直接显示
        if self.bg_image:
            # 创建背景标签，直接平铺原图
            bg_label = tk.Label(self.root, image=self.bg_image)
            bg_label.place(x=0, y=0, width=self.window_width, height=self.window_height)
        else:
            # 备用方案：使用纯色背景
            self.root.configure(bg='#E8F4FD')
        
        # 在原图上添加UI元素
        self.add_main_ui_elements()
        
    def add_main_ui_elements(self):
        """在原图背景上添加UI元素"""
        # 标题区域 - 半透明背景
        title_frame = tk.Frame(self.root, bg='#FFE4E1', relief='raised', bd=3)
        title_frame.place(relx=0.1, rely=0.05, relwidth=0.8, relheight=0.12)
        
        # 主标题
        title_size = max(16, int(self.window_width / 40))
        title_label = tk.Label(title_frame,
                              text="🌸 退货数据处理工具 🌸",
                              font=('Microsoft YaHei UI', title_size, 'bold'),
                              fg='#4169E1',
                              bg='#FFE4E1')
        title_label.pack(expand=True)
        
        # 副标题
        subtitle_size = max(10, int(self.window_width / 80))
        subtitle_label = tk.Label(title_frame,
                                 text="✨ 原神风格 · 轻量化数据处理系统 ✨",
                                 font=('Microsoft YaHei UI', subtitle_size),
                                 fg='#FF69B4',
                                 bg='#FFE4E1')
        subtitle_label.pack()
        
        # 中央按钮区域
        button_frame = tk.Frame(self.root, bg='', relief='raised', bd=0)
        button_frame.place(relx=0.3, rely=0.45, relwidth=0.4, relheight=0.15)
        
        # 主要操作按钮
        main_button = ttk.Button(button_frame,
                                text="📂 选择文件夹开始处理",
                                style='MainGenshin.TButton',
                                command=self.open_file_selection)
        main_button.pack(pady=15)
        
        # 底部状态栏
        status_frame = tk.Frame(self.root, bg='#4169E1', relief='raised', bd=2)
        status_frame.place(relx=0.05, rely=0.9, relwidth=0.9, relheight=0.08)
        
        status_size = max(10, int(self.window_width / 100))
        status_label = tk.Label(status_frame,
                               text="🌟 准备就绪 - 原神风格数据处理工具 🌟",
                               font=('Microsoft YaHei UI', status_size, 'bold'),
                               fg='white',
                               bg='#4169E1')
        status_label.pack(expand=True)
        
    def open_file_selection(self):
        """打开文件选择界面"""
        # 隐藏主界面元素
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # 创建文件选择界面
        self.create_file_selection_interface()
        
    def create_file_selection_interface(self):
        """创建文件选择和处理界面"""
        # 二级窗口使用固定的合适大小，不受一级窗口影响
        secondary_width = self.default_width  # 900
        secondary_height = self.default_height  # 600
        
        # 调整窗口大小为二级窗口的合适尺寸
        self.root.geometry(f"{secondary_width}x{secondary_height}")
        
        # 重新居中窗口
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (secondary_width // 2)
        y = (self.root.winfo_screenheight() // 2) - (secondary_height // 2)
        self.root.geometry(f"{secondary_width}x{secondary_height}+{x}+{y}")
        
        # 设置背景色（二级窗口不使用图片背景）
        self.root.configure(bg='#E8F4FD')
        
        # 创建主容器 - 使用滚动框架确保所有内容都能显示
        main_canvas = tk.Canvas(self.root, bg='#FFFAF0', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg='#FFFAF0')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        # 放置画布和滚动条
        main_canvas.place(relx=0.02, rely=0.02, relwidth=0.94, relheight=0.96)
        scrollbar.place(relx=0.96, rely=0.02, relheight=0.96)
        
        # 在滚动框架中创建内容
        self.create_interface_content(scrollable_frame)
        
    def create_interface_content(self, parent):
        """在滚动框架中创建界面内容"""
        # 顶部标题和返回按钮
        top_frame = tk.Frame(parent, bg='#FFE4E1', relief='raised', bd=2)
        top_frame.pack(fill='x', pady=(10, 10), padx=10)
        
        # 返回按钮
        back_btn = ttk.Button(top_frame,
                             text="← 返回主界面",
                             style='Back.TButton',
                             command=self.back_to_main)
        back_btn.pack(side='left', padx=10, pady=10)
        
        # 标题
        title_label = tk.Label(top_frame,
                              text="📁 文件夹选择与处理",
                              font=('Microsoft YaHei UI', 14, 'bold'),
                              fg='#4169E1',
                              bg='#FFE4E1')
        title_label.pack(side='left', padx=20, pady=10)
        
        # 文件选择区域
        self.create_file_selection_section(parent)
        
        # 文件列表区域
        self.create_file_list_section(parent)
        
        # 处理按钮区域 - 确保这个区域被创建
        self.create_process_section(parent)
        
        # 状态栏
        self.create_status_section(parent)
        
        # 自动加载当前目录
        self.load_current_directory()
        
    def create_file_selection_section(self, parent):
        """创建文件选择区域"""
        selection_frame = tk.LabelFrame(parent,
                                       text="📂 选择文件夹",
                                       font=('Microsoft YaHei UI', 11, 'bold'),
                                       fg='#4169E1',
                                       bg='#FFFAF0',
                                       relief='groove', bd=2)
        selection_frame.pack(fill='x', pady=10, padx=10)
        
        # 输入路径显示
        path_frame = tk.Frame(selection_frame, bg='#FFFAF0')
        path_frame.pack(fill='x', padx=15, pady=10)
        
        tk.Label(path_frame, text="输入路径:",
                font=('Microsoft YaHei UI', 10, 'bold'),
                fg='#2F4F4F',
                bg='#FFFAF0').pack(side='left')
        
        self.path_var = tk.StringVar(value=self.selected_folder)
        path_label = tk.Label(path_frame, textvariable=self.path_var,
                             font=('Microsoft YaHei UI', 9),
                             fg='#4169E1',
                             bg='#FFFAF0',
                             wraplength=600)
        path_label.pack(side='left', padx=(10, 0))
        
        # 输出路径显示
        output_frame = tk.Frame(selection_frame, bg='#FFFAF0')
        output_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(output_frame, text="输出路径:",
                font=('Microsoft YaHei UI', 10, 'bold'),
                fg='#2F4F4F',
                bg='#FFFAF0').pack(side='left')
        
        self.output_path_var = tk.StringVar(value=self.output_folder)
        output_label = tk.Label(output_frame, textvariable=self.output_path_var,
                               font=('Microsoft YaHei UI', 9),
                               fg='#FF69B4',
                               bg='#FFFAF0',
                               wraplength=600)
        output_label.pack(side='left', padx=(10, 0))
        
        # 按钮区域
        button_frame = tk.Frame(selection_frame, bg='#FFFAF0')
        button_frame.pack(fill='x', padx=15, pady=(0, 15))
        
        current_btn = ttk.Button(button_frame,
                                text="📂 当前目录",
                                style='Process.TButton',
                                command=self.load_current_directory)
        current_btn.pack(side='left', padx=(0, 10))
        
        browse_btn = ttk.Button(button_frame,
                               text="🔍 选择输入目录",
                               style='Process.TButton',
                               command=self.browse_directory)
        browse_btn.pack(side='left', padx=(0, 10))
        
        output_btn = ttk.Button(button_frame,
                               text="📁 选择输出目录",
                               style='Process.TButton',
                               command=self.browse_output_directory)
        output_btn.pack(side='left')
        
    def create_file_list_section(self, parent):
        """创建文件列表区域"""
        list_frame = tk.LabelFrame(parent,
                                  text="📋 文件列表",
                                  font=('Microsoft YaHei UI', 11, 'bold'),
                                  fg='#4169E1',
                                  bg='#FFFAF0',
                                  relief='groove', bd=2)
        list_frame.pack(fill='x', pady=10, padx=10)
        
        # 表格容器 - 设置固定高度
        tree_frame = tk.Frame(list_frame, bg='#FFFAF0')
        tree_frame.pack(fill='x', padx=15, pady=15)
        
        # 滚动条
        scrollbar_y = ttk.Scrollbar(tree_frame)
        scrollbar_y.pack(side='right', fill='y')
        
        scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # 文件表格 - 设置固定高度
        self.file_tree = ttk.Treeview(tree_frame,
                                     style='Genshin.Treeview',
                                     yscrollcommand=scrollbar_y.set,
                                     xscrollcommand=scrollbar_x.set,
                                     height=8)  # 固定高度为8行
        
        # 配置列
        self.file_tree['columns'] = ('type', 'size', 'modified')
        self.file_tree.column('#0', width=300, minwidth=200)
        self.file_tree.column('type', width=80, minwidth=60)
        self.file_tree.column('size', width=100, minwidth=60)
        self.file_tree.column('modified', width=150, minwidth=80)
        
        # 配置标题
        self.file_tree.heading('#0', text='📄 文件名', anchor='w')
        self.file_tree.heading('type', text='📝 类型', anchor='center')
        self.file_tree.heading('size', text='📊 大小', anchor='center')
        self.file_tree.heading('modified', text='🕒 修改时间', anchor='center')
        
        self.file_tree.pack(fill='x')
        
        # 配置滚动条
        scrollbar_y.config(command=self.file_tree.yview)
        scrollbar_x.config(command=self.file_tree.xview)
        
        # 绑定双击事件
        self.file_tree.bind('<Double-1>', self.on_file_double_click)
        
    def create_process_section(self, parent):
        """创建处理按钮区域 - 确保按钮可见"""
        print("正在创建处理按钮区域...")  # 调试信息
        
        process_frame = tk.Frame(parent, bg='#FFE4E1', relief='raised', bd=3)
        process_frame.pack(fill='x', pady=20, padx=10)
        
        # 添加一个标题来确保区域可见
        title_label = tk.Label(process_frame,
                              text="🚀 数据处理操作",
                              font=('Microsoft YaHei UI', 12, 'bold'),
                              fg='#4169E1',
                              bg='#FFE4E1')
        title_label.pack(pady=(10, 5))
        
        # 处理按钮
        self.process_btn = ttk.Button(process_frame,
                                     text="✨ 开始处理数据 ✨",
                                     style='Process.TButton',
                                     command=self.start_processing,
                                     state='disabled')
        self.process_btn.pack(pady=(5, 15))
        
        print("处理按钮已创建")  # 调试信息
        
    def create_status_section(self, parent):
        """创建状态栏"""
        status_frame = tk.Frame(parent, bg='#4169E1', relief='sunken', bd=1)
        status_frame.pack(fill='x', pady=10, padx=10)
        
        self.status_var = tk.StringVar(value="🌟 请选择包含退货入库时效分析文件的文件夹")
        status_label = tk.Label(status_frame, textvariable=self.status_var,
                               font=('Microsoft YaHei UI', 9),
                               fg='white',
                               bg='#4169E1',
                               anchor='w')
        status_label.pack(fill='x', padx=10, pady=5)
        
    def back_to_main(self):
        """返回主界面"""
        # 清除文件选择界面
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # 恢复一级窗口的大小和位置
        self.root.geometry(f"{self.window_width}x{self.window_height}")
        self.center_window()
        
        # 重新创建主界面
        self.create_main_interface()
        
    def load_current_directory(self):
        """读取当前目录"""
        self.selected_folder = os.getcwd()
        self.path_var.set(self.selected_folder)
        self.load_files()
        
    def browse_directory(self):
        """浏览选择输入目录"""
        folder = filedialog.askdirectory(
            title="选择包含退货入库时效分析文件的文件夹",
            initialdir=self.selected_folder
        )
        if folder:
            self.selected_folder = folder
            self.path_var.set(folder)
            self.load_files()
            
    def browse_output_directory(self):
        """浏览选择输出目录"""
        folder = filedialog.askdirectory(
            title="选择输出文件保存的文件夹",
            initialdir=self.output_folder
        )
        if folder:
            self.output_folder = folder
            self.output_path_var.set(folder)
            
    def load_files(self):
        """加载文件列表"""
        self.status_var.set("🔄 正在扫描文件...")
        self.process_btn.config(state='disabled')
        
        # 清空现有列表
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
            
        # 在新线程中加载文件
        threading.Thread(target=self._load_files_worker, daemon=True).start()
        
    def _load_files_worker(self):
        """在工作线程中加载文件"""
        try:
            files = []
            for item in os.listdir(self.selected_folder):
                item_path = os.path.join(self.selected_folder, item)
                if os.path.isfile(item_path):
                    stat = os.stat(item_path)
                    files.append({
                        'name': item,
                        'path': item_path,
                        'type': self.get_file_type(item),
                        'size': self.format_file_size(stat.st_size),
                        'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M')
                    })
            
            # 在主线程中更新UI
            self.root.after(0, self._update_file_list_ui, files)
            
        except Exception as e:
            self.root.after(0, lambda: self.status_var.set(f"❌ 扫描失败: {str(e)}"))
            
    def _update_file_list_ui(self, files):
        """在主线程中更新文件列表UI"""
        self.files_data = files
        
        for file_info in files:
            self.file_tree.insert('', 'end',
                                 text=file_info['name'],
                                 values=(file_info['type'],
                                        file_info['size'],
                                        file_info['modified']))
        
        self.status_var.set(f"✅ 发现 {len(files)} 个文件")
        
        # 检查是否有可处理的文件
        processable_files = [f for f in files if f['name'].startswith('退货入库时效分析') and f['name'].endswith('.xlsx')]
        if processable_files:
            self.process_btn.config(state='normal')
            self.status_var.set(f"✅ 发现 {len(files)} 个文件，其中 {len(processable_files)} 个可处理")
            print(f"找到可处理文件，启用处理按钮")  # 调试信息
        else:
            self.status_var.set(f"⚠️ 发现 {len(files)} 个文件，但没有退货入库时效分析文件")
            
    def get_file_type(self, filename):
        """获取文件类型"""
        ext = os.path.splitext(filename)[1].lower()
        type_map = {
            '.xlsx': 'Excel',
            '.xls': 'Excel',
            '.csv': 'CSV',
            '.txt': '文本',
            '.pdf': 'PDF',
            '.py': 'Python',
            '.md': 'Markdown'
        }
        return type_map.get(ext, '其他')
        
    def format_file_size(self, size):
        """格式化文件大小"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} TB"
        
    def on_file_double_click(self, event):
        """文件双击事件"""
        selection = self.file_tree.selection()
        if selection:
            item = self.file_tree.item(selection[0])
            filename = item['text']
            messagebox.showinfo("📄 文件详情",
                              f"文件名: {filename}\n"
                              f"类型: {item['values'][0]}\n"
                              f"大小: {item['values'][1]}\n"
                              f"修改时间: {item['values'][2]}")
            
    def start_processing(self):
        """开始处理数据"""
        print("开始处理数据按钮被点击")  # 调试信息
        
        if not self.files_data:
            messagebox.showwarning("⚠️ 警告", "没有可处理的文件！")
            return
            
        # 显示确认对话框
        result = messagebox.askyesno("🌸 确认处理",
                                   f"确定要处理退货数据吗？\n\n"
                                   f"📁 输入路径: {self.selected_folder}\n"
                                   f"📁 输出路径: {self.output_folder}\n"
                                   f"📄 文件数量: {len(self.files_data)} 个\n\n"
                                   f"处理完成后将生成新的Excel分析报表。")
        
        if result:
            self.status_var.set("🚀 正在处理数据，请稍候...")
            self.process_btn.config(state='disabled')
            
            # 在新线程中处理数据
            threading.Thread(target=self._process_data_worker, daemon=True).start()
            
    def _process_data_worker(self):
        """在工作线程中处理数据"""
        try:
            # 使用新的数据处理器
            success, message = self.processor.process(self.selected_folder, self.output_folder)
            
            if success:
                self.root.after(0, lambda: self._processing_success(message))
            else:
                self.root.after(0, lambda: self._processing_error(message))
                
        except Exception as e:
            error_msg = f"处理过程中出现异常: {str(e)}"
            self.root.after(0, lambda: self._processing_error(error_msg))
            
    def _processing_success(self, message):
        """处理成功"""
        messagebox.showinfo("🎉 处理完成",
                          f"数据处理完成！\n\n"
                          f"✅ 已清理自营仓主体数据\n"
                          f"✅ 已创建时效分析表\n"
                          f"✅ 已生成数据透视表\n"
                          f"✅ 已计算满足率\n\n"
                          f"{message}")
        self.status_var.set("🌟 处理完成！可以选择新文件夹继续处理")
        self.process_btn.config(state='normal')
        
    def _processing_error(self, error_msg):
        """处理失败"""
        messagebox.showerror("❌ 处理失败", 
                           f"数据处理过程中出现错误：\n\n{error_msg}\n\n"
                           f"请检查文件格式是否正确。")
        self.status_var.set("❌ 处理失败，请检查文件格式")
        self.process_btn.config(state='normal')
        
    def run(self):
        """运行应用程序"""
        self.root.mainloop()

if __name__ == "__main__":
    print("启动退货数据处理工具...")
    app = GenshinSimpleApp()
    app.run()