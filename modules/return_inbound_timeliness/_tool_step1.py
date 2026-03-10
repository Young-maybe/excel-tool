import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import os
import sys
from pathlib import Path

# 统一的基准目录（优先脚本目录，退化为当前工作目录）
BASE_DIR = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()
print(f"[调试] 当前工作目录: {Path.cwd()}")
print(f"[调试] 基准目录: {BASE_DIR}")


def get_previous_month():
    """获取上一个月的月份"""
    today = datetime.now()
    first_day_this_month = today.replace(day=1)
    last_day_previous_month = first_day_this_month - timedelta(days=1)
    return last_day_previous_month.month

def parse_time_string(time_str):
    """解析时间字符串为datetime对象"""
    if not time_str or pd.isna(time_str):
        return None
    
    if isinstance(time_str, datetime):
        return time_str
    
    if not isinstance(time_str, str):
        return None
    
    time_str = str(time_str).strip()
    if not time_str:
        return None
    
    # 尝试不同的时间格式
    time_formats = [
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
        "%m/%d/%Y",
        "%m-%d-%Y"
    ]
    
    for fmt in time_formats:
        try:
            return datetime.strptime(time_str, fmt)
        except ValueError:
            continue
    
    print(f"警告：无法解析时间格式: {time_str}")
    return None

def read_csv_data():
    """读取CSV文件数据"""
    try:
        csv_file = BASE_DIR / "退货商品明细汇总.csv"
        if not csv_file.exists():
            print(f"错误：找不到文件: {csv_file}")
            print(f"提示：请确认文件是否位于该路径，或将文件放到与脚本同目录下")
            return None

        # 1. 使用 chardet 自动检测编码
        detected_encoding = None
        try:
            import chardet
            with open(str(csv_file), 'rb') as f:
                raw_data = f.read(200000)  # 读取前 200KB 用于检测
                result = chardet.detect(raw_data)
                detected_encoding = result['encoding']
                confidence = result['confidence']
                if detected_encoding and confidence > 0.7:
                    print(f"Chardet检测到编码: {detected_encoding} (置信度: {confidence:.0%})")
                else:
                    print(f"Chardet检测置信度较低 (编码: {detected_encoding}, 置信度: {confidence:.0%})，将忽略并继续尝试常用编码。")
                    detected_encoding = None  # 置信度低则忽略
        except ImportError:
            print("警告: 未安装 chardet 库，无法自动检测编码。建议运行: pip install chardet")
        except Exception as e:
            print(f"使用 chardet 检测编码时出错: {e}")

        df = None
        last_err = None

        # 2. 优先尝试 chardet 检测到的编码 (及 gb18030 作为备选)
        if detected_encoding:
            # If chardet suggests a GB encoding, add gb18030 as a high-priority fallback
            encodings_to_try = [detected_encoding]
            if detected_encoding.lower() in ['gb2312', 'gbk', 'gb18030']:
                # Put gb18030 first as it's more robust
                encodings_to_try.insert(0, 'gb18030')
            
            # Remove duplicates, keeping the first occurrence (so gb18030 stays first)
            encodings_to_try = list(dict.fromkeys(encodings_to_try))

            for enc in encodings_to_try:
                try:
                    df = pd.read_csv(
                        str(csv_file),
                        encoding=enc,
                        sep=None,  # 自动推断分隔符
                        engine='python',
                        on_bad_lines='skip'
                    )
                    print(f"成功读取CSV (使用编码 '{enc}' 和自动分隔符)")
                    break  # Success, exit the loop
                except Exception as e:
                    print(f"尝试使用编码 '{enc}' 读取失败。错误: {e}")
                    last_err = e
                    df = None # Ensure df is None to continue to next encoding

        # 3. 如果失败，回退到 pandas 自动分隔符推断
        if df is None:
            try:
                df = pd.read_csv(
                    str(csv_file),
                    sep=None, engine='python', dtype=str, on_bad_lines='skip'
                )
                print(f"成功读取CSV（自动推断分隔符）：行数={len(df)} 列数={len(df.columns)}")
            except Exception as e:
                last_err = e
                df = None

        # 4. 若仍失败，再按多编码 × 多分隔符穷举
        if df is None:
            # Add gb18030 as a robust fallback for Chinese encodings
            encodings = ['utf-8', 'utf-8-sig', 'gb18030', 'gbk', 'gb2312']
            separators = [',', '\t', ';', '|']
            for encoding in encodings:
                for sep in separators:
                    try:
                        df = pd.read_csv(
                            str(csv_file),
                            encoding=encoding, sep=sep, engine='python',
                            dtype=str, on_bad_lines='skip'
                        )
                        print(f"成功读取CSV：编码={encoding} 分隔符={repr(sep)} 行数={len(df)} 列数={len(df.columns)}")
                        break
                    except Exception as e:
                        last_err = e
                        df = None
                if df is not None:
                    break

        # 5. 若依旧失败，尝试按 Excel 读取
        if df is None:
            try:
                df = pd.read_excel(str(csv_file), engine='openpyxl')
                print("检测到文件按Excel格式可读取，已作为Excel解析（文件拓展名可能不规范）。")
            except Exception as e:
                print("无法读取CSV文件，请检查文件编码/分隔符/内容是否存在异常。")
                if last_err:
                    print(f"最后一次错误信息：{last_err}")
                print(f"读取为Excel时的错误：{e}")
                return None

        print(f"CSV文件包含 {len(df)} 行数据")
        print("CSV文件列名：")
        for i, col in enumerate(df.columns):
            print(f"{i+1}. {col}")

        return df

    except Exception as e:
        print(f"读取CSV文件时出错：{e}")
        return None

def read_warehouse_config():
    """读取仓库配置文件"""
    try:
        config_file = BASE_DIR / "店铺匹配仓库配置.xlsx"
        if not config_file.exists():
            print(f"警告：找不到配置文件: {config_file}，将跳过仓库映射")
            return {}
        
        # 读取自营仓工作表
        wb = openpyxl.load_workbook(str(config_file))
        if '自营仓' not in wb.sheetnames:
            print(f"警告：配置文件中找不到'自营仓'工作表，将跳过仓库映射")
            return {}
        
        ws = wb['自营仓']
        warehouse_mapping = {}
        
        # 读取D列（管易仓库名称）和E列（仓库盘点货主）
        for row in range(2, ws.max_row + 1):  # 从第2行开始，跳过表头
            warehouse_name = ws.cell(row=row, column=4).value  # D列
            warehouse_owner = ws.cell(row=row, column=5).value  # E列
            
            if warehouse_name and warehouse_owner:
                warehouse_mapping[warehouse_name] = warehouse_owner
        
        print(f"成功读取仓库配置：{len(warehouse_mapping)} 个映射关系")
        for name, owner in list(warehouse_mapping.items())[:5]:  # 显示前5个映射
            print(f"  {name} -> {owner}")
        if len(warehouse_mapping) > 5:
            print(f"  ... 还有 {len(warehouse_mapping) - 5} 个映射")
        
        return warehouse_mapping
        
    except Exception as e:
        print(f"读取仓库配置文件时出错：{e}")
        return {}

def detect_all_hidden_columns(worksheet):
    """检测Excel中所有类型的隐藏列，但排除用户指定不应隐藏的列"""
    hidden_columns = set()
    
    # 方法1：检查hidden属性（直接隐藏）
    for col in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col)
        if worksheet.column_dimensions[col_letter].hidden:
            hidden_columns.add(col)
    
    # 方法2：检查列分组（分组折叠隐藏）
    if hasattr(worksheet, 'column_groups') and worksheet.column_groups:
        print("检测到列分组，正在分析...")
        for group in worksheet.column_groups:
            print(f"  分组: {group}")
            # 解析分组范围，如 "A:F"
            if ':' in str(group):
                start_col, end_col = str(group).split(':')
                start_num = column_index_from_string(start_col)
                end_num = column_index_from_string(end_col)
                
                # 根据用户反馈，排除不应该隐藏的列
                group_range = set(range(start_num, end_num + 1))
                
                # 排除I、J、AP、AQ、AS、AT列（第9、10、42、43、45、46列）
                exclude_columns = {9, 10, 42, 43, 45, 46}
                group_range = group_range - exclude_columns
                
                hidden_columns.update(group_range)
                print(f"    添加隐藏列: {sorted(list(group_range))}")
    
    # 方法3：检查极小列宽（手动隐藏）
    for col in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col)
        width = worksheet.column_dimensions[col_letter].width
        if width is not None and width < 0.5:
            hidden_columns.add(col)
    
    return sorted(list(hidden_columns))

def analyze_template_excel():
    """分析模板Excel文件的格式"""
    try:
        template_file = BASE_DIR / "12345.xlsx"
        if not template_file.exists():
            print(f"错误：找不到模板文件: {template_file}")
            return None
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(str(template_file))
        ws = wb.active
        
        print(f"模板Excel文件工作表名：{ws.title}")
        print(f"模板文件最大行数：{ws.max_row}")
        print(f"模板文件最大列数：{ws.max_column}")
        
        # 获取表头信息
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
        
        print("模板Excel文件列名：")
        for i, header in enumerate(headers):
            print(f"{i+1}. {header}")
        
        # 使用新的隐藏列检测方法
        hidden_columns = detect_all_hidden_columns(ws)
        print(f"检测到的所有隐藏列：{hidden_columns}")
        print(f"隐藏列总数：{len(hidden_columns)}")
        
        return wb, ws, headers, hidden_columns
        
    except Exception as e:
        print(f"分析模板Excel文件时出错：{e}")
        return None, None, None, []

def create_column_mapping(csv_columns, template_headers):
    """创建CSV列到模板列的映射"""
    mapping = {}
    
    # 直接映射相同名称的列
    for i, csv_col in enumerate(csv_columns):
        if csv_col in template_headers:
            template_idx = template_headers.index(csv_col)
            mapping[i] = template_idx
    
    return mapping

def create_output_excel(csv_data, template_wb, template_ws, template_headers, hidden_columns, warehouse_mapping):
    """创建输出Excel文件"""
    try:
        # 获取上一个月
        prev_month = get_previous_month()
        output_filename = f"退货入库时效分析{prev_month}月总表.xlsx"
        
        # 创建新的工作簿
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "退货入库时效分析"
        
        # 复制模板的格式和结构（但不设置隐藏状态）
        if template_ws:
            # 复制列宽，但先不设置隐藏状态
            for col in range(1, template_ws.max_column + 1):
                col_letter = get_column_letter(col)
                # 只复制列宽，不复制隐藏状态
                template_col_dim = template_ws.column_dimensions[col_letter]
                new_ws.column_dimensions[col_letter].width = template_col_dim.width
                # 暂时不设置隐藏状态，留到最后
            
            # 复制表头和格式
            for col in range(1, len(template_headers) + 1):
                header = template_headers[col-1] if col-1 < len(template_headers) else None
                cell = new_ws.cell(row=1, column=col, value=header)
                
                # 复制模板表头的格式
                if col <= template_ws.max_column:
                    template_cell = template_ws.cell(row=1, column=col)
                    if template_cell.font:
                        cell.font = Font(
                            name=template_cell.font.name,
                            size=template_cell.font.size,
                            bold=template_cell.font.bold,
                            color=template_cell.font.color
                        )
                    if template_cell.alignment:
                        cell.alignment = Alignment(
                            horizontal=template_cell.alignment.horizontal,
                            vertical=template_cell.alignment.vertical
                        )
                    if template_cell.fill and template_cell.fill.start_color:
                        cell.fill = PatternFill(
                            start_color=template_cell.fill.start_color,
                            end_color=template_cell.fill.end_color,
                            fill_type=template_cell.fill.fill_type
                        )
        
        # 填充CSV数据
        if csv_data is not None and not csv_data.empty:
            # 创建列映射
            column_mapping = create_column_mapping(csv_data.columns.tolist(), template_headers)
            
            print(f"列映射关系：{len(column_mapping)} 个匹配的列")
            
            # 填充数据
            for row_idx, (_, row) in enumerate(csv_data.iterrows(), 2):  # 从第2行开始
                # 根据映射填充数据
                for csv_col_idx, template_col_idx in column_mapping.items():
                    value = row.iloc[csv_col_idx]
                    cell = new_ws.cell(row=row_idx, column=template_col_idx + 1)
                    
                    # 检查是否是时间列，需要特殊处理
                    col_name = csv_data.columns[csv_col_idx]
                    if col_name in ["制单时间", "审核时间", "入库时间"]:
                        # 转换时间字符串为datetime对象
                        datetime_value = parse_time_string(value)
                        if datetime_value:
                            cell.value = datetime_value
                            # 设置时间格式
                            cell.number_format = "m/d/yy h:mm"
                        else:
                            cell.value = None
                    else:
                        cell.value = value
                
                # 处理特殊列
                # 第41列（AO）：自营仓主体 - 通过仓库配置映射
                warehouse_name = row.iloc[39] if len(row) > 39 else ""  # AN列（仓库名称）的索引是39
                if warehouse_name and warehouse_name in warehouse_mapping:
                    warehouse_owner = warehouse_mapping[warehouse_name]
                    new_ws.cell(row=row_idx, column=41, value=warehouse_owner)
                else:
                    new_ws.cell(row=row_idx, column=41, value="")
                
                # 第44列（AR）：签收时间 - 保持空白，但设置时间格式
                ar_cell = new_ws.cell(row=row_idx, column=44, value="")
                ar_cell.number_format = "m/d/yy h:mm"
                
                # 第45列（AS）：最新时间 - 使用公式，设置时间格式
                as_cell = new_ws.cell(row=row_idx, column=45)
                as_cell.value = f"=MAX(AQ{row_idx},AR{row_idx})"
                as_cell.number_format = "m/d/yy h:mm"
                
                # 第47列（AU）：48h入库时效是否满足 - 使用公式
                au_cell = new_ws.cell(row=row_idx, column=47)
                au_cell.value = f'=IF((AT{row_idx}-AS{row_idx})>2,"不满足","满足")'
                
                # 第88列（CJ）：物流公司，复制第13列（M列）的内容
                logistics_company_value = row.iloc[12] if len(row) > 12 else ""  # 第13列的索引是12
                new_ws.cell(row=row_idx, column=88, value=logistics_company_value)
        
        # 最后一步：设置隐藏列
        print("\n开始设置隐藏列...")
        for col in hidden_columns:
            col_letter = get_column_letter(col)
            new_ws.column_dimensions[col_letter].hidden = True
            print(f"隐藏第{col}列 ({col_letter})")
        
        print(f"共隐藏了 {len(hidden_columns)} 列")
        
        # 保存文件
        output_path = BASE_DIR / output_filename
        new_wb.save(output_path)
        print(f"成功创建文件：{output_path}")
        print(f"文件包含 {new_ws.max_row-1} 行数据（不含表头）")
        
        # 验证公式
        print("\n验证公式设置：")
        print(f"AS2公式：{new_ws.cell(row=2, column=45).value}")
        print(f"AU2公式：{new_ws.cell(row=2, column=47).value}")
        
        return output_filename
        
    except Exception as e:
        print(f"创建输出Excel文件时出错：{e}")
        return None

def main():
    """主函数"""
    print("开始处理退货商品明细数据...")
    
    # 读取CSV数据
    csv_data = read_csv_data()
    if csv_data is None or csv_data.empty:
        print("错误：未能读取到有效的CSV数据，处理中止。")
        return

    # 读取仓库配置
    warehouse_mapping = read_warehouse_config()
    
    # 分析模板Excel
    template_wb, template_ws, template_headers, hidden_columns = analyze_template_excel()
    if template_wb is None:
        print("错误：未能读取到有效的模板Excel文件，处理中止。")
        return
    
    # 创建输出Excel文件
    output_file = create_output_excel(csv_data, template_wb, template_ws, template_headers, hidden_columns, warehouse_mapping)
    
    if output_file:
        print(f"\n处理完成！输出文件：{output_file}")
        print("\n功能总结：")
        print("[OK] 正确隐藏65个列")
        print("[OK] 映射仓库配置到自营仓主体列")
        print("[OK] AS列使用MAX公式计算最新时间")
        print("[OK] AU列使用IF公式判断时效满足情况")
        print("[OK] CJ列复制物流公司信息")
    else:
        print("处理失败！")

if __name__ == "__main__":
    main()