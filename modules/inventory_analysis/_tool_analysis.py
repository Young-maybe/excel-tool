import os
import re
import sys
from pathlib import Path
import difflib
import warnings
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore", category=UserWarning)

# 假设程序与数据在同一目录下运行
try:
    BASE_DIR = Path(__file__).resolve().parent
except NameError:
    # 在 Jupyter/交互环境下 __file__ 不存在，退回到当前工作目录
    BASE_DIR = Path.cwd()

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 工具函数

def list_excel_files() -> List[Path]:
    return [p for p in BASE_DIR.glob("*.xlsx")]


def find_first(pattern: str) -> Optional[Path]:
    # 使用通配符匹配，返回第一个匹配项（按名称排序以保证稳定）
    candidates = sorted(BASE_DIR.glob(pattern))
    return candidates[0] if candidates else None


def find_by_regex(regex: str) -> Optional[Path]:
    r = re.compile(regex)
    cands = sorted([p for p in list_excel_files() if r.search(p.name)])
    return cands[0] if cands else None


def best_match_file(target_name: str, excludes: List[Path]) -> Optional[Path]:
    target = target_name.lower()
    best: Tuple[float, Optional[Path]] = (0.0, None)
    exclude_names = {p.name for p in excludes}
    for p in list_excel_files():
        if p.name in exclude_names:
            continue
        name_no_ext = p.stem.lower()
        score = difflib.SequenceMatcher(None, target, name_no_ext).ratio()
        if score > best[0]:
            best = (score, p)
    return best[1]


def get_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    # 根据候选列名列表，返回最匹配（忽略大小写、空格）的列
    norm = {c: re.sub(r"\s", "", str(c)).lower() for c in df.columns}
    cand_norm = [re.sub(r"\s", "", c).lower() for c in candidates]
    # 完全匹配优先
    for i, cn in enumerate(cand_norm):
        for c, nc in norm.items():
            if nc == cn:
                return c
    # 包含匹配
    for i, cn in enumerate(cand_norm):
        for c, nc in norm.items():
            if cn in nc:
                return c
    # 模糊相似度
    best = (0.0, None)
    for c, nc in norm.items():
        for cn in cand_norm:
            score = difflib.SequenceMatcher(None, cn, nc).ratio()
            if score > best[0]:
                best = (score, c)
    return best[1]


def safe_read_excel(path: Path, sheet_name=None, header=0) -> Optional[pd.DataFrame]:
    try:
        return pd.read_excel(path, sheet_name=sheet_name, header=header, dtype=str)
    except Exception:
        return None


def to_number(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype=float)
    s = series.astype(str)
    # 去除千分位逗号与括号负号格式
    s = s.str.replace(",", "", regex=False)
    s = s.str.replace("(", "-", regex=False).str.replace(")", "", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0)


def build_mapping(df: pd.DataFrame, key_candidates: List[str], value_candidates: List[str]) -> Dict[str, any]:
    if df is None or df.empty:
        return {}
    key_col = get_col(df, key_candidates)
    val_col = get_col(df, value_candidates)
    if not key_col or not val_col:
        return {}
    s = df[[key_col, val_col]].dropna()
    return dict(zip(s[key_col].astype(str), s[val_col]))


def sum_by_key(df: pd.DataFrame, key_candidates: List[str], value_candidates: List[str]) -> Dict[str, float]:
    if df is None or df.empty:
        return {}
    key_col = get_col(df, key_candidates)
    val_col = get_col(df, value_candidates)
    if not key_col or not val_col:
        return {}
    tmp = df[[key_col, val_col]].copy()
    tmp[key_col] = tmp[key_col].astype(str)
    tmp[val_col] = to_number(tmp[val_col])
    grp = tmp.groupby(key_col, as_index=False)[val_col].sum()
    return dict(zip(grp[key_col], grp[val_col]))


def process_workbook():
    # 1) 找到主文件：附件一：自营仓盘点表?月.xlsx
    main_file = find_by_regex(r"附件一：自营仓盘点表.月\.xlsx$") or find_first("附件一：自营仓盘点表*月.xlsx")
    if not main_file:
        print("未找到主文件：附件一：自营仓盘点表?月.xlsx")
        return

    # 2) 另存为同名修改后的文件（追加 -修改后）
    out_file = main_file.with_name(f"{main_file.stem}-修改后{main_file.suffix}")

    # 3) 预加载对照源文件
    sku_export = find_first("商品库存导出*.xlsx")
    snapshot = find_by_regex(r"库存快照明细.*修改后\.xlsx$")
    costbook = find_first("商品+供应商库持续更新版*.xlsx")

    # 3.1 商品库存导出：sheet1（名称、条码），与各sheet同名（可配数）
    sku_sheet1 = safe_read_excel(sku_export, sheet_name=0, header=0) if sku_export else None

    # 映射：商品名称、商品条码
    map_name = build_mapping(
        sku_sheet1,
        key_candidates=["商品规格代码", "规格代码", "SKU", "规格编码"],
        value_candidates=["商品名称", "名称", "品名"]
    )
    map_barcode = build_mapping(
        sku_sheet1,
        key_candidates=["商品规格代码", "规格代码", "SKU", "规格编码"],
        value_candidates=["商品条码", "条码", "UPC", "EAN"]
    )

    # 3.2 百合截单数量：库存快照明细*修改后.xlsx 的 透视表sheet（规格代码-总计）
    snap_df = safe_read_excel(snapshot, sheet_name="透视表", header=0) if snapshot else None
    map_baihe_total = sum_by_key(
        snap_df,
        key_candidates=["规格代码", "商品规格代码", "SKU"],
        value_candidates=["总计", "合计", "数量"]
    )

    # 3.3 成本：商品+供应商库持续更新版*.xlsx 的 商品库 sheet（clean-成本）
    cost_df = safe_read_excel(costbook, sheet_name="商品库", header=0) if costbook else None
    map_cost = build_mapping(
        cost_df,
        key_candidates=["规格代码", "商品规格代码", "SKU", "规格编码"],
        value_candidates=["clean-成本", "成本", "单位成本"]
    )

    # 4) 打开主工作簿并处理每个sheet
    wb = load_workbook(filename=str(main_file))

    # 记住每个sheet名字
    sheet_names = wb.sheetnames
    print("处理工作表：", sheet_names)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # 读取A列（规格代码），表头在第二行，数据从第3行开始
        header_row = 2
        start_row = 3

        # 找到最后一行（按A列非空）
        last_row = start_row - 1
        for r in range(start_row, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() != "":
                last_row = r

        if last_row < start_row:
            # 当前sheet没有数据，跳过
            continue

        # 准备与当前sheet同名的可配数来源
        sku_same_df = safe_read_excel(sku_export, sheet_name=sheet_name, header=0) if sku_export else None
        map_kep = sum_by_key(
            sku_same_df,
            key_candidates=["商品规格代码", "规格代码", "SKU", "规格编码"],
            value_candidates=["可配数(求和)", "可配数", "可配"]
        )

        # 为WMS/实物选择最相似excel并读取 盘点表 sheet
        excludes = [p for p in [main_file, sku_export, snapshot, costbook] if p is not None]
        wms_file = best_match_file(sheet_name, excludes)
        print(f"[{sheet_name}] 匹配到WMS来源文件：{wms_file.name if wms_file else '未找到'}")
        wms_df = safe_read_excel(wms_file, sheet_name="盘点表", header=0) if wms_file else None
        # 仅使用盘点表从 M 列开始的区域进行匹配（忽略 M 列之前的信息）
        if wms_df is not None and not wms_df.empty and wms_df.shape[1] > 12:
            wms_df = wms_df.iloc[:, 12:].copy()
        map_wms = sum_by_key(
            wms_df,
            key_candidates=["规格代码", "商品规格代码", "SKU", "规格编码"],
            value_candidates=["在库数量(求和)", "在库数量", "在库", "合计数量", "合计在库", "库存数量"]
        )
        map_real = sum_by_key(
            wms_df,
            key_candidates=["规格代码", "商品规格代码", "SKU", "规格编码"],
            value_candidates=["盘点数(求和)", "盘点数", "实物盘点数量", "实物", "实盘数量", "实盘"]
        )

        # 在第二行写入（或更新）表头至指定列位
        headers = {
            "B": "商品名称",
            "C": "货品条码",
            "D": "百合截单数量",
            "E": "WMS结存数量",
            "F": "实物盘点数量",
            "G": "差异1=实物-WMS",
            "H": "商品库单位成本",
            "I": "库存成本",
            "J": "管易可配数",
            "L": "实际差异=实物-可配",
            "M": "盘盈亏金额",
        }
        for col_letter, title in headers.items():
            ws[f"{col_letter}{header_row}"] = title

        # 填充数据
        for r in range(start_row, last_row + 1):
            key = ws.cell(row=r, column=1).value
            key = "" if key is None else str(key).strip()

            # 跳过汇总行，避免覆盖总计/合计等公式行
            if key and ("合计" in key or "总计" in key):
                continue

            # 基础映射
            name = map_name.get(key, "")
            barcode = map_barcode.get(key, "")
            baihe = map_baihe_total.get(key, 0)
            wms_qty = map_wms.get(key, 0)
            real_qty = map_real.get(key, 0)
            unit_cost = map_cost.get(key, 0)
            kep = map_kep.get(key, 0)

            # 写入基础列（若目标单元已是公式则跳过）
            def write_if_not_formula(addr, val):
                cell = ws[addr]
                cv = cell.value
                if isinstance(cv, str) and cv.startswith("="):
                    return
                cell.value = val

            write_if_not_formula(f"B{r}", name)
            write_if_not_formula(f"C{r}", barcode)
            write_if_not_formula(f"D{r}", float(baihe) if baihe != "" else 0)
            write_if_not_formula(f"E{r}", float(wms_qty) if wms_qty != "" else 0)
            write_if_not_formula(f"F{r}", float(real_qty) if real_qty != "" else 0)
            write_if_not_formula(f"H{r}", float(unit_cost) if unit_cost != "" else 0)
            write_if_not_formula(f"J{r}", float(kep) if kep != "" else 0)

            # 计算列（同样尊重已有公式）
            try:
                diff1 = float(real_qty) - float(wms_qty)
            except Exception:
                diff1 = 0
            write_if_not_formula(f"G{r}", diff1)
            if diff1 != 0:
                ws[f"G{r}"].fill = YELLOW_FILL

            try:
                stock_cost = float(unit_cost) * float(real_qty)
            except Exception:
                stock_cost = 0
            write_if_not_formula(f"I{r}", stock_cost)

            try:
                diff_actual = float(real_qty) - float(kep)
            except Exception:
                diff_actual = 0
            write_if_not_formula(f"L{r}", diff_actual)
            if diff_actual != 0:
                ws[f"L{r}"].fill = YELLOW_FILL

            pnl_amt = abs(float(unit_cost) * float(diff_actual))
            write_if_not_formula(f"M{r}", pnl_amt)

    # 保存到修改后文件
    wb.save(str(out_file))
    print(f"已生成：{out_file.name}")


if __name__ == "__main__":
    process_workbook()
