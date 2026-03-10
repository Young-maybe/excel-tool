"""
猴面包树B2B发货 - 箱唛转换

严格保持原脚本的核心处理逻辑，仅做适配：
- 输入：从 input_dir/猴面包树B2B箱单.xlsx 读取
- 输出：写入 output_dir/猴面包树箱唛-转换后.xlsx

说明：原作者脚本为两段输出：
1) 先生成临时箱唛文件（test11.xlsx）→ 修正箱号 → 保存 猴面包树箱唛.xlsx
2) 再执行格式转换 → 生成 猴面包树箱唛-转换后.xlsx
本模块以第二段“转换后”的结果作为最终输出。
"""
from __future__ import annotations

import os
import re
from typing import Callable, List, Optional

import openpyxl
from openpyxl.styles import Border, Font, Side
import pandas as pd
import xlsxwriter


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> List[str]:
    def report(p: int, s: str) -> None:
        if progress_cb:
            progress_cb(p, s)

    source_path = os.path.join(input_dir, "猴面包树B2B箱单.xlsx")
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"未找到源文件：{source_path}")

    os.makedirs(output_dir, exist_ok=True)

    report(5, "读取箱单 Excel…")
    raw_data = pd.read_excel(source_path)

    tmp_path = os.path.join(output_dir, "_tmp_猴面包树箱唛_raw.xlsx")
    base_path = os.path.join(output_dir, "猴面包树箱唛.xlsx")
    converted_path = os.path.join(output_dir, "猴面包树箱唛-转换后.xlsx")

    wb = xlsxwriter.Workbook(tmp_path)

    lines: List[List] = []
    sheet_name: Optional[str] = None
    created_sheets = set()
    count1 = 0

    def flush_sheet(name: str, data: List[List]) -> None:
        if not data:
            return
        if name in created_sheets:
            return
        ws = wb.add_worksheet(name[:31])
        created_sheets.add(name)
        for r, line in enumerate(data):
            for c, item in enumerate(line):
                ws.write(r, c, item)

    total_rows = max(len(raw_data) - 1, 1)
    report(12, f"生成临时箱唛（共 {total_rows} 行）…")
    for idx, (_, row) in enumerate(raw_data[1:].iterrows(), start=1):
        if progress_cb and (idx % 500 == 0 or idx == total_rows):
            report(12 + int(idx / total_rows * 38), f"写入临时箱唛… {idx}/{total_rows}")
        if pd.isnull(row["主订单-渠道订单号"]):
            if sheet_name is not None and lines:
                flush_sheet(sheet_name, lines)
                lines = []
            continue

        text = str(row["主订单-买家留言"])
        first_four = text[:4]

        if "-" in first_four or "（" in first_four or "(" in first_four:
            index = text.find("）")
            if index == -1:
                index = text.find(")")
            if index != -1:
                first_four = text[: index + 1]
            else:
                first_four = text[:5]

        cleaned_buyer_message = re.sub(r'[\/:*?"<>|（）()]', "", first_four)

        a = row["商品编码"]
        c = row["商品名称"]
        d = row["主订单-渠道订单号"]
        f = row["SDO"]
        name_str = str(row["主订单-收件信息-收货人"])
        split_chars = ["/", "、", "和", "及", "，", ","]
        for char in split_chars:
            if char in name_str:
                wu = name_str.split(char)[0]
                break
        else:
            wu = name_str

        g = cleaned_buyer_message + "-" + wu
        h = int(float(str(row["商品数量"])))
        zhengshu = int(row["总箱数"])

        new_sheet_name = f"箱唛-{g}"

        if sheet_name != new_sheet_name:
            if sheet_name is not None and lines:
                flush_sheet(sheet_name, lines)
                lines = []
            sheet_name = new_sheet_name

        pattern_pinx1 = r"拼箱\*(\d+)"
        pattern_a_b_c_d = r"^(\d+)\*(\d+)\+(\d+)\*(\d+)$"
        pattern_e_f = r"^(\d+)\*([^\+\*]+)$"

        match = re.match(pattern_pinx1, str(row["箱单"]))

        if match:  # 拼箱
            pinx_number = int(match.group(1))
            pinx_info = f"拼箱*{pinx_number}"
            pinx_xiangh = str(pinx_number) + "/" + str(zhengshu)
            lines.append(["条码", "{}".format(a), "箱号", "{}".format(pinx_xiangh)])
            lines.append(["名称", "{}".format(c), "采购单号", "{}".format(d)])
            lines.append(["数量", "{}".format(h), "发货单号", "{}".format(f)])
            lines.append(["收货地", "{}".format(g), "备注", "{}".format(pinx_info)])
            lines.append(["", "", "", ""])

        if re.match(pattern_a_b_c_d, str(row["箱单"])):  # a*b+c*d
            match_abcd = re.match(pattern_a_b_c_d, str(row["箱单"]))
            if match_abcd:
                a1 = int(match_abcd.group(1))
                b1 = int(match_abcd.group(2))
                c1 = int(match_abcd.group(3))
                count2 = count1
                d1 = count1 + b1 + 2
                for e in range(count1 + 1, d1):
                    q_val = str(count2 + 1)
                    b = "{}/{}".format(q_val, str(zhengshu))
                    if e == d1 - 1:
                        quantity = c1
                    else:
                        quantity = a1
                    lines.append(["条码", "{}".format(a), "箱号", "{}".format(b)])
                    lines.append(["名称", "{}".format(c), "采购单号", "{}".format(d)])
                    lines.append(["数量", "{}".format(quantity), "发货单号", "{}".format(f)])
                    lines.append(["收货地", "{}".format(g), "备注", "{}".format("")])
                    lines.append(["", "", "", ""])
                    count2 += 1

        if re.match(pattern_e_f, str(row["箱单"])):  # e*f
            match_ef = re.match(pattern_e_f, str(row["箱单"]))
            if match_ef:
                e1 = int(match_ef.group(1))
                f1 = int(match_ef.group(2))
                for z in range(1, f1 + 1):
                    b = "{}/{}".format(str(z), str(zhengshu))
                    lines.append(["条码", "{}".format(a), "箱号", "{}".format(b)])
                    lines.append(["名称", "{}".format(c), "采购单号", "{}".format(d)])
                    lines.append(["数量", "{}".format(e1), "发货单号", "{}".format(f)])
                    lines.append(["收货地", "{}".format(g), "备注", "{}".format("")])
                    lines.append(["", "", "", ""])

        count1 += 1

    if sheet_name is not None and lines:
        flush_sheet(sheet_name, lines)

    wb.close()

    def update_box_numbers(sheet) -> None:
        pinxiang_rows = []
        pinxiang_numbers = []
        for row_index, _remark_cell in enumerate(sheet["D"], start=1):
            cell_value = sheet.cell(row=row_index, column=4).value
            match = re.search(r"拼箱\*(\d+)", str(cell_value))
            if match:
                pinxiang_number = int(match.group(1))
                pinxiang_numbers.append(pinxiang_number)
                pinxiang_rows.append(row_index - 2)

        current_box_number = max(pinxiang_numbers) + 1 if pinxiang_numbers else 1

        for row_index in range(1, sheet.max_row + 1, 5):
            if row_index in pinxiang_rows:
                continue

            box_number_cell = sheet.cell(row=row_index, column=4)
            box_number_value = box_number_cell.value
            remark_cell = sheet.cell(row=row_index + 3, column=4)
            remark_value = remark_cell.value.strip() if remark_cell.value else ""

            if re.match(r".*拼箱.*", remark_value):
                continue

            box_number_parts = str(box_number_value).split("/")
            if len(box_number_parts) == 2:
                box_number_cell.value = f"{current_box_number}/{box_number_parts[1]}"
                current_box_number += 1

    def convert_format(input_file: str) -> str:
        wb_local = openpyxl.load_workbook(input_file)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        bold_font = Font(bold=True)

        for sheet_name_local in wb_local.sheetnames:
            ws = wb_local[sheet_name_local]

            # 插入左侧空列和顶部空行
            ws.insert_cols(1)
            ws.insert_rows(1)

            # 应用格式（与原脚本一致）
            start_row = 2
            while start_row <= ws.max_row:
                if ws.cell(row=start_row, column=2).value is None:
                    start_row += 1
                    continue

                end_row = start_row + 3
                for row_i in range(start_row, end_row + 1):
                    for col_i in range(2, 6):  # B到E列
                        cell = ws.cell(row=row_i, column=col_i)
                        cell.font = bold_font
                        cell.border = thin_border

                start_row = end_row + 2

        output_file = input_file.replace(".xlsx", "-转换后.xlsx")
        wb_local.save(output_file)
        return output_file

    try:
        report(55, "保存临时箱唛文件…")
        wb.close()

        # 第二段前置：修正箱号并生成 猴面包树箱唛.xlsx
        report(65, "修正箱号并生成箱唛.xlsx…")
        wb_fix = openpyxl.load_workbook(tmp_path)
        report(70, "逐个sheet修正箱号…")
        for s in wb_fix.sheetnames:
            update_box_numbers(wb_fix[s])
        wb_fix.save(base_path)
        wb_fix.close()

        # 第二段：格式转换，产出最终文件
        report(80, "执行格式转换（生成“转换后”）…")
        final_path = convert_format(base_path)

        report(90, "清理临时文件并统一命名…")
        # 只保留最终输出（符合“转换后作为输出文件”）
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        if os.path.exists(base_path):
            os.remove(base_path)

        # final_path 会是 base_path 的 -转换后
        if final_path != converted_path and os.path.exists(final_path):
            # 统一命名到固定输出名，避免重复后缀
            if os.path.exists(converted_path):
                os.remove(converted_path)
            os.replace(final_path, converted_path)
            final_path = converted_path

        report(100, "完成")
        return [final_path]
    finally:
        # 若中途异常，尽量清理临时文件（保留 base 方便排查）
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


