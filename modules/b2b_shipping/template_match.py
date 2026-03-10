"""
猴面包树B2B发货 - 送货单与模板匹配（需要备货单的规格箱数）

严格保持原脚本的核心处理逻辑，仅做适配：
- 模板目录：在 input_dir/送货单模板
- 待处理送货单：从 output_dir 查找（因为步骤1输出统一到 output_dir）
- 输出：在 output_dir 原地更新送货单文件
"""
from __future__ import annotations

import glob
import os
import re
from copy import copy
from difflib import SequenceMatcher
from typing import Callable, List, Optional

import openpyxl


def normalize_header(text):
    """最终版：清洗并标准化表头文本，用于模糊匹配。"""
    if not text:
        return ""
    cleaned_text = re.sub(r"[\s/()（）]+", "", str(text)).lower()
    alias_map = {
        "主订单-渠道订单号": "purchase_order",
        "采购订单": "purchase_order",
        "采购单号": "purchase_order",
        "采购订单号": "purchase_order",
        "商品编码": "book_id",
        "书号条码isbn": "book_id",
        "书号条码": "book_id",
        "商品名称": "book_name",
        "书名": "book_name",
        "商品数量": "quantity",
        "数量本": "quantity",
        "sdo单号": "sdo",
        "sdo": "sdo",
        "送货时间": "delivery_time",
        "预约时间": "delivery_time",
        "码洋": "mayang",
        "实洋": "shiyang",
        "折扣": "discount",
        "定价": "price",
    }
    return alias_map.get(cleaned_text, cleaned_text)


def get_text_similarity(a, b):
    a = re.sub(r"送货单|模板|系统专用|\.xlsx", "", a)
    b = re.sub(r"送货单|模板|系统专用|\.xlsx", "", b)
    return SequenceMatcher(None, a, b).ratio()


def find_data_blocks(worksheet, key_column_for_matching):
    blocks = []
    start_row = 0
    key_column_for_block_detection = "Q"
    for row in range(3, worksheet.max_row + 2):
        is_data_row = worksheet[f"{key_column_for_block_detection}{row}"].value is not None
        if is_data_row and start_row == 0:
            start_row = row
        elif not is_data_row and start_row != 0:
            end_row = row - 1
            key_text = worksheet[f"{key_column_for_matching}{start_row}"].value or ""
            blocks.append({"start": start_row, "end": end_row, "key_text": str(key_text)})
            start_row = 0
    return blocks


def clone_sheet(source_ws, target_wb, new_title):
    target_ws = target_wb.create_sheet(title=new_title)
    for r_idx, row in enumerate(source_ws.iter_rows(), 1):
        for c_idx, cell in enumerate(row, 1):
            new_cell = target_ws.cell(row=r_idx, column=c_idx)
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.protection, new_cell.alignment = (
                    copy(cell.font),
                    copy(cell.border),
                    copy(cell.fill),
                    cell.number_format,
                    copy(cell.protection),
                    copy(cell.alignment),
                )
    for row_id, dim in source_ws.row_dimensions.items():
        if dim.height is not None:
            target_ws.row_dimensions[row_id].height = dim.height
    for col_id, dim in source_ws.column_dimensions.items():
        if dim.width is not None:
            target_ws.column_dimensions[col_id].width = dim.width
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    return target_ws


def find_address_column(worksheet) -> Optional[str]:
    header_to_find = "主订单-收件信息-收货地址"
    for row in [1, 2]:
        for cell in worksheet[row]:
            if cell.value and str(cell.value).strip() == header_to_find:
                return cell.column_letter
    return None


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> List[str]:
    def report(p: int, s: str) -> None:
        if progress_cb:
            progress_cb(p, s)

    TEMPLATE_SUBFOLDER, DELIVERY_NOTE_SUFFIX = "送货单模板", "送货单.xlsx"

    template_directory = os.path.join(input_dir, TEMPLATE_SUBFOLDER)
    if not os.path.isdir(template_directory):
        raise FileNotFoundError(f"未找到模板子文件夹：{template_directory}")

    delivery_notes = [f for f in os.listdir(output_dir) if f.endswith(DELIVERY_NOTE_SUFFIX)]
    if not delivery_notes:
        raise FileNotFoundError("未在输出目录中找到任何“送货单.xlsx”。请先运行“送货单、备货单生成”。")

    report(10, "扫描模板文件…")
    all_template_files = glob.glob(os.path.join(template_directory, "*.xlsx"))
    if not all_template_files:
        raise FileNotFoundError(f"模板文件夹中未找到任何 .xlsx：{template_directory}")

    outputs: List[str] = []
    total = len(delivery_notes)
    report(5, f"扫描输出目录送货单…（{total} 个）")
    for idx, note_filename in enumerate(sorted(delivery_notes), start=1):
        base = 10 + int((idx - 1) / max(total, 1) * 80)
        report(base, f"匹配模板并回写：{idx}/{total} {note_filename}")
        note_path = os.path.join(output_dir, note_filename)

        report(min(base + 3, 92), "为送货单选择最相近模板…")
        best_template_match = {"score": -1, "path": None}
        for template_path in all_template_files:
            similarity = get_text_similarity(note_filename, os.path.basename(template_path))
            if similarity > best_template_match["score"]:
                best_template_match.update({"score": similarity, "path": template_path})
        if not best_template_match["path"]:
            continue

        report(min(base + 6, 92), "读取送货单与模板…")
        template_path = best_template_match["path"]
        wb_note = openpyxl.load_workbook(note_path)
        ws_note = wb_note["箱单"]

        report(min(base + 8, 92), "定位收货地址列并切块…")
        address_column_letter = find_address_column(ws_note)
        if not address_column_letter:
            wb_note.close()
            continue

        blocks = find_data_blocks(ws_note, address_column_letter)
        if not blocks:
            wb_note.close()
            continue

        report(min(base + 10, 92), f"找到 {len(blocks)} 个地址块，开始填充…")
        wb_template_for_copy = openpyxl.load_workbook(template_path)
        source_header_map = {normalize_header(c.value): c.column for r in [1, 2] for c in ws_note[r] if c.value}

        for block in blocks:
            report(min(base + 12, 93), f"匹配模板sheet并复制：{block['key_text']}")
            best_sheet_match = {"score": -1, "sheet_name": None}
            for sheet_name in wb_template_for_copy.sheetnames:
                row7_text = "".join([str(c.value or "") for c in wb_template_for_copy[sheet_name][7]])
                similarity = get_text_similarity(block["key_text"], row7_text)
                if similarity > best_sheet_match["score"]:
                    best_sheet_match.update({"score": similarity, "sheet_name": sheet_name})

            sheet_name = best_sheet_match["sheet_name"]
            if not sheet_name:
                continue

            new_sheet_title = sheet_name
            counter = 2
            while new_sheet_title in wb_note.sheetnames:
                new_sheet_title = f"{sheet_name} ({counter})"
                counter += 1

            ws_dest = clone_sheet(wb_template_for_copy[sheet_name], wb_note, new_sheet_title)

            dest_header_map = {normalize_header(c.value): c.column for c in ws_dest[8] if c.value}
            transfer_map = {}
            for dest_norm_header, dest_col in dest_header_map.items():
                if dest_norm_header == normalize_header("进价"):
                    continue
                for source_norm_header, source_col in source_header_map.items():
                    if dest_norm_header == source_norm_header:
                        transfer_map[dest_col] = {"source_col": source_col, "header": dest_norm_header}
                        break

            for r_offset in range(block["end"] - block["start"] + 1):
                source_row, dest_row = block["start"] + r_offset, 9 + r_offset
                for dest_col, transfer_info in transfer_map.items():
                    if transfer_info["header"] in ["mayang", "shiyang"]:
                        continue
                    source_col = transfer_info["source_col"]
                    source_cell = ws_note.cell(row=source_row, column=source_col)
                    target_cell = ws_dest.cell(row=dest_row, column=dest_col)

                    write_cell = target_cell
                    if isinstance(target_cell, openpyxl.cell.cell.MergedCell):
                        for merged_range in ws_dest.merged_cells.ranges:
                            if target_cell.coordinate in merged_range:
                                write_cell = ws_dest.cell(row=merged_range.min_row, column=merged_range.min_col)
                                break

                    write_cell.value = source_cell.value
                    if transfer_info["header"] == "discount":
                        write_cell.number_format = copy(source_cell.number_format)

            dest_norm_map = {normalize_header(c.value): c.column_letter for c in ws_dest[8] if c.value}
            qty_col, price_col, discount_col, mayang_col, shiyang_col = (
                dest_norm_map.get("quantity"),
                dest_norm_map.get("price"),
                dest_norm_map.get("discount"),
                dest_norm_map.get("mayang"),
                dest_norm_map.get("shiyang"),
            )

            if all([qty_col, price_col, mayang_col]):
                for r_offset in range(block["end"] - block["start"] + 1):
                    dest_row = 9 + r_offset
                    ws_dest[f"{mayang_col}{dest_row}"] = f"={qty_col}{dest_row}*{price_col}{dest_row}"
                    if all([shiyang_col, discount_col]):
                        ws_dest[f"{shiyang_col}{dest_row}"] = f"={mayang_col}{dest_row}*{discount_col}{dest_row}"

            last_filled_row = 9 + (block["end"] - block["start"])
            total_row_finder = -1
            for row in range(last_filled_row + 1, ws_dest.max_row + 1):
                cell_value = ws_dest[f"B{row}"].value
                if cell_value and "总计" in str(cell_value):
                    total_row_finder = row
                    break

            if total_row_finder != -1:
                if total_row_finder > last_filled_row + 1:
                    ws_dest.delete_rows(last_filled_row + 1, total_row_finder - (last_filled_row + 1))

                new_total_row = last_filled_row + 1
                for cell in ws_dest[new_total_row]:
                    if isinstance(cell.value, str) and "SUM" in cell.value.upper():
                        col_letter = cell.column_letter
                        cell.value = f"=SUM({col_letter}9:{col_letter}{last_filled_row})"

        wb_template_for_copy.close()
        wb_note.save(note_path)
        outputs.append(note_path)

        report(min(base + 20, 95), f"已完成：{note_filename}")

    report(100, "完成")
    return outputs


