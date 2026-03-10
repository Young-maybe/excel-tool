"""
猴面包树B2B发货 - 提货单生成（记得填sdo和箱数）

严格保持原脚本的核心处理逻辑，仅做适配：
- 模板文件：在 input_dir/AAA提货单模板系统专用.xlsx
- 待处理送货单：从 output_dir 查找（因为步骤1输出统一到 output_dir）
- 输出：提货单写入 output_dir
"""
from __future__ import annotations

import os
import re
import shutil
from copy import copy
from typing import Callable, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment


def normalize_header(text):
    """清洗并标准化表头文本，用于模糊匹配。"""
    if not text:
        return ""
    cleaned_text = re.sub(r"\s+", "", str(text)).lower()
    alias_map = {
        "sdo单号": "sdo",
        "sdo": "sdo",
        "送货时间": "delivery_time",
        "预约时间": "delivery_time",
    }
    return alias_map.get(cleaned_text, cleaned_text)


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> List[str]:
    def report(p: int, s: str) -> None:
        if progress_cb:
            progress_cb(p, s)

    TEMPLATE_FILENAME = "AAA提货单模板系统专用.xlsx"
    DELIVERY_NOTE_SUFFIX = "送货单.xlsx"
    PICKING_SLIP_SUFFIX = "提货单.xlsx"

    template_path = os.path.join(input_dir, TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到模板文件：{template_path}")

    report(10, f"读取提货单模板：{TEMPLATE_FILENAME}")
    delivery_notes = [f for f in os.listdir(output_dir) if f.endswith(DELIVERY_NOTE_SUFFIX)]
    if not delivery_notes:
        raise FileNotFoundError("未在输出目录中找到任何“送货单.xlsx”。请先运行“送货单、备货单生成”。")

    outputs: List[str] = []
    total = len(delivery_notes)
    report(5, f"扫描送货单…（{total} 个）")
    for idx, delivery_note_filename in enumerate(sorted(delivery_notes), start=1):
        base = 10 + int((idx - 1) / max(total, 1) * 80)
        report(base, f"生成提货单：{idx}/{total} {delivery_note_filename}")
        delivery_note_path = os.path.join(output_dir, delivery_note_filename)

        picking_slip_filename = delivery_note_filename.replace(DELIVERY_NOTE_SUFFIX, PICKING_SLIP_SUFFIX)
        picking_slip_path = os.path.join(output_dir, picking_slip_filename)
        report(min(base + 3, 92), "复制模板生成提货单副本…")
        shutil.copy(template_path, picking_slip_path)

        report(min(base + 6, 92), "读取送货单/提货单工作簿…")
        wb_delivery = openpyxl.load_workbook(delivery_note_path)
        ws_delivery = wb_delivery.active
        wb_picking = openpyxl.load_workbook(picking_slip_path)
        ws_picking = wb_picking.active

        report(min(base + 9, 92), "识别表头并建立列映射…")
        delivery_header_map: Dict[str, int] = {}
        for row in [1, 2]:
            for cell in ws_delivery[row]:
                if cell.value:
                    normalized = normalize_header(cell.value)
                    if normalized:
                        delivery_header_map[normalized] = cell.column

        col_mapping: Dict[int, int] = {}
        for row_num in [1, 2]:
            for header_cell in ws_picking[row_num]:
                if not header_cell.value:
                    continue
                normalized = normalize_header(header_cell.value)
                if not normalized:
                    continue

                source_col_idx = delivery_header_map.get(normalized)
                if source_col_idx:
                    target_col_idx = header_cell.column
                    col_mapping[source_col_idx] = target_col_idx
                    report(min(base + 14, 95), "回填数据与样式…")
                    for row_idx in range(3, ws_delivery.max_row + 1):
                        source_cell = ws_delivery.cell(row=row_idx, column=source_col_idx)
                        target_cell = ws_picking.cell(row=row_idx, column=target_col_idx)
                        target_cell.value = source_cell.value
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.fill, target_cell.number_format, target_cell.protection, target_cell.alignment = (
                                copy(source_cell.font),
                                copy(source_cell.border),
                                copy(source_cell.fill),
                                source_cell.number_format,
                                copy(source_cell.protection),
                                copy(source_cell.alignment),
                            )

        report(min(base + 16, 96), "处理合并单元格 & 合计区域…")
        for merged_range in ws_delivery.merged_cells.ranges:
            if merged_range.min_col in col_mapping:
                target_col = col_mapping[merged_range.min_col]
                if merged_range.min_col == merged_range.max_col:
                    start_row, end_row = merged_range.min_row, merged_range.max_row
                    ws_picking.merge_cells(start_row=start_row, start_column=target_col, end_row=end_row, end_column=target_col)

        last_data_row = ws_delivery.max_row

        # 提取合计区域信息（4行高）
        total_area_info = []
        template_start_row, template_end_row = 111, 114
        for r in range(template_start_row, template_end_row + 1):
            row_info = [copy(ws_picking.cell(row=r, column=c)) for c in range(1, ws_picking.max_column + 1)]
            total_area_info.append(row_info)

        if last_data_row + 1 <= ws_picking.max_row:
            ws_picking.delete_rows(last_data_row + 1, ws_picking.max_row - last_data_row)

        new_total_start_row = last_data_row + 1
        for r_offset, row_data in enumerate(total_area_info):
            current_row = new_total_start_row + r_offset
            for c_offset, source_cell in enumerate(row_data):
                new_cell = ws_picking.cell(row=current_row, column=c_offset + 1)
                new_cell.value = source_cell.value
                if source_cell.has_style:
                    new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.protection, new_cell.alignment = (
                        copy(source_cell.font),
                        copy(source_cell.border),
                        copy(source_cell.fill),
                        source_cell.number_format,
                        copy(source_cell.protection),
                        copy(source_cell.alignment),
                    )

        formula_col_letter = "C"
        formula_col_idx = openpyxl.utils.column_index_from_string(formula_col_letter)
        formula_cell = ws_picking.cell(row=new_total_start_row, column=formula_col_idx)
        formula_cell.value = f"=SUM(C3:C{last_data_row})"

        start_row, end_row = new_total_start_row, new_total_start_row + 3
        ws_picking.merge_cells(start_row=start_row, start_column=formula_col_idx, end_row=end_row, end_column=formula_col_idx)
        formula_cell.alignment = Alignment(horizontal="center", vertical="center")

        report(min(base + 18, 98), "保存提货单…")
        wb_picking.save(picking_slip_path)
        outputs.append(picking_slip_path)

        report(min(base + 20, 95), f"已完成：{picking_slip_filename}")

    report(100, "完成")
    return outputs


