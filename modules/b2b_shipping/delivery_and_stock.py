"""
猴面包树B2B发货 - 送货单、备货单生成

严格保持原脚本的核心处理逻辑，仅做适配：
- 输入：由 UI 传入 input_dir（原脚本为当前目录）
- 输出：统一写入 output_dir（原脚本写回当前目录）
"""
from __future__ import annotations

import os
import re
import shutil
import sys
from copy import copy
from typing import Callable, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange


def convert_xls_to_xlsx(xls_path: str, xlsx_path: str) -> bool:
    """使用Windows COM接口将.xls文件转换为.xlsx文件。"""
    if sys.platform != "win32":
        return False
    excel = None
    try:
        import win32com.client as win32  # type: ignore

        xls_path_abs, xlsx_path_abs = os.path.abspath(xls_path), os.path.abspath(xlsx_path)
        excel = win32.Dispatch("Excel.Application")
        excel.Visible, excel.DisplayAlerts = False, False
        workbook = excel.Workbooks.Open(xls_path_abs)
        workbook.SaveAs(xlsx_path_abs, FileFormat=51)
        workbook.Close()
        return True
    except Exception:
        return False
    finally:
        if excel:
            excel.Quit()


def generate_base_file(filepath: str, columns_to_hide_ranges, move_spec) -> None:
    """
    模块a+b: 负责生成包含所有列和格式的基础文件。
    （核心处理逻辑保持与原脚本一致）
    """
    source_wb = openpyxl.load_workbook(filepath)
    source_ws = source_wb.active
    sheet_data = [[copy(cell) for cell in row] for row in source_ws.iter_rows()]
    merged_cells = [m.coord for m in source_ws.merged_cells.ranges]
    row_heights = {i: dim.height for i, dim in source_ws.row_dimensions.items()}
    col_widths = [source_ws.column_dimensions[get_column_letter(i)].width for i in range(1, source_ws.max_column + 1)]

    source_idx = column_index_from_string(move_spec["source"]) - 1
    dest_idx = column_index_from_string(move_spec["dest"]) - 1
    for row in sheet_data:
        if len(row) > source_idx:
            moved_cell = row.pop(source_idx)
            row.insert(dest_idx, moved_cell)
    if len(col_widths) > source_idx:
        moved_width = col_widths.pop(source_idx)
        col_widths.insert(dest_idx, moved_width)

    def transform_col_index(c_idx: int) -> int:
        if c_idx < dest_idx:
            return c_idx
        if c_idx == source_idx:
            return dest_idx
        if dest_idx <= c_idx < source_idx:
            return c_idx + 1
        if c_idx > source_idx:
            return c_idx - 1
        return c_idx

    new_merged_cells = []
    for mc_range_str in merged_cells:
        mc = CellRange(mc_range_str)
        mc.min_col, mc.max_col = transform_col_index(mc.min_col - 1) + 1, transform_col_index(mc.max_col - 1) + 1
        new_merged_cells.append(mc.coord)

    target_wb = openpyxl.Workbook()
    target_ws = target_wb.active
    for r_idx, row_data in enumerate(sheet_data, 1):
        for c_idx, source_cell in enumerate(row_data, 1):
            target_cell = target_ws.cell(row=r_idx, column=c_idx)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font, target_cell.border, target_cell.fill, target_cell.number_format, target_cell.protection, target_cell.alignment = (
                    copy(source_cell.font),
                    copy(source_cell.border),
                    copy(source_cell.fill),
                    source_cell.number_format,
                    copy(source_cell.protection),
                    copy(source_cell.alignment),
                )
    for r, h in row_heights.items():
        if h is not None:
            target_ws.row_dimensions[r].height = h
    for c_idx, width in enumerate(col_widths, 1):
        if width is not None:
            target_ws.column_dimensions[get_column_letter(c_idx)].width = width
    for mc_range in new_merged_cells:
        target_ws.merge_cells(mc_range)

    last_row = target_ws.max_row
    for row in range(3, last_row + 1):
        if target_ws[f"Q{row}"].value:
            target_ws[f"AR{row}"].value = f"=AQ{row}*S{row}"

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    headers_to_add = [("AU", "sdo单号"), ("AV", "物流单号"), ("AW", "预约时间"), ("AX", "规格"), ("AY", "总箱数")]
    for col_letter, header_text in headers_to_add:
        cell = target_ws[f"{col_letter}2"]
        cell.value, cell.border, cell.alignment = header_text, thin_border, center_align

    start_row = 0
    for row in range(3, target_ws.max_row + 2):
        is_data_row = target_ws[f"Q{row}"].value is not None
        if is_data_row and start_row == 0:
            start_row = row
        elif not is_data_row and start_row != 0:
            end_row = row - 1
            data_height = end_row - start_row + 1
            cols_to_border = ["AU", "AV", "AW", "AX", "AY"]
            for r in range(start_row, end_row + 1):
                for c_letter in cols_to_border:
                    target_ws[f"{c_letter}{r}"].border = thin_border
            if data_height > 1:
                cols_to_merge = ["AU", "AV", "AW", "AY"]
                for col_letter in cols_to_merge:
                    target_ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                    target_ws[f"{col_letter}{start_row}"].alignment = center_align
            start_row = 0

    initial_hide_indices = {
        i
        for r in columns_to_hide_ranges
        for i in (range(column_index_from_string(r.split("-")[0]), column_index_from_string(r.split("-")[1]) + 1) if "-" in r else [column_index_from_string(r)])
    }
    final_hide_indices = {transform_col_index(i - 1) + 1 for i in initial_hide_indices}
    final_hide_indices.add(source_idx + 1)
    for col_idx in final_hide_indices:
        target_ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    target_wb.save(filepath)


def process_folder(
    input_dir: str,
    output_dir: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> List[str]:
    """从 input_dir 读取源文件，输出送货单/备货单到 output_dir。"""
    def report(p: int, s: str) -> None:
        if progress_cb:
            progress_cb(p, s)

    os.makedirs(output_dir, exist_ok=True)

    COLUMNS_TO_HIDE_RANGES = ["A", "D-P", "S-Z", "AB", "AH-AO"]
    COLUMN_MOVE_SPEC = {"source": "AP", "dest": "R"}

    pattern = re.compile(r"^易快报导单--猴面包树-(.*?)(\d{8}).*?\.(xlsx?)$", re.IGNORECASE)
    outputs: List[str] = []

    report(2, "扫描输入目录并匹配易快报导单命名规则…")
    candidates = [fn for fn in os.listdir(input_dir) if pattern.search(fn)]
    total = len(candidates)
    report(5, f"扫描输入文件…（匹配到 {total} 个）")

    for idx, filename in enumerate(sorted(candidates), start=1):
        base = 10 + int((idx - 1) / max(total, 1) * 80)
        report(base, f"处理 {idx}/{total}：{filename}")
        match = pattern.search(filename)
        if not match:
            continue
        company_part, date_part, extension = match.groups()
        company_name = company_part.strip("-_ \t()（）")
        if not company_name:
            continue

        base_name = f"{company_name}{date_part[4:6]}{date_part[6:8]}"
        delivery_note_filename = f"{base_name}送货单.xlsx"
        stocking_list_filename = f"{base_name}备货单.xlsx"
        delivery_note_filepath = os.path.join(output_dir, delivery_note_filename)
        stocking_list_filepath = os.path.join(output_dir, stocking_list_filename)

        source_path = os.path.join(input_dir, filename)
        temp_base_path = os.path.join(output_dir, f"temp_{base_name}.xlsx")

        if extension.lower() == "xls":
            report(min(base + 5, 90), "检测到 .xls，调用 Excel 转换为 .xlsx…")
            ok = convert_xls_to_xlsx(source_path, temp_base_path)
            if not ok:
                raise FileNotFoundError("检测到 .xls 但无法通过 Excel 转换，请确认已安装 Excel 且已安装 pywin32。")
        else:
            report(min(base + 5, 90), "复制源文件到工作副本…")
            shutil.copy(source_path, temp_base_path)

        report(min(base + 12, 92), "生成基础文件（列移动/隐藏/公式等）…")
        generate_base_file(temp_base_path, COLUMNS_TO_HIDE_RANGES, COLUMN_MOVE_SPEC)

        # 派生备货单
        report(min(base + 18, 94), "生成备货单…")
        shutil.copy(temp_base_path, stocking_list_filepath)
        report(min(base + 20, 95), "备货单：隐藏列并保存…")
        wb_stock = openpyxl.load_workbook(stocking_list_filepath)
        ws_stock = wb_stock.active
        for col_idx in range(column_index_from_string("AB"), column_index_from_string("AW") + 1):
            ws_stock.column_dimensions[get_column_letter(col_idx)].hidden = True
        wb_stock.save(stocking_list_filepath)

        # 派生送货单
        report(min(base + 24, 96), "生成送货单…")
        shutil.copy(temp_base_path, delivery_note_filepath)
        report(min(base + 26, 97), "送货单：删除列并保存…")
        wb_delivery = openpyxl.load_workbook(delivery_note_filepath)
        ws_delivery = wb_delivery.active
        ws_delivery.delete_cols(column_index_from_string("AY"))
        ws_delivery.delete_cols(column_index_from_string("AX"))
        wb_delivery.save(delivery_note_filepath)

        # 重命名 sheet 为 箱单
        report(min(base + 28, 98), "送货单：重命名sheet为“箱单”…")
        wb_delivery_final = openpyxl.load_workbook(delivery_note_filepath)
        ws_delivery_final = wb_delivery_final.active
        ws_delivery_final.title = "箱单"
        wb_delivery_final.save(delivery_note_filepath)

        if os.path.exists(temp_base_path):
            os.remove(temp_base_path)

        outputs.extend([delivery_note_filepath, stocking_list_filepath])

    if not outputs:
        raise FileNotFoundError("未在输入文件夹中找到符合命名规则的易快报导单文件。")

    report(100, "完成")
    return outputs


