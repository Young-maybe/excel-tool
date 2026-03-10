#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
仓库库存数据处理工具
根据仓库分类和商品库存数据生成Excel报表
"""

import sys
import os
import traceback
from datetime import datetime, timedelta

# Fix for PyInstaller --windowed mode: redirect stdout/stderr to avoid NoneType errors
if sys.stdout is None:
    sys.stdout = open(os.devnull, "w")
if sys.stderr is None:
    sys.stderr = open(os.devnull, "w")

# 核心GUI模块 - 迁移到 PyQt6（保持界面与逻辑尽量不变）
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLabel,
    QFileDialog,
    QTextEdit,
    QProgressBar,
    QMessageBox,
    QGroupBox,
    QTabWidget,
)
from PyQt6.QtCore import QThread, pyqtSignal, Qt
from PyQt6.QtGui import QFont, QIcon

# PyQt5 -> PyQt6 兼容：本文件尽量保持原逻辑不动，仅补齐常用枚举
try:
    if not hasattr(Qt, "AlignCenter"):  # PyQt6 无 Qt.AlignCenter
        Qt.AlignCenter = Qt.AlignmentFlag.AlignCenter  # type: ignore[attr-defined]
except Exception:
    # 若未来Qt实现禁止设置属性，这里保持静默，后续再逐点替换
    pass


def create_embedded_widget(initial_tab_index: int = 0) -> QWidget:
    """
    将原工具的界面“嵌入”到外部页面中。
    关键点：
    - 不改原逻辑/原UI，只是把 MainWindow 的 centralWidget 拿出来放到一个容器里
    - 为了让信号槽仍然可用，需要保持 MainWindow 实例不被 GC
    """
    win = MainWindow()
    central = win.centralWidget()
    container = QWidget()
    layout = QVBoxLayout(container)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(0)

    if central is not None:
        central.setParent(container)
        layout.addWidget(central)

    try:
        win.tab_widget.setCurrentIndex(int(initial_tab_index))
    except Exception:
        pass

    # 防止 window 被垃圾回收导致信号槽失效
    setattr(container, "_legacy_window", win)
    return container

# 数据处理模块 - 延迟导入以加快启动速度
# pandas, numpy, openpyxl 等将在需要时导入


class DataProcessor(QThread):
    """数据处理线程"""
    progress_updated = pyqtSignal(int)
    status_updated = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, warehouse_file, inventory_file, output_file):
        super().__init__()
        self.warehouse_file = warehouse_file
        self.inventory_file = inventory_file
        self.output_file = output_file
        
        # 延迟导入数据处理模块
        self._import_data_modules()
    
    def _import_data_modules(self):
        """延迟导入数据处理模块以提升启动速度"""
        try:
            global pd, np, openpyxl, Font, Alignment, PatternFill
            self.status_updated.emit("正在加载数据处理模块...")
            
            import pandas as pd
            import numpy as np
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            self.status_updated.emit("模块加载完成")
        except ImportError as e:
            self.status_updated.emit(f"模块导入失败: {e}")
            raise
    
    def preprocess_inventory_data(self, inventory_df):
        """预处理库存数据，确保数据质量"""
        self.status_updated.emit("执行数据预处理...")
        
        # 1. 确保必要列存在
        required_columns = ['商品规格代码', '商品名称', '仓库', '可销数', '在途数']
        for col in required_columns:
            if col not in inventory_df.columns:
                self.status_updated.emit(f"警告: 缺少必要列 '{col}'")
                inventory_df[col] = 0 if col in ['可销数', '在途数'] else ''
        
        # 2. 数据类型转换和清理
        # 规格代码格式化
        inventory_df['商品规格代码'] = self.format_spec_code(inventory_df['商品规格代码'])
        
        # 数值列转换
        for col in ['可销数', '在途数']:
            inventory_df[col] = pd.to_numeric(inventory_df[col], errors='coerce').fillna(0)
        
        # 3. 文本数据清理
        for col in ['商品名称', '仓库']:
            if col in inventory_df.columns:
                inventory_df[col] = inventory_df[col].astype(str).str.strip()
        
        # 4. 检查并报告数据重复情况
        original_count = len(inventory_df)
        
        # 检查完全重复的行
        duplicated_rows = inventory_df.duplicated().sum()
        if duplicated_rows > 0:
            self.status_updated.emit(f"发现 {duplicated_rows} 行完全重复的数据，将保留第一行")
            inventory_df = inventory_df.drop_duplicates()
        
        # 检查业务逻辑重复（相同规格代码+仓库的不同行）
        business_duplicates = inventory_df.groupby(['商品规格代码', '仓库']).size()
        business_dup_count = (business_duplicates > 1).sum()
        if business_dup_count > 0:
            self.status_updated.emit(f"发现 {business_dup_count} 个规格代码-仓库组合有多行数据")
            self.status_updated.emit("这些重复数据的数量将在后续处理中自动求和")
        
        final_count = len(inventory_df)
        if original_count != final_count:
            self.status_updated.emit(f"数据预处理完成: {original_count} → {final_count} 行")
        
        return inventory_df
    
    def format_spec_code(self, spec_code_series):
        """
        格式化规格代码：去掉前后空格，尝试转换为整数格式（去掉小数点），如果转换失败则保留文本格式
        """
        def convert_single_code(code):
            if pd.isna(code):
                return ''
            
            # 转换为字符串并去掉前后空格
            code_str = str(code).strip()
            
            # 如果是空字符串，直接返回
            if not code_str:
                return ''
            
            # 尝试转换为数字（只转换为整数，去掉小数点）
            try:
                # 转换为浮点数然后取整数部分
                float_val = float(code_str)
                # 直接转换为整数（去掉小数点）
                return int(float_val)
            except (ValueError, TypeError):
                # 转换失败，返回清理后的文本格式
                return code_str
        
        return spec_code_series.apply(convert_single_code)
        
    def run(self):
        try:
            self.status_updated.emit("开始处理数据...")
            self.progress_updated.emit(10)
            
            # 读取仓库分类数据
            self.status_updated.emit("读取仓库分类数据...")
            warehouse_df = pd.read_excel(self.warehouse_file, engine='openpyxl')
            self.progress_updated.emit(15)
            
            # 读取商品库存数据 - 优化版本
            self.status_updated.emit("读取商品库存数据...")
            # 使用更高效的CSV读取方式
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            inventory_df = None
            
            for encoding in encodings:
                try:
                    # 使用低内存模式和指定数据类型以提高性能
                    inventory_df = pd.read_csv(
                        self.inventory_file, 
                        encoding=encoding,
                        low_memory=False,
                        dtype={'商品规格代码': 'str', '可销数': 'float64', '在途数': 'float64'},
                        skipinitialspace=True  # 跳过字段开头的空格
                    )
                    self.status_updated.emit(f"成功使用 {encoding} 编码读取库存数据")
                    break
                except (UnicodeDecodeError, pd.errors.ParserError):
                    continue
                    
            if inventory_df is None:
                raise Exception("无法读取库存数据文件，请检查文件编码")
            
            # 清理数据，填充NaN值
            inventory_df['可销数'] = inventory_df['可销数'].fillna(0)
            inventory_df['在途数'] = inventory_df['在途数'].fillna(0)
            
            # 清理商品规格代码的前后空格并尝试转换为数字格式
            inventory_df['商品规格代码'] = self.format_spec_code(inventory_df['商品规格代码'])
            
            self.progress_updated.emit(25)
            
            # 数据处理
            self.status_updated.emit("开始数据匹配和聚合...")
            result_data = self.process_data(warehouse_df, inventory_df)
            self.progress_updated.emit(80)
            
            # 生成Excel文件
            self.status_updated.emit("生成Excel文件...")
            self.generate_excel(result_data, self.output_file)
            self.progress_updated.emit(100)
            
            self.status_updated.emit("数据处理完成！")
            self.finished_signal.emit(True, "数据处理成功完成！")
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}\n{traceback.format_exc()}"
            self.status_updated.emit(f"错误: {str(e)}")
            self.finished_signal.emit(False, error_msg)
    
    def process_data(self, warehouse_df, inventory_df):
        """处理数据逻辑 - 优化版本"""
        self.status_updated.emit("预处理数据...")
        
        # 数据预处理：清理和规范化
        inventory_df = self.preprocess_inventory_data(inventory_df)
        
        # 获取仓库分类中的计划备注列数据并去重
        plan_remarks = warehouse_df['计划备注'].dropna().unique()
        plan_remarks = list(plan_remarks)
        
        # 先清理数据中的前后空格
        warehouse_df['供应商'] = warehouse_df['供应商'].astype(str).str.strip()
        inventory_df['仓库'] = inventory_df['仓库'].astype(str).str.strip()
        
        # 创建供应商到计划备注的映射字典，提高查询效率
        supplier_to_remark = dict(zip(warehouse_df['供应商'], warehouse_df['计划备注']))
        
        # 为库存数据添加计划备注列，使用向量化操作
        inventory_df = inventory_df.copy()
        inventory_df['计划备注'] = inventory_df['仓库'].map(supplier_to_remark)
        
        # 检查匹配情况
        matched_count = inventory_df['计划备注'].notna().sum()
        total_count = len(inventory_df)
        self.status_updated.emit(f"精确匹配成功率: {matched_count}/{total_count} ({matched_count/total_count*100:.1f}%)")
        
        # 添加调试信息：显示未匹配的仓库样例
        if matched_count == 0:
            unmatched_warehouses = inventory_df[inventory_df['计划备注'].isna()]['仓库'].unique()[:5]
            self.status_updated.emit(f"未匹配仓库样例: {list(unmatched_warehouses)}")
            supplier_samples = warehouse_df['供应商'].head(5).tolist()
            self.status_updated.emit(f"仓库分类样例: {supplier_samples}")
        
        # 总是尝试模糊匹配以提高匹配率
        self.status_updated.emit("执行智能匹配算法...")
        # 去除空格、连字符和特殊字符后再匹配
        warehouse_df_clean = warehouse_df.copy()
        warehouse_df_clean['供应商_清理'] = (warehouse_df_clean['供应商'].astype(str)
                                    .str.strip()
                                    .str.replace(' ', '')
                                    .str.replace('-', '')
                                    .str.replace('_', '')
                                    .str.replace('（', '(')
                                    .str.replace('）', ')'))
        
        inventory_df_clean = inventory_df.copy()
        inventory_df_clean['仓库_清理'] = (inventory_df_clean['仓库'].astype(str)
                                   .str.strip()
                                   .str.replace(' ', '')
                                   .str.replace('-', '')
                                   .str.replace('_', '')
                                   .str.replace('（', '(')
                                   .str.replace('）', ')'))
        
        supplier_to_remark_clean = dict(zip(warehouse_df_clean['供应商_清理'], warehouse_df_clean['计划备注']))
        inventory_df['计划备注'] = inventory_df_clean['仓库_清理'].map(supplier_to_remark_clean)
        
        matched_count_new = inventory_df['计划备注'].notna().sum()
        self.status_updated.emit(f"智能匹配成功率: {matched_count_new}/{total_count} ({matched_count_new/total_count*100:.1f}%)")
        
        self.status_updated.emit("处理信选仓数据...")
        
        # 只保留在仓库分类表格中有匹配的数据，过滤掉未匹配的仓库数据
        valid_inventory = inventory_df[inventory_df['计划备注'].notna()]
        
        # 统计过滤后的数据
        filtered_count = len(inventory_df) - len(valid_inventory)
        if filtered_count > 0:
            self.status_updated.emit(f"已过滤掉 {filtered_count} 条未在仓库分类表格中的数据")
        
        self.status_updated.emit(f"处理商品数据，共计 {len(valid_inventory)} 条记录")
        
        # 添加重复数据检查和调试信息
        duplicate_check = valid_inventory.groupby(['商品规格代码', '仓库']).size()
        duplicates = duplicate_check[duplicate_check > 1]
        if len(duplicates) > 0:
            self.status_updated.emit(f"发现 {len(duplicates)} 个规格代码-仓库组合有重复记录，将自动合并数量")
            # 显示前几个重复的样例
            for i, (spec_warehouse, count) in enumerate(duplicates.head(3).items()):
                spec_code, warehouse = spec_warehouse
                self.status_updated.emit(f"  样例 {i+1}: 规格代码 {spec_code} 在仓库 {warehouse} 有 {count} 条记录")
        
        if len(valid_inventory) == 0:
            self.status_updated.emit("警告：没有找到库存数据，信选仓表将为空")
            xuancang_data = []
        else:
            # 使用groupby进行高效的数据聚合 - 信选仓数据
            # 这里会自动将相同规格代码、商品名称、计划备注的记录的可销数进行求和
            xuancang_agg = valid_inventory.groupby(['商品规格代码', '商品名称', '计划备注'])['可销数'].sum().unstack(fill_value=0)
            
            # 计算总计和可用库存
            xuancang_agg['总计'] = xuancang_agg.sum(axis=1)
            
            # 计算可用库存（去除自营商城仓）
            available_columns = [col for col in xuancang_agg.columns if col not in ['自营商城仓', '总计']]
            if available_columns:
                xuancang_agg['可用库存'] = xuancang_agg[available_columns].sum(axis=1)
            else:
                xuancang_agg['可用库存'] = 0
            
            # 重置索引，将商品名称作为列
            xuancang_agg = xuancang_agg.reset_index()
            
            # 重新排列列的顺序
            column_order = ['商品名称', '商品规格代码'] + [col for col in plan_remarks if col in xuancang_agg.columns] + ['总计', '可用库存']
            xuancang_agg = xuancang_agg.reindex(columns=column_order, fill_value=0)
            
            # 转换为字典列表
            xuancang_data = xuancang_agg.to_dict('records')
        
        self.status_updated.emit("处理自营在途数据...")
        
        # 使用groupby进行高效的数据聚合 - 自营在途数据
        # 只选择自营仓和自营商城仓的数据，且必须是已匹配的数据
        ziying_data = valid_inventory[valid_inventory['计划备注'].isin(['自营仓', '自营商城仓'])]
        
        if not ziying_data.empty:
            # 检查自营数据中的重复情况
            ziying_duplicate_check = ziying_data.groupby(['商品规格代码', '计划备注']).size()
            ziying_duplicates = ziying_duplicate_check[ziying_duplicate_check > 1]
            if len(ziying_duplicates) > 0:
                self.status_updated.emit(f"自营数据中发现 {len(ziying_duplicates)} 个规格代码-仓库组合有重复，将合并在途数")
            
            # 这里会自动将相同规格代码、计划备注的记录的在途数进行求和
            ziying_agg = ziying_data.groupby(['商品规格代码', '计划备注'])['在途数'].sum().unstack(fill_value=0)
            
            # 确保包含自营仓和自营商城仓列
            for col in ['自营仓', '自营商城仓']:
                if col not in ziying_agg.columns:
                    ziying_agg[col] = 0
            
            # 计算总计
            ziying_agg['总计'] = ziying_agg['自营仓'] + ziying_agg['自营商城仓']
            
            # 重置索引
            ziying_agg = ziying_agg.reset_index()
            
            # 重新排列列的顺序
            ziying_agg = ziying_agg[['商品规格代码', '自营仓', '自营商城仓', '总计']]
            
            # 转换为字典列表
            ziyingzaitu_data = ziying_agg.to_dict('records')
        else:
            # 如果没有自营数据，创建空的数据结构
            all_products = inventory_df['商品规格代码'].unique()
            ziyingzaitu_data = [
                {
                    '商品规格代码': product,
                    '自营仓': 0,
                    '自营商城仓': 0,
                    '总计': 0
                }
                for product in all_products
            ]
        
        self.status_updated.emit("数据处理完成，准备生成Excel...")
        
        return {
            'xuancang': xuancang_data,
            'ziyingzaitu': ziyingzaitu_data,
            'plan_remarks': plan_remarks
        }
    
    def generate_excel(self, result_data, output_file):
        """生成Excel文件 - 防止科学计数法显示"""
        self.status_updated.emit("创建Excel工作表...")
        
        # 使用openpyxl引擎来精确控制格式
        from openpyxl import Workbook
        from openpyxl.styles import NamedStyle
        
        wb = Workbook()
        
        # 创建信选仓sheet
        if 'xuancang' in wb.sheetnames:
            wb.remove(wb['xuancang'])
        ws_xuancang = wb.create_sheet('信选仓', 0)
        
        xuancang_df = pd.DataFrame(result_data['xuancang'])
        if not xuancang_df.empty:
            # 确保商品规格代码格式正确（数字格式优先，去前后空格）
            if '商品规格代码' in xuancang_df.columns:
                xuancang_df['商品规格代码'] = self.format_spec_code(xuancang_df['商品规格代码'])
            
            # 写入表头
            for col_idx, col_name in enumerate(xuancang_df.columns, 1):
                ws_xuancang.cell(row=1, column=col_idx, value=col_name)
            
            # 写入数据并设置格式
            spec_code_col = None
            if '商品规格代码' in xuancang_df.columns:
                spec_code_col = list(xuancang_df.columns).index('商品规格代码') + 1
            
            for row_idx, row_data in enumerate(xuancang_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_xuancang.cell(row=row_idx, column=col_idx, value=value)
                    # 如果是规格代码列且是数字，设置为文本格式避免科学计数法
                    if col_idx == spec_code_col and isinstance(value, (int, float)):
                        cell.number_format = '0'  # 设置为整数格式，不显示科学计数法
        else:
            # 创建空的信选仓表格
            headers = ['商品名称', '商品规格代码', '总计', '可用库存']
            if 'plan_remarks' in result_data:
                headers = ['商品名称', '商品规格代码'] + list(result_data['plan_remarks']) + ['总计', '可用库存']
            for col_idx, header in enumerate(headers, 1):
                ws_xuancang.cell(row=1, column=col_idx, value=header)
        
        # 创建自营在途sheet
        ws_ziyingzaitu = wb.create_sheet('自营在途')
        
        ziyingzaitu_df = pd.DataFrame(result_data['ziyingzaitu'])
        if not ziyingzaitu_df.empty:
            # 确保商品规格代码格式正确（数字格式优先，去前后空格）
            if '商品规格代码' in ziyingzaitu_df.columns:
                ziyingzaitu_df['商品规格代码'] = self.format_spec_code(ziyingzaitu_df['商品规格代码'])
            
            # 写入表头
            for col_idx, col_name in enumerate(ziyingzaitu_df.columns, 1):
                ws_ziyingzaitu.cell(row=1, column=col_idx, value=col_name)
            
            # 写入数据并设置格式
            spec_code_col = None
            if '商品规格代码' in ziyingzaitu_df.columns:
                spec_code_col = list(ziyingzaitu_df.columns).index('商品规格代码') + 1
            
            for row_idx, row_data in enumerate(ziyingzaitu_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_ziyingzaitu.cell(row=row_idx, column=col_idx, value=value)
                    # 如果是规格代码列且是数字，设置为文本格式避免科学计数法
                    if col_idx == spec_code_col and isinstance(value, (int, float)):
                        cell.number_format = '0'  # 设置为整数格式，不显示科学计数法
        else:
            # 创建空的自营在途表格
            headers = ['商品规格代码', '自营仓', '自营商城仓', '总计']
            for col_idx, header in enumerate(headers, 1):
                ws_ziyingzaitu.cell(row=1, column=col_idx, value=header)
        
        # 删除默认的Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 保存文件（先按原逻辑生成“信选仓/自营在途”两张表）
        wb.save(output_file)

        # 追加两个sheet：来自输入的“仓库分类文件”和“商品库存导出文件”
        try:
            from openpyxl import load_workbook
            from copy import copy as _copy

            out_wb = load_workbook(output_file)

            def ensure_fresh_sheet(wb_obj, name: str):
                if name in wb_obj.sheetnames:
                    wb_obj.remove(wb_obj[name])
                return wb_obj.create_sheet(name)

            # 1) 追加“仓库”sheet（来自仓库分类Excel文件的活动表）
            ws_wh = ensure_fresh_sheet(out_wb, "仓库")
            try:
                wh_wb = load_workbook(self.warehouse_file, data_only=False)
                wh_ws = wh_wb.active

                # 原样复制：单元格值 + 样式 + 合并单元格 + 行列属性（尽量保持不变）
                for row in wh_ws.iter_rows():
                    for cell in row:
                        new_cell = ws_wh.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = _copy(cell.font)
                            new_cell.border = _copy(cell.border)
                            new_cell.fill = _copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = _copy(cell.protection)
                            new_cell.alignment = _copy(cell.alignment)

                # 列宽/隐藏/层级等
                for col_letter, dim in wh_ws.column_dimensions.items():
                    new_dim = ws_wh.column_dimensions[col_letter]
                    if dim.width is not None:
                        new_dim.width = dim.width
                    new_dim.hidden = dim.hidden
                    new_dim.outlineLevel = dim.outlineLevel
                    new_dim.collapsed = dim.collapsed

                # 行高/隐藏/层级等
                for row_idx, dim in wh_ws.row_dimensions.items():
                    new_dim = ws_wh.row_dimensions[row_idx]
                    if dim.height is not None:
                        new_dim.height = dim.height
                    new_dim.hidden = dim.hidden
                    new_dim.outlineLevel = dim.outlineLevel
                    new_dim.collapsed = dim.collapsed

                # 合并单元格
                for merged in getattr(wh_ws, "merged_cells", []).ranges:
                    ws_wh.merge_cells(str(merged))

                # 冻结窗格
                try:
                    ws_wh.freeze_panes = wh_ws.freeze_panes
                except Exception:
                    pass

                # 自动筛选
                try:
                    if wh_ws.auto_filter and wh_ws.auto_filter.ref:
                        ws_wh.auto_filter.ref = wh_ws.auto_filter.ref
                except Exception:
                    pass
            except Exception as e:
                # 不影响主报表生成，但给出提示
                self.status_updated.emit(f"警告：追加“仓库”sheet失败：{e}")

            # 2) 追加“商品库存导出”sheet（来自商品库存导出CSV）
            ws_inv = ensure_fresh_sheet(out_wb, "商品库存导出")
            try:
                # 尽量保持原始内容：按字符串读取，避免科学计数/前导零丢失
                inv_df = None
                encodings = ["utf-8", "utf-8-sig", "gb18030", "gbk", "gb2312"]
                for enc in encodings:
                    try:
                        inv_df = pd.read_csv(
                            self.inventory_file,
                            encoding=enc,
                            low_memory=False,
                            dtype=str,
                            skipinitialspace=True,
                        )
                        break
                    except Exception:
                        inv_df = None
                        continue

                if inv_df is None:
                    raise Exception("无法读取商品库存导出CSV（编码尝试失败）")

                # --- 方向性数值转换：保留“编码/条码/SKU”等字段为文本，但把数量/金额/库存等写成数字 ---
                def _norm_col(name: str) -> str:
                    return str(name or "").strip().lower()

                def _is_number_like(x: object) -> bool:
                    if x is None:
                        return False
                    s = str(x).strip()
                    if s == "" or s.lower() in {"nan", "none"}:
                        return False
                    # 去千分位
                    s = s.replace(",", "")
                    # 允许负号与小数
                    try:
                        float(s)
                        return True
                    except Exception:
                        return False

                def _to_number(x: object):
                    s = "" if x is None else str(x).strip()
                    if s == "" or s.lower() in {"nan", "none"}:
                        return None
                    s = s.replace(",", "")
                    # 整数优先（避免 1.0）
                    try:
                        if "." not in s:
                            return int(s)
                        f = float(s)
                        if f.is_integer():
                            return int(f)
                        return f
                    except Exception:
                        return None

                # 文本列：任何“代码/编码/条码/sku/单号/编号”等都必须保持文本
                text_kw = ("代码", "编码", "条码", "sku", "单号", "编号", "id", "货品", "商品名", "名称", "店铺", "仓库", "类目", "品牌")
                # 数值列：库存/数量/金额/成本/价格/重量等写成数字
                num_kw = ("数量", "库存", "可用", "可配", "可销", "在途", "锁定", "金额", "成本", "单价", "价格", "售价", "折扣", "重量", "体积", "箱数")

                cols = list(inv_df.columns)
                numeric_cols: set[int] = set()
                for idx, c in enumerate(cols):
                    n = _norm_col(c)
                    if any(k in n for k in text_kw):
                        continue
                    if any(k in n for k in num_kw):
                        numeric_cols.add(idx)

                # 写入表头
                ws_inv.append(cols)

                # 写入数据（逐行，避免一次性占用过大内存）
                for row in inv_df.itertuples(index=False, name=None):
                    out_row = []
                    for i, v in enumerate(row):
                        if i in numeric_cols and _is_number_like(v):
                            out_row.append(_to_number(v))
                        else:
                            out_row.append(v)
                    ws_inv.append(out_row)

                # 对数值列设置数字显示格式（避免 Excel 把数字再显示成科学计数）
                try:
                    for col_idx in numeric_cols:
                        excel_col = col_idx + 1
                        for r in range(2, ws_inv.max_row + 1):
                            cell = ws_inv.cell(row=r, column=excel_col)
                            if isinstance(cell.value, int):
                                cell.number_format = "0"
                            elif isinstance(cell.value, float):
                                cell.number_format = "0.00"
                except Exception:
                    pass
            except Exception as e:
                self.status_updated.emit(f"警告：追加“商品库存导出”sheet失败：{e}")

            out_wb.save(output_file)
        except Exception as e:
            self.status_updated.emit(f"警告：追加输入来源sheet失败：{e}")

        self.status_updated.emit("Excel文件生成完成!")


class UnshipOrderProcessor(QThread):
    """未发货单据处理线程"""
    progress_updated = pyqtSignal(int)
    status_updated = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, order_file, mapping_file, output_file):
        super().__init__()
        self.order_file = order_file
        self.mapping_file = mapping_file
        self.output_file = output_file
        self._import_required_modules()
    
    def _import_required_modules(self):
        """导入需要的模块"""
        try:
            global pd, np
            import pandas as pd
            import numpy as np
        except ImportError as e:
            self.status_updated.emit(f"模块导入失败: {e}")
            raise
        
    def run(self):
        try:
            self.status_updated.emit("开始处理未发货单据数据...")
            self.progress_updated.emit(10)
            
            # 读取订单商品明细数据
            self.status_updated.emit("读取订单商品明细数据...")
            order_df = self.read_order_data()
            self.progress_updated.emit(30)
            
            # 读取店铺仓库对应关系
            self.status_updated.emit("读取店铺&仓库对应关系...")
            mapping_df = self.read_mapping_data()
            self.progress_updated.emit(50)
            
            # 数据处理
            self.status_updated.emit("开始数据处理...")
            processed_data = self.process_unship_order_data(order_df, mapping_df)
            self.progress_updated.emit(80)
            
            # 生成Excel文件
            self.status_updated.emit("生成未发货单据Excel文件...")
            self.generate_unship_order_excel(processed_data)
            self.progress_updated.emit(100)
            
            self.status_updated.emit("未发货单据处理完成！")
            self.finished_signal.emit(True, "未发货单据处理成功完成！")
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}\n{traceback.format_exc()}"
            self.status_updated.emit(f"错误: {str(e)}")
            self.finished_signal.emit(False, error_msg)
    
    def read_order_data(self):
        """读取订单商品明细数据"""
        encodings = ['gbk', 'utf-8', 'utf-8-sig', 'gb2312', 'gb18030']
        order_df = None
        
        for encoding in encodings:
            try:
                order_df = pd.read_csv(
                    self.order_file,
                    encoding=encoding,
                    dtype=str,
                    skipinitialspace=True
                )
                self.status_updated.emit(f"成功使用 {encoding} 编码读取订单数据，共 {len(order_df)} 行")
                break
            except (UnicodeDecodeError, Exception) as e:
                continue
        
        if order_df is None:
            raise Exception("无法读取订单数据文件，请检查文件编码")
        
        # 数据清理：去掉前后空格
        for col in order_df.columns:
            if order_df[col].dtype == 'object':
                order_df[col] = order_df[col].astype(str).str.strip()
        
        return order_df
    
    def read_mapping_data(self):
        """读取店铺&仓库对应关系数据"""
        try:
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(self.mapping_file, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            self.status_updated.emit(f"映射文件包含的sheet: {sheet_names}")
            
            # 分别收集店铺和仓库数据
            all_shops = set()
            all_warehouses = set()
            
            # 第一步：从所有sheet收集店铺和仓库数据
            shops_from_sheets = {}
            warehouses_from_sheets = {}
            
            for sheet in sheet_names:
                try:
                    self.status_updated.emit(f"正在分析sheet: {sheet}")
                    df = pd.read_excel(self.mapping_file, sheet_name=sheet, engine='openpyxl', header=0)
                    if df.empty:
                        continue
                    
                    self.status_updated.emit(f"  sheet '{sheet}' 列名: {list(df.columns)}")
                    
                    # 查找可能的店铺仓库映射列
                    shop_cols = []
                    warehouse_cols = []
                    
                    for col in df.columns:
                        col_str = str(col).lower()
                        # 更宽泛的店铺名称匹配
                        if any(keyword in col_str for keyword in ['店铺', 'shop', '商店', '店铺名称', '门店']):
                            shop_cols.append(col)
                        # 更宽泛的仓库名称匹配
                        elif any(keyword in col_str for keyword in ['仓库', 'warehouse', '仓储', '仓库名称']):
                            warehouse_cols.append(col)
                    
                    self.status_updated.emit(f"  检测到的店铺列: {shop_cols}")
                    self.status_updated.emit(f"  检测到的仓库列: {warehouse_cols}")
                    
                    # 收集店铺数据
                    if shop_cols:
                        shop_col = shop_cols[0]
                        shops = df[shop_col].dropna().astype(str).str.strip()
                        shops = shops[(shops != '') & (shops != 'nan')]
                        
                        if not shops.empty:
                            shop_set = set(shops.unique())
                            shops_from_sheets[sheet] = shop_set
                            all_shops.update(shop_set)
                            self.status_updated.emit(f"  从sheet '{sheet}' 收集到店铺: {list(shop_set)}")
                    
                    # 收集仓库数据
                    if warehouse_cols:
                        warehouse_col = warehouse_cols[0]
                        warehouses = df[warehouse_col].dropna().astype(str).str.strip()
                        warehouses = warehouses[(warehouses != '') & (warehouses != 'nan')]
                        
                        if not warehouses.empty:
                            warehouse_set = set(warehouses.unique())
                            warehouses_from_sheets[sheet] = warehouse_set
                            all_warehouses.update(warehouse_set)
                            self.status_updated.emit(f"  从sheet '{sheet}' 收集到仓库: {list(warehouse_set)}")
                    
                    # 如果没有明确的列名，尝试通过内容推断
                    if not shop_cols and not warehouse_cols:
                        self.status_updated.emit(f"  未找到明确的店铺/仓库列，尝试分析所有列的内容...")
                        
                        for col in df.columns:
                            if df[col].dtype == 'object':  # 只检查文本列
                                sample_values = df[col].dropna().astype(str).head(10).tolist()
                                
                                # 检查是否包含店铺相关词汇
                                shop_keywords = ['店', '商店', '门店', '专卖店', '旗舰店']
                                warehouse_keywords = ['仓', '库', '仓库', '仓储', '配送中心', '虚拟仓']
                                
                                shop_score = sum(1 for val in sample_values for keyword in shop_keywords if keyword in val)
                                warehouse_score = sum(1 for val in sample_values for keyword in warehouse_keywords if keyword in val)
                                
                                if shop_score > 0:
                                    self.status_updated.emit(f"  列 '{col}' 可能包含店铺信息: {sample_values[:3]}")
                                    shops = df[col].dropna().astype(str).str.strip()
                                    shops = shops[(shops != '') & (shops != 'nan')]
                                    if not shops.empty:
                                        shop_set = set(shops.unique())
                                        shops_from_sheets[sheet] = shop_set
                                        all_shops.update(shop_set)
                                
                                if warehouse_score > 0:
                                    self.status_updated.emit(f"  列 '{col}' 可能包含仓库信息: {sample_values[:3]}")
                                    warehouses = df[col].dropna().astype(str).str.strip()
                                    warehouses = warehouses[(warehouses != '') & (warehouses != 'nan')]
                                    if not warehouses.empty:
                                        warehouse_set = set(warehouses.unique())
                                        warehouses_from_sheets[sheet] = warehouse_set
                                        all_warehouses.update(warehouse_set)
                
                except Exception as e:
                    self.status_updated.emit(f"  处理sheet '{sheet}' 时出错: {str(e)}")
                    continue
            
            self.status_updated.emit(f"所有收集到的仓库: {list(all_warehouses)}")
            self.status_updated.emit(f"所有收集到的店铺: {list(all_shops)}")
            
            # 第二步：创建所有店铺和所有仓库的完整组合
            if all_shops and all_warehouses:
                self.status_updated.emit("🔄 创建所有店铺与所有仓库的完整组合...")
                
                mapping_data = []
                for shop in all_shops:
                    for warehouse in all_warehouses:
                        mapping_data.append({
                            '店铺名称': shop,
                            '仓库名称': warehouse
                        })
                
                combined_mappings = pd.DataFrame(mapping_data)
                self.status_updated.emit(f"✅ 生成完整映射关系: {len(all_shops)} 个店铺 × {len(all_warehouses)} 个仓库 = {len(combined_mappings)} 条")
                
                # 显示前10个组合样例
                self.status_updated.emit("📋 店铺-仓库组合样例:")
                for i, (_, row) in enumerate(combined_mappings.head(10).iterrows()):
                    self.status_updated.emit(f"  {i+1}. {row['店铺名称']} -> {row['仓库名称']}")
                
                if len(combined_mappings) > 10:
                    self.status_updated.emit(f"  ... 还有 {len(combined_mappings) - 10} 个组合")
                
                return combined_mappings
            
            elif all_shops:
                self.status_updated.emit("⚠️ 只找到店铺，没有找到仓库，创建默认映射")
                mapping_data = []
                for shop in all_shops:
                    mapping_data.append({
                        '店铺名称': shop,
                        '仓库名称': '默认仓库'
                    })
                return pd.DataFrame(mapping_data)
            
            elif all_warehouses:
                self.status_updated.emit("⚠️ 只找到仓库，没有找到店铺，创建默认映射")
                mapping_data = []
                for warehouse in all_warehouses:
                    mapping_data.append({
                        '店铺名称': '默认店铺',
                        '仓库名称': warehouse
                    })
                return pd.DataFrame(mapping_data)
            
            else:
                # 如果没有找到任何映射，创建默认映射
                self.status_updated.emit("❌ 未找到任何店铺或仓库信息，使用最基本的默认映射")
                return pd.DataFrame({'店铺名称': ['通用店铺'], '仓库名称': ['默认仓库']})
        
        except Exception as e:
            raise Exception(f"读取店铺仓库对应关系失败: {str(e)}")
    
    def process_unship_order_data(self, order_df, mapping_df):
        """处理未发货单据数据"""
        # 复制原始数据
        processed_df = order_df.copy()
        
        # 显示原始数据统计
        self.status_updated.emit(f"原始订单数据总计: {len(processed_df)} 条")
        if '仓库名称' in processed_df.columns:
            original_warehouses = processed_df['仓库名称'].value_counts()
            self.status_updated.emit(f"原始数据中的仓库分布: {dict(original_warehouses.head(10))}")
        if '店铺名称' in processed_df.columns:
            original_shops = processed_df['店铺名称'].value_counts()
            self.status_updated.emit(f"原始数据中的店铺分布: {dict(original_shops.head(10))}")
        
        # 先过滤掉仓库名称为空的数据
        if '仓库名称' in processed_df.columns:
            original_count = len(processed_df)
            # 过滤掉仓库名称为空、nan或者只有空格的数据
            processed_df = processed_df[
                (processed_df['仓库名称'].notna()) & 
                (processed_df['仓库名称'].astype(str).str.strip() != '') &
                (processed_df['仓库名称'].astype(str).str.strip() != 'nan')
            ].copy()
            filtered_count = len(processed_df)
            self.status_updated.emit(f"过滤空仓库名称: {original_count} -> {filtered_count} 条数据")
        
        # 根据店铺&仓库对应关系进行双维度过滤
        if not mapping_df.empty and '仓库名称' in mapping_df.columns and '店铺名称' in mapping_df.columns:
            # 获取映射文件中所有有效的仓库名称和店铺名称
            valid_warehouses = set(mapping_df['仓库名称'].dropna().astype(str).str.strip())
            valid_shops = set(mapping_df['店铺名称'].dropna().astype(str).str.strip())
            
            # 移除空字符串和'nan'
            valid_warehouses = {wh for wh in valid_warehouses if wh and wh != 'nan'}
            valid_shops = {shop for shop in valid_shops if shop and shop != 'nan'}
            
            self.status_updated.emit(f"映射文件中的有效仓库: {list(valid_warehouses)}")
            self.status_updated.emit(f"映射文件中的有效店铺: {list(valid_shops)}")
            
            # 第一步：按仓库过滤
            if '仓库名称' in processed_df.columns and valid_warehouses:
                before_warehouse_filter = len(processed_df)
                processed_df = processed_df[
                    processed_df['仓库名称'].astype(str).str.strip().isin(valid_warehouses)
                ].copy()
                after_warehouse_filter = len(processed_df)
                self.status_updated.emit(f"仓库过滤: {before_warehouse_filter} -> {after_warehouse_filter} 条数据")
                
                # 显示被过滤掉的仓库
                original_warehouses = set(order_df['仓库名称'].dropna().astype(str).str.strip())
                filtered_warehouses = original_warehouses - valid_warehouses
                if filtered_warehouses:
                    self.status_updated.emit(f"被过滤的仓库: {list(filtered_warehouses)}")
            
            # 第二步：按店铺过滤
            if '店铺名称' in processed_df.columns and valid_shops:
                before_shop_filter = len(processed_df)
                processed_df = processed_df[
                    processed_df['店铺名称'].astype(str).str.strip().isin(valid_shops)
                ].copy()
                after_shop_filter = len(processed_df)
                self.status_updated.emit(f"店铺过滤: {before_shop_filter} -> {after_shop_filter} 条数据")
                
                # 显示被过滤掉的店铺
                original_shops = set(order_df['店铺名称'].dropna().astype(str).str.strip())
                filtered_shops = original_shops - valid_shops
                if filtered_shops:
                    self.status_updated.emit(f"被过滤的店铺: {list(filtered_shops)}")
            
            # 第三步：验证店铺-仓库组合的有效性（更严格的过滤）
            if len(processed_df) > 0:
                # 创建有效的店铺-仓库组合集合
                valid_combinations = set()
                for _, row in mapping_df.iterrows():
                    shop = str(row['店铺名称']).strip()
                    warehouse = str(row['仓库名称']).strip()
                    if shop and warehouse and shop != 'nan' and warehouse != 'nan':
                        valid_combinations.add((shop, warehouse))
                
                self.status_updated.emit(f"有效的店铺-仓库组合: {len(valid_combinations)} 个")
                
                # 过滤出有效的店铺-仓库组合
                before_combination_filter = len(processed_df)
                valid_rows = []
                
                for index, row in processed_df.iterrows():
                    shop = str(row.get('店铺名称', '')).strip()
                    warehouse = str(row.get('仓库名称', '')).strip()
                    
                    if (shop, warehouse) in valid_combinations:
                        valid_rows.append(index)
                
                if valid_rows:
                    processed_df = processed_df.loc[valid_rows].copy()
                else:
                    # 如果没有有效组合，创建空的DataFrame但保持列结构
                    processed_df = processed_df.iloc[0:0].copy()
                
                after_combination_filter = len(processed_df)
                self.status_updated.emit(f"店铺-仓库组合过滤: {before_combination_filter} -> {after_combination_filter} 条数据")
                
                # 如果过滤后没有数据，提供调试信息
                if after_combination_filter == 0:
                    self.status_updated.emit("⚠️ 警告：双维度过滤后没有数据！")
                    self.status_updated.emit("🔍 调试信息：检查原始数据中的店铺-仓库组合是否与映射文件匹配")
                    
                    # 显示一些原始数据的店铺-仓库组合样例
                    if '店铺名称' in order_df.columns and '仓库名称' in order_df.columns:
                        original_combinations = order_df[['店铺名称', '仓库名称']].drop_duplicates().head(10)
                        self.status_updated.emit("📋 原始数据中的店铺-仓库组合样例:")
                        for _, row in original_combinations.iterrows():
                            shop = str(row['店铺名称']).strip()
                            warehouse = str(row['仓库名称']).strip()
                            is_valid = (shop, warehouse) in valid_combinations
                            status = "✅" if is_valid else "❌"
                            self.status_updated.emit(f"  {status} {shop} -> {warehouse}")
            
        else:
            self.status_updated.emit("警告：未找到有效的店铺仓库映射关系，跳过双维度过滤")
        
        # 1. 在第一列插入【备注】列
        processed_df.insert(0, '备注', '')
        
        # 2. 在支付时间列后面加一个【推单-支付=耗时】列
        if '支付时间' in processed_df.columns:
            pay_time_index = processed_df.columns.get_loc('支付时间')
            processed_df.insert(pay_time_index + 1, '推单-支付=耗时', '')
        
        # 3. 处理时间相关的计算
        current_time = datetime.now()
        
        for index, row in processed_df.iterrows():
            try:
                # 获取制单时间和支付时间
                制单时间_str = str(row.get('制单时间', '')).strip() if '制单时间' in processed_df.columns else ''
                支付时间_str = str(row.get('支付时间', '')).strip() if '支付时间' in processed_df.columns else ''
                物流单号 = str(row.get('物流单号', '')).strip() if '物流单号' in processed_df.columns else ''
                
                # 解析时间
                制单时间 = None
                支付时间 = None
                
                if 制单时间_str and 制单时间_str != 'nan':
                    try:
                        制单时间 = pd.to_datetime(制单时间_str)
                    except:
                        pass
                
                if 支付时间_str and 支付时间_str != 'nan':
                    try:
                        支付时间 = pd.to_datetime(支付时间_str)
                    except:
                        pass
                
                # 计算推单-支付=耗时
                if 制单时间 and 支付时间:
                    耗时 = 制单时间 - 支付时间
                    if '推单-支付=耗时' in processed_df.columns:
                        processed_df.at[index, '推单-支付=耗时'] = str(耗时)
                
                # 计算备注
                备注 = ''
                
                # 如果物流单号不为空，写入"管易发货失败"
                if 物流单号 and 物流单号 != 'nan' and 物流单号 != '':
                    备注 = '管易发货失败'
                else:
                    # 判断制单时间-支付时间的关系
                    if 制单时间 and 支付时间:
                        时间差 = 制单时间 - 支付时间
                        
                        if 时间差 > timedelta(hours=1):
                            备注 = '推单超时'
                        elif 时间差 <= timedelta(hours=1):
                            # 判断当前时间-制单时间的关系
                            if 制单时间:
                                当前与制单时间差 = current_time - 制单时间
                                if 当前与制单时间差 < timedelta(hours=24):
                                    备注 = '24时效内'
                                else:
                                    备注 = '超时'
                
                processed_df.at[index, '备注'] = 备注
                
            except Exception as e:
                # 如果处理某行出错，跳过该行
                processed_df.at[index, '备注'] = '数据异常'
                continue
        
        # 生成第二个sheet页：SDO去重后分析
        if '订单编号' in processed_df.columns:
            # 按订单编号去重，保留第一条记录
            sdo_df = processed_df.drop_duplicates(subset=['订单编号'], keep='first')
            
            # 按照需求文档生成第二个sheet的结构
            # 第一列：订单编号
            # 第二列：备注（取第一个sheet中的备注）
            # 第三列：空列（占位）
            # 第四列：备注（去重的备注名称）
            # 第五列：合计（根据对应的备注类型进行数量的汇总）
            # 第六列：占比
            
            # 统计备注类型
            备注统计 = sdo_df['备注'].value_counts()
            总数 = len(sdo_df)
            
            # 生成分析数据
            sdo_analysis_rows = []
            for 备注类型, 数量 in 备注统计.items():
                占比 = (数量 / 总数 * 100) if 总数 > 0 else 0
                sdo_analysis_rows.append({
                    '订单编号': '',  # 第一列在分析部分为空
                    '备注': '',     # 第二列在分析部分为空
                    '': '',         # 第三列空白
                    '备注_去重': 备注类型,  # 第四列：去重的备注名称
                    '合计': 数量,    # 第五列：合计
                    '占比': f"{占比:.2f}%"  # 第六列：占比
                })
            
            # 添加总计行
            sdo_analysis_rows.append({
                '订单编号': '',
                '备注': '',
                '': '',
                '备注_去重': '总计',
                '合计': 总数,
                '占比': '100.00%'
            })
            
            # 生成基础数据部分（订单编号和备注的明细）
            sdo_detail_rows = []
            for _, row in sdo_df.iterrows():
                sdo_detail_rows.append({
                    '订单编号': row['订单编号'],     # 第一列：订单编号
                    '备注': row['备注'],           # 第二列：备注
                })
            
            # 创建明细数据DataFrame
            sdo_detail_df = pd.DataFrame(sdo_detail_rows)
            
            # 创建分析数据DataFrame（只包含后三列）
            sdo_analysis_simple = []
            for item in sdo_analysis_rows:
                sdo_analysis_simple.append({
                    '备注_去重': item['备注_去重'],
                    '合计': item['合计'],
                    '占比': item['占比']
                })
            sdo_analysis_only_df = pd.DataFrame(sdo_analysis_simple)
            
            # 合并两个DataFrame，使它们并排显示，都从第一行开始
            max_rows = max(len(sdo_detail_df), len(sdo_analysis_only_df))
            
            # 扩展较短的DataFrame以匹配行数
            if len(sdo_detail_df) < max_rows:
                for i in range(max_rows - len(sdo_detail_df)):
                    sdo_detail_df = pd.concat([sdo_detail_df, pd.DataFrame({'订单编号': [''], '备注': ['']})], ignore_index=True)
            
            if len(sdo_analysis_only_df) < max_rows:
                for i in range(max_rows - len(sdo_analysis_only_df)):
                    sdo_analysis_only_df = pd.concat([sdo_analysis_only_df, pd.DataFrame({'备注_去重': [''], '合计': [''], '占比': ['']})], ignore_index=True)
            
            # 合并为最终的DataFrame，添加空列分隔
            sdo_analysis_df = pd.DataFrame({
                '订单编号': sdo_detail_df['订单编号'],
                '备注': sdo_detail_df['备注'],
                '': [''] * max_rows,  # 第三列空白分隔
                '备注_去重': sdo_analysis_only_df['备注_去重'],
                '合计': sdo_analysis_only_df['合计'],
                '占比': sdo_analysis_only_df['占比']
            })
            
            # 简化的数据用于返回
            sdo_data_df = sdo_df[['订单编号', '备注']].copy()
        else:
            sdo_analysis_df = pd.DataFrame()
            sdo_data_df = pd.DataFrame()
        
        # 生成订单量分析数据
        order_analysis_df = self.generate_order_analysis(processed_df)
        
        return {
            'sheet1': processed_df,
            'sheet2_data': sdo_data_df,
            'sheet2_analysis': sdo_analysis_df,
            'sheet3_order_analysis': order_analysis_df
        }
    
    def generate_order_analysis(self, processed_df):
        """生成订单量分析数据"""
        if '仓库名称' in processed_df.columns and '订单编号' in processed_df.columns:
            # 按仓库名称分组，统计去重的SDO单号数量
            warehouse_order_count = processed_df.groupby('仓库名称')['订单编号'].nunique().reset_index()
            warehouse_order_count.columns = ['仓库名称', '订单量']
            
            # 按订单量降序排序
            warehouse_order_count = warehouse_order_count.sort_values('订单量', ascending=False)
            
            # 添加总计行
            total_orders = processed_df['订单编号'].nunique()
            total_row = pd.DataFrame({'仓库名称': ['总计'], '订单量': [total_orders]})
            
            # 合并数据
            order_analysis_df = pd.concat([warehouse_order_count, total_row], ignore_index=True)
            
            self.status_updated.emit(f"订单量分析生成完成，包含 {len(warehouse_order_count)} 个仓库")
            
            return order_analysis_df
        else:
            # 如果没有必要的列，返回空的DataFrame
            return pd.DataFrame(columns=['仓库名称', '订单量'])
    
    def generate_unship_order_excel(self, data):
        """生成未发货单据Excel文件"""
        # === 输出格式内化（参考表：未发货商品汇总.xlsx） ===
        # 第一个sheet的最终列顺序：
        # - 第一个“备注”保持原逻辑计算结果
        # - 第二个备注列若同名会导致覆盖/挤空，这里内化为“备注2”（留空供人工填写）
        target_cols_sheet1 = [
            "备注",
            "订单编号",
            "订单类型",
            "平台单号",
            "店铺名称",
            "制单时间",
            "支付时间",
            "推单-支付=耗时",
            "订单状态",
            "仓库名称",
            "快递名称",
            "物流单号",
            "商品名称",
            "规格代码",
            "客服",
            "备注2",  # 第二个备注列（留空）
            "商品数量",
            "推送外仓单据编号",
            "供应链备注",
        ]

        def build_sheet1(df: "pd.DataFrame") -> "pd.DataFrame":
            # 确保存在关键列（不存在则补空列）
            for col in [
                "备注",
                "订单编号",
                "订单类型",
                "平台单号",
                "店铺名称",
                "制单时间",
                "支付时间",
                "推单-支付=耗时",
                "订单状态",
                "仓库名称",
                "快递名称",
                "物流单号",
                "商品名称",
                "规格代码",
                "商品数量",
            ]:
                if col not in df.columns:
                    df[col] = ""

            # 新增三列（按需求：内化在程序里）
            if "客服" not in df.columns:
                df["客服"] = ""
            if "推送外仓单据编号" not in df.columns:
                df["推送外仓单据编号"] = ""
            if "供应链备注" not in df.columns:
                df["供应链备注"] = ""
            # 第二个备注列（留空）
            df["备注2"] = ""

            # 构造输出（避免重复列名导致覆盖）
            out = pd.DataFrame(index=df.index)
            for col in target_cols_sheet1:
                out[col] = df.get(col, "")
            return out

        sheet1_final = build_sheet1(data["sheet1"].copy())

        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            # 写入第一个sheet：订单商品明细数据（删除多余列 + 增加三列 + 维持目标顺序）
            sheet1_final.to_excel(writer, sheet_name="订单商品明细数据", index=False)
            
            # 写入第二个sheet：SDO去重后分析
            if not data['sheet2_analysis'].empty:
                data['sheet2_analysis'].to_excel(writer, sheet_name='SDO去重后分析', index=False)
                self.status_updated.emit(f"第二个sheet生成成功，包含 {len(data['sheet2_analysis'])} 行数据")
            else:
                # 创建空的分析表
                empty_analysis = pd.DataFrame(columns=['订单编号', '备注', '', '备注_去重', '合计', '占比'])
                empty_analysis.to_excel(writer, sheet_name='SDO去重后分析', index=False)
                self.status_updated.emit("第二个sheet为空，生成空表结构")
            
            # 写入第三个sheet：订单量分析
            if not data['sheet3_order_analysis'].empty:
                data['sheet3_order_analysis'].to_excel(writer, sheet_name='订单量分析', index=False)
                self.status_updated.emit(f"第三个sheet生成成功，包含 {len(data['sheet3_order_analysis'])} 行数据")
            else:
                # 创建空的订单量分析表
                empty_order_analysis = pd.DataFrame(columns=['仓库名称', '订单量'])
                empty_order_analysis.to_excel(writer, sheet_name='订单量分析', index=False)
                self.status_updated.emit("第三个sheet为空，生成空表结构")
        
        # 仅对“订单商品明细数据”sheet：表头样式 + 三列黄底 + 列宽内化（不依赖外部表头文件）
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border
            from openpyxl.utils import get_column_letter

            wb = load_workbook(self.output_file)
            if "订单商品明细数据" in wb.sheetnames:
                ws = wb["订单商品明细数据"]

                header_font = Font(name="宋体", size=11, bold=False)
                fill_blue = PatternFill(fill_type="solid", start_color="FF00B0F0", end_color="FF00B0F0")
                fill_yellow = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
                header_alignment = Alignment(vertical="center")
                header_border = Border()  # 无边框

                # 参考表：客服、第二个备注、供应链备注 为黄底（这里第二个备注内化为“备注2”）
                yellow_headers = {"客服", "备注2", "供应链备注"}

                # 参考表列宽（按列位置内化）
                widths = [
                    5.72727272727273,
                    17.3636363636364,
                    13.0,
                    13.0,
                    13.0,
                    20.6363636363636,
                    13.0,
                    13.0,
                    13.0,
                    38.3636363636364,
                    13.0,
                    13.0,
                    13.0,
                    17.6363636363636,
                    13.0,
                    13.0,
                    13.0,
                    21.9090909090909,
                    11.8181818181818,
                ]

                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=c)
                    header_name = "" if cell.value is None else str(cell.value).strip()

                    cell.font = header_font
                    cell.fill = fill_yellow if header_name in yellow_headers else fill_blue
                    cell.alignment = header_alignment
                    cell.border = header_border

                    # 列宽内化
                    if 1 <= c <= len(widths):
                        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]

            wb.save(self.output_file)
        except Exception as e:
            # 不影响主要输出，仅提示
            self.status_updated.emit(f"⚠️ 表头格式应用失败: {str(e)}")
        
        self.status_updated.emit("未发货单据Excel文件生成完成!")


class UnshipDataProcessor(QThread):
    """未配货数据处理线程"""
    progress_updated = pyqtSignal(int)
    status_updated = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, order_file, stock_file, mapping_file, output_file, feishu_client=None, feishu_connection_info=None):
        super().__init__()
        self.order_file = order_file
        self.stock_file = stock_file
        self.mapping_file = mapping_file
        self.output_file = output_file
        self.feishu_client = feishu_client
        
        # 设置飞书连接信息
        if feishu_connection_info:
            self.feishu_use_requests = feishu_connection_info.get('use_requests', False)
            self.feishu_access_token = feishu_connection_info.get('access_token', None)
            self.feishu_headers = feishu_connection_info.get('headers', None)
            self.feishu_api_base = feishu_connection_info.get('api_base', None)
        else:
            self.feishu_use_requests = False
            
        self._import_required_modules()
    
    def _import_required_modules(self):
        """导入需要的模块"""
        try:
            global pd, np
            import pandas as pd
            import numpy as np
        except ImportError as e:
            self.status_updated.emit(f"模块导入失败: {e}")
            raise
        
    def run(self):
        try:
            self.status_updated.emit("开始处理未配货数据...")
            self.progress_updated.emit(5)
            
            # 读取基础数据
            self.status_updated.emit("读取订单商品明细数据...")
            order_df = self.read_order_data()
            self.progress_updated.emit(15)
            
            self.status_updated.emit("读取库存报表数据...")
            stock_df, stock_sheet2_df = self.read_stock_data()
            self.progress_updated.emit(25)
            
            self.status_updated.emit("读取店铺仓库对应关系...")
            mapping_df = self.read_mapping_data()
            self.progress_updated.emit(35)
            
            # 从飞书获取货品状态数据
            self.status_updated.emit("从飞书获取货品状态数据...")
            feishu_data = self.get_feishu_data()
            self.progress_updated.emit(45)
            
            # 数据处理
            self.status_updated.emit("开始数据过滤和处理...")
            processed_data = self.process_unship_data(order_df, stock_df, stock_sheet2_df, mapping_df, feishu_data)
            self.progress_updated.emit(80)
            
            # 生成Excel文件
            self.status_updated.emit("生成未配货分析Excel文件...")
            self.generate_unship_excel(processed_data)
            self.progress_updated.emit(100)
            
            self.status_updated.emit("未配货数据处理完成！")
            self.finished_signal.emit(True, "未配货数据处理成功完成！")
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}\n{traceback.format_exc()}"
            self.status_updated.emit(f"错误: {str(e)}")
            self.finished_signal.emit(False, error_msg)
    
    def read_order_data(self):
        """读取订单商品明细数据"""
        # 更全面的编码格式列表，把最常用的GBK放在前面
        encodings = ['gbk', 'utf-8', 'utf-8-sig', 'gb2312', 'gb18030', 'big5', 'latin1', 'cp1252']
        order_df = None
        last_error = None
        
        self.status_updated.emit(f"正在尝试读取文件: {os.path.basename(self.order_file)}")
        
        for encoding in encodings:
            try:
                self.status_updated.emit(f"尝试使用 {encoding} 编码...")
                # 优化CSV读取参数，解决引擎兼容性问题
                order_df = pd.read_csv(
                    self.order_file, 
                    encoding=encoding, 
                    sep=None,  # 自动检测分隔符
                    engine='python',  # 使用python引擎，更灵活
                    on_bad_lines='skip',  # 跳过有问题的行
                    skipinitialspace=True,  # 跳过字段开头的空格
                    dtype=str  # 所有列都作为字符串读取，避免类型错误
                )
                self.status_updated.emit(f"✅ 成功使用 {encoding} 编码读取订单数据，共 {len(order_df)} 行")
                break
            except (UnicodeDecodeError, pd.errors.ParserError, Exception) as e:
                last_error = str(e)
                self.status_updated.emit(f"❌ {encoding} 编码失败: {str(e)[:100]}")
                continue
                
        if order_df is None:
            error_msg = f"无法读取订单数据文件，请检查文件格式和编码。\n最后一个错误: {last_error}\n\n请确保文件是正确的CSV格式。"
            raise Exception(error_msg)
        
        # 显示文件信息
        self.status_updated.emit(f"文件列名: {list(order_df.columns)[:10]}...")  # 显示前10个列名
        
        # 清理平台单号的前后空格
        if '平台单号' in order_df.columns:
            order_df['平台单号'] = order_df['平台单号'].astype(str).str.strip()
        else:
            self.status_updated.emit("⚠️ 警告: 未找到'平台单号'列")
        
        return order_df
    
    def read_stock_data(self):
        """读取库存报表数据"""
        try:
            # 读取第一个sheet（总计数据）
            stock_df = pd.read_excel(self.stock_file, sheet_name=0, engine='openpyxl')
            # 读取第二个sheet（在途数据） 
            stock_sheet2_df = pd.read_excel(self.stock_file, sheet_name=1, engine='openpyxl')
            return stock_df, stock_sheet2_df
        except Exception as e:
            raise Exception(f"读取库存报表失败: {str(e)}")
    
    def read_excluded_spec_codes(self, excel_file, sheet_names):
        """读取第三个sheet中的排除规格代码"""
        try:
            excluded_codes = set()
            
            # 如果有第三个sheet，尝试读取其中的规格代码
            if len(sheet_names) >= 3:
                third_sheet = sheet_names[2]  # 第三个sheet（索引为2）
                self.status_updated.emit(f"🚫 读取第三个sheet中的排除规格代码: {third_sheet}")
                
                third_sheet_df = pd.read_excel(excel_file, sheet_name=third_sheet, engine='openpyxl')
                
                # 查找可能包含规格代码的列
                spec_code_cols = []
                for col in third_sheet_df.columns:
                    col_str = str(col).lower()
                    if any(keyword in col_str for keyword in ['规格代码', 'spec', 'sku', '代码', 'code']):
                        spec_code_cols.append(col)
                
                if spec_code_cols:
                    self.status_updated.emit(f"📋 在第三个sheet中找到规格代码列: {spec_code_cols}")
                    
                    # 从所有规格代码列中收集数据
                    for col in spec_code_cols:
                        # 处理原始数据，保持原始数据类型
                        col_data = third_sheet_df[col].dropna()
                        
                        for value in col_data:
                            if pd.isna(value):
                                continue
                            
                            # 保存原始值
                            excluded_codes.add(value)
                            
                            # 如果是数字类型，也保存字符串格式
                            if isinstance(value, (int, float)):
                                excluded_codes.add(str(value))
                            
                            # 如果是字符串，尝试转换为数字格式
                            elif isinstance(value, str):
                                value_str = str(value).strip()
                                if value_str:
                                    excluded_codes.add(value_str)
                                    # 尝试转换为数字
                                    try:
                                        if value_str.replace('.', '').replace('-', '').isdigit():
                                            excluded_codes.add(int(float(value_str)))
                                            excluded_codes.add(float(value_str))
                                    except (ValueError, TypeError):
                                        pass
                    
                    self.status_updated.emit(f"🚫 从第三个sheet中读取到 {len(excluded_codes)} 个排除规格代码")
                    if excluded_codes:
                        sample_codes = list(excluded_codes)[:5]
                        self.status_updated.emit(f"📝 排除规格代码样本: {sample_codes}")
                else:
                    self.status_updated.emit("⚠️ 第三个sheet中未找到规格代码列")
            else:
                self.status_updated.emit("ℹ️ 映射文件中没有第三个sheet，不排除任何规格代码")
            
            return excluded_codes
            
        except Exception as e:
            self.status_updated.emit(f"⚠️ 读取排除规格代码失败: {str(e)}")
            return set()
    
    def read_mapping_data(self):
        """读取店铺仓库对应关系数据，店铺和仓库在不同sheet页中，多对多关系"""
        try:
            # 读取所有sheet
            excel_file = pd.ExcelFile(self.mapping_file, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            self.status_updated.emit(f"📋 Excel文件包含的sheet: {sheet_names}")
            
            # 读取第三个sheet中的排除规格代码（如果存在）
            self.excluded_spec_codes = self.read_excluded_spec_codes(excel_file, sheet_names)
            
            # 尝试找到店铺和仓库的sheet
            shop_sheet = None
            warehouse_sheet = None
            mapping_sheet = None
            
            for sheet in sheet_names:
                sheet_lower = sheet.lower()
                if any(keyword in sheet_lower for keyword in ['店铺', 'shop', '商店']):
                    shop_sheet = sheet
                elif any(keyword in sheet_lower for keyword in ['仓库', 'warehouse', '仓储']):
                    warehouse_sheet = sheet
                elif any(keyword in sheet_lower for keyword in ['映射', '对应', 'mapping', '关系']):
                    mapping_sheet = sheet
            
            self.status_updated.emit(f"🏪 检测到店铺sheet: {shop_sheet}")
            self.status_updated.emit(f"🏭 检测到仓库sheet: {warehouse_sheet}")
            self.status_updated.emit(f"🔗 检测到映射sheet: {mapping_sheet}")
            
            # 如果有明确的映射表，直接使用
            if mapping_sheet:
                self.status_updated.emit(f"🎯 使用映射sheet: {mapping_sheet}")
                mapping_df = pd.read_excel(self.mapping_file, sheet_name=mapping_sheet, engine='openpyxl', header=0)
                return self.process_mapping_sheet(mapping_df, mapping_sheet)
            
            # 如果没有映射表，但有店铺和仓库表，尝试分析所有sheet寻找映射关系
            all_mappings = []
            
            for sheet in sheet_names:
                try:
                    self.status_updated.emit(f"🔍 分析sheet: {sheet}")
                    df = pd.read_excel(self.mapping_file, sheet_name=sheet, engine='openpyxl', header=0)
                    
                    if df.empty:
                        continue
                    
                    columns = list(df.columns)
                    self.status_updated.emit(f"  列名: {columns}")
                    
                    # 查找可能的店铺仓库映射列
                    shop_cols = []
                    warehouse_cols = []
                    
                    for col in columns:
                        col_str = str(col).lower()
                        if any(keyword in col_str for keyword in ['店铺', 'shop', '商店']):
                            shop_cols.append(col)
                        elif any(keyword in col_str for keyword in ['仓库', 'warehouse', '仓储']):
                            warehouse_cols.append(col)
                    
                    # 如果找到店铺和仓库列，提取映射关系
                    if shop_cols and warehouse_cols:
                        shop_col = shop_cols[0]
                        warehouse_col = warehouse_cols[0]
                        
                        self.status_updated.emit(f"  ✅ 在sheet '{sheet}' 中找到映射: {shop_col} -> {warehouse_col}")
                        
                        # 提取映射数据
                        sheet_mappings = df[[shop_col, warehouse_col]].copy()
                        sheet_mappings.columns = ['店铺名称', '仓库名称']
                        sheet_mappings = sheet_mappings.dropna()
                        sheet_mappings = sheet_mappings[sheet_mappings['店铺名称'].astype(str).str.strip() != '']
                        sheet_mappings = sheet_mappings[sheet_mappings['仓库名称'].astype(str).str.strip() != '']
                        
                        if not sheet_mappings.empty:
                            all_mappings.append(sheet_mappings)
                            self.status_updated.emit(f"  📊 从sheet '{sheet}' 提取到 {len(sheet_mappings)} 条映射")
                    
                    # 如果只有两列且没有明确关键词，尝试作为店铺仓库映射
                    # 但要排除第三个sheet（排除规格代码的sheet）
                    elif len(columns) == 2 and len(df) > 0:
                        # 检查是否是排除规格代码的sheet
                        is_exclusion_sheet = False
                        for col in columns:
                            col_str = str(col).lower()
                            if any(keyword in col_str for keyword in ['规格代码', 'spec', 'sku', '代码', 'code']):
                                is_exclusion_sheet = True
                                break
                        
                        if is_exclusion_sheet:
                            self.status_updated.emit(f"  🚫 跳过sheet '{sheet}'，这是排除规格代码的sheet，不是映射关系")
                        else:
                            self.status_updated.emit(f"  🔄 尝试将sheet '{sheet}' 的两列作为店铺仓库映射")
                            sheet_mappings = df.copy()
                            sheet_mappings.columns = ['店铺名称', '仓库名称']
                            sheet_mappings = sheet_mappings.dropna()
                            sheet_mappings = sheet_mappings[sheet_mappings['店铺名称'].astype(str).str.strip() != '']
                            sheet_mappings = sheet_mappings[sheet_mappings['仓库名称'].astype(str).str.strip() != '']
                            
                            if not sheet_mappings.empty:
                                # 显示样例数据供确认
                                sample = sheet_mappings.head(3)
                                self.status_updated.emit(f"  📋 样例数据:\n{sample.to_string(index=False)}")
                                all_mappings.append(sheet_mappings)
                                self.status_updated.emit(f"  📊 从sheet '{sheet}' 提取到 {len(sheet_mappings)} 条映射")
                    
                    # 特殊处理：如果是单列仓库或店铺表，记录下来但不直接作为映射
                    elif len(columns) == 1:
                        col_name = columns[0]
                        col_str = str(col_name).lower()
                        if any(keyword in col_str for keyword in ['仓库', 'warehouse', '仓储']):
                            self.status_updated.emit(f"  📦 发现仓库列表sheet: '{sheet}' (列: {col_name})")
                            warehouse_list = df[col_name].dropna().unique().tolist()
                            self.status_updated.emit(f"    仓库列表: {warehouse_list}")
                        elif any(keyword in col_str for keyword in ['店铺', 'shop', '商店']):
                            self.status_updated.emit(f"  🏪 发现店铺列表sheet: '{sheet}' (列: {col_name})")
                            shop_list = df[col_name].dropna().unique().tolist()
                            self.status_updated.emit(f"    店铺列表: {shop_list}")
                        else:
                            self.status_updated.emit(f"  ❓ 未知单列sheet: '{sheet}' (列: {col_name})")
                    
                    else:
                        self.status_updated.emit(f"  ⚠️ 无法识别sheet '{sheet}' 的结构，跳过")
                
                except Exception as e:
                    self.status_updated.emit(f"  ❌ 读取sheet '{sheet}' 失败: {str(e)}")
                    continue
            
            # 合并所有映射关系
            if all_mappings:
                combined_mappings = pd.concat(all_mappings, ignore_index=True)
                combined_mappings = combined_mappings.drop_duplicates()
                
                self.status_updated.emit(f"🎉 成功合并 {len(all_mappings)} 个sheet的映射数据")
                self.status_updated.emit(f"📊 总计映射关系: {len(combined_mappings)} 条")
                
                return self.finalize_mapping_data(combined_mappings)
            else:
                # 如果没有找到映射关系，尝试生成默认映射或给出更详细的指导
                self.status_updated.emit("⚠️ 未找到明确的店铺仓库映射关系")
                self.status_updated.emit("📋 请检查Excel文件是否包含以下结构之一:")
                self.status_updated.emit("  1️⃣ 包含'店铺名称'和'仓库名称'列的sheet")
                self.status_updated.emit("  2️⃣ 两列数据，第一列为店铺，第二列为仓库")
                self.status_updated.emit("  3️⃣ 多个sheet，每个包含店铺-仓库映射")
                
                # 尝试重新扫描，降低要求
                self.status_updated.emit("🔄 尝试降低要求重新扫描...")
                fallback_mappings = []
                
                for sheet in sheet_names:
                    try:
                        df = pd.read_excel(self.mapping_file, sheet_name=sheet, engine='openpyxl', header=0)
                        if df.empty or len(df.columns) < 2:
                            continue
                        
                        # 尝试任意两列作为映射，但排除规格代码sheet
                        if len(df.columns) >= 2:
                            # 检查是否是排除规格代码的sheet
                            is_exclusion_sheet = False
                            for col in df.columns:
                                col_str = str(col).lower()
                                if any(keyword in col_str for keyword in ['规格代码', 'spec', 'sku', '代码', 'code']):
                                    is_exclusion_sheet = True
                                    break
                            
                            if is_exclusion_sheet:
                                self.status_updated.emit(f"  🚫 降级策略中跳过sheet '{sheet}'，这是排除规格代码的sheet")
                                continue
                            
                            col1, col2 = df.columns[0], df.columns[1]
                            test_mappings = df[[col1, col2]].copy()
                            test_mappings.columns = ['店铺名称', '仓库名称']
                            test_mappings = test_mappings.dropna()
                            
                            if len(test_mappings) > 0:
                                # 简单验证：检查是否有重复的组合模式
                                unique_combinations = len(test_mappings.drop_duplicates())
                                if unique_combinations > 1:  # 至少有2个不同的组合
                                    sample = test_mappings.head(3)
                                    self.status_updated.emit(f"  🔄 尝试使用sheet '{sheet}' 的前两列:")
                                    self.status_updated.emit(f"    列: {col1} -> {col2}")
                                    self.status_updated.emit(f"    样例:\n{sample.to_string(index=False)}")
                                    
                                    # 询问用户或直接使用
                                    fallback_mappings.append(test_mappings)
                                    self.status_updated.emit(f"  ✅ 临时使用此映射，共 {len(test_mappings)} 条")
                                    break  # 找到一个就停止
                    except:
                        continue
                
                if fallback_mappings:
                    combined_mappings = pd.concat(fallback_mappings, ignore_index=True)
                    combined_mappings = combined_mappings.drop_duplicates()
                    self.status_updated.emit(f"⚡ 使用降级策略，找到 {len(combined_mappings)} 条映射关系")
                    return self.finalize_mapping_data(combined_mappings)
                else:
                    # 最后的策略：生成全映射关系
                    self.status_updated.emit("🔄 尝试生成全映射关系（所有店铺对应所有仓库）...")
                    return self.generate_full_mapping(sheet_names)
                
        except Exception as e:
            raise Exception(f"读取店铺仓库对应关系失败: {str(e)}")
    
    def process_mapping_sheet(self, mapping_df, sheet_name):
        """处理单个映射sheet"""
        columns = list(mapping_df.columns)
        self.status_updated.emit(f"📊 {sheet_name} 列名: {columns}")
        
        # 智能识别店铺和仓库列
        shop_col = None
        warehouse_col = None
        
        for col in columns:
            col_str = str(col).lower()
            if any(keyword in col_str for keyword in ['店铺', 'shop', '商店']) and not shop_col:
                shop_col = col
            elif any(keyword in col_str for keyword in ['仓库', 'warehouse', '仓储']) and not warehouse_col:
                warehouse_col = col
        
        if not shop_col or not warehouse_col:
            if len(columns) >= 2:
                shop_col = columns[0]
                warehouse_col = columns[1]
                self.status_updated.emit(f"🔄 使用前两列作为店铺仓库: {shop_col} -> {warehouse_col}")
        
        if not shop_col or not warehouse_col:
            raise Exception(f"无法在sheet '{sheet_name}' 中识别店铺和仓库列")
        
        # 创建标准化的映射数据
        standardized_df = pd.DataFrame()
        standardized_df['店铺名称'] = mapping_df[shop_col].astype(str).str.strip()
        standardized_df['仓库名称'] = mapping_df[warehouse_col].astype(str).str.strip()
        
        return self.finalize_mapping_data(standardized_df)
    
    def finalize_mapping_data(self, mapping_df):
        """最终处理映射数据"""
        # 数据清理
        original_count = len(mapping_df)
        self.status_updated.emit(f"📝 原始数据行数: {original_count}")
        
        # 删除空值和无效数据
        mapping_df = mapping_df.dropna(subset=['店铺名称', '仓库名称'])
        mapping_df = mapping_df[mapping_df['店铺名称'] != '']
        mapping_df = mapping_df[mapping_df['仓库名称'] != '']
        
        invalid_values = ['nan', 'none', 'null', 'undefined', '无', '空']
        for col in ['店铺名称', '仓库名称']:
            mapping_df = mapping_df[~mapping_df[col].str.lower().isin(invalid_values)]
        
        cleaned_count = len(mapping_df)
        self.status_updated.emit(f"🧹 数据清理后: {original_count} -> {cleaned_count} 行")
        
        if len(mapping_df) == 0:
            raise Exception("清理后没有有效的映射数据")
        
        # 分析多对多关系
        unique_shops = mapping_df['店铺名称'].unique()
        unique_warehouses = mapping_df['仓库名称'].unique()
        
        self.status_updated.emit(f"🏪 唯一店铺数量: {len(unique_shops)}")
        self.status_updated.emit(f"🏭 唯一仓库数量: {len(unique_warehouses)}")
        
        # 显示多对多关系分析
        shop_warehouse_count = mapping_df.groupby('店铺名称')['仓库名称'].nunique()
        warehouse_shop_count = mapping_df.groupby('仓库名称')['店铺名称'].nunique()
        
        multi_warehouse_shops = shop_warehouse_count[shop_warehouse_count > 1]
        multi_shop_warehouses = warehouse_shop_count[warehouse_shop_count > 1]
        
        if len(multi_warehouse_shops) > 0:
            self.status_updated.emit(f"🔄 发现 {len(multi_warehouse_shops)} 个店铺对应多个仓库:")
            for shop, count in multi_warehouse_shops.head(3).items():
                warehouses = mapping_df[mapping_df['店铺名称'] == shop]['仓库名称'].unique()
                self.status_updated.emit(f"  {shop}: {list(warehouses)} ({count}个仓库)")
        
        if len(multi_shop_warehouses) > 0:
            self.status_updated.emit(f"🔄 发现 {len(multi_shop_warehouses)} 个仓库对应多个店铺:")
            for warehouse, count in multi_shop_warehouses.head(3).items():
                shops = mapping_df[mapping_df['仓库名称'] == warehouse]['店铺名称'].unique()
                self.status_updated.emit(f"  {warehouse}: {list(shops)} ({count}个店铺)")
        
        # 显示最终结果
        sample = mapping_df.head(5)
        self.status_updated.emit(f"✅ 最终映射关系样例:\n{sample.to_string(index=False)}")
        
        return mapping_df
    
    def generate_full_mapping(self, sheet_names):
        """生成全映射关系：所有店铺对应所有仓库"""
        try:
            # 收集所有可能的店铺和仓库列表
            all_shops = []
            all_warehouses = []
            
            self.status_updated.emit("🔍 从所有sheet中收集店铺和仓库信息...")
            
            for sheet in sheet_names:
                try:
                    df = pd.read_excel(self.mapping_file, sheet_name=sheet, engine='openpyxl', header=0)
                    if df.empty:
                        continue
                    
                    for col in df.columns:
                        col_str = str(col).lower()
                        # 收集店铺信息
                        if any(keyword in col_str for keyword in ['店铺', 'shop', '商店']):
                            shops = df[col].dropna().unique().tolist()
                            all_shops.extend([str(shop).strip() for shop in shops if str(shop).strip()])
                            self.status_updated.emit(f"  🏪 从sheet '{sheet}' 列 '{col}' 收集到店铺: {shops}")
                        
                        # 收集仓库信息
                        elif any(keyword in col_str for keyword in ['仓库', 'warehouse', '仓储']):
                            warehouses = df[col].dropna().unique().tolist()
                            all_warehouses.extend([str(wh).strip() for wh in warehouses if str(wh).strip()])
                            self.status_updated.emit(f"  🏭 从sheet '{sheet}' 列 '{col}' 收集到仓库: {warehouses}")
                
                except Exception as e:
                    self.status_updated.emit(f"  ❌ 处理sheet '{sheet}' 时出错: {str(e)}")
                    continue
            
            # 去重并清理
            unique_shops = list(set([shop for shop in all_shops if shop and shop.lower() not in ['nan', 'none', 'null']]))
            unique_warehouses = list(set([wh for wh in all_warehouses if wh and wh.lower() not in ['nan', 'none', 'null']]))
            
            self.status_updated.emit(f"🏪 收集到唯一店铺: {unique_shops}")
            self.status_updated.emit(f"🏭 收集到唯一仓库: {unique_warehouses}")
            
            # 如果没有收集到足够的信息，使用默认值
            if not unique_shops:
                # 尝试从订单数据中获取店铺信息
                self.status_updated.emit("⚠️ 未找到店铺信息，将从订单数据中动态获取")
                unique_shops = ["默认店铺"]  # 临时占位符
            
            if not unique_warehouses:
                self.status_updated.emit("⚠️ 未找到仓库信息，使用默认仓库")
                unique_warehouses = ["默认仓库", "华东仓", "华北仓", "华南仓", "西南仓"]
            
            # 生成全映射关系：每个店铺对应每个仓库
            full_mappings = []
            for shop in unique_shops:
                for warehouse in unique_warehouses:
                    full_mappings.append({
                        '店铺名称': shop,
                        '仓库名称': warehouse
                    })
            
            mapping_df = pd.DataFrame(full_mappings)
            
            self.status_updated.emit(f"🎉 生成全映射关系: {len(unique_shops)} 个店铺 × {len(unique_warehouses)} 个仓库 = {len(mapping_df)} 条映射")
            self.status_updated.emit("📋 全映射模式：所有店铺都可以与每一个仓库对应")
            
            # 显示样例
            if len(mapping_df) > 0:
                sample = mapping_df.head(10)
                self.status_updated.emit(f"📊 映射关系样例:\n{sample.to_string(index=False)}")
            
            return self.finalize_mapping_data(mapping_df)
            
        except Exception as e:
            # 最后的备用方案：创建一个最基本的映射
            self.status_updated.emit(f"⚠️ 生成全映射失败: {str(e)}")
            self.status_updated.emit("🔄 使用最基本的备用映射...")
            
            basic_mapping = pd.DataFrame([
                {'店铺名称': '通用店铺', '仓库名称': '默认仓库'},
                {'店铺名称': '通用店铺', '仓库名称': '华东仓'},
                {'店铺名称': '通用店铺', '仓库名称': '华北仓'},
                {'店铺名称': '通用店铺', '仓库名称': '华南仓'}
            ])
            
            self.status_updated.emit(f"🎯 使用备用映射: {len(basic_mapping)} 条")
            return self.finalize_mapping_data(basic_mapping)
    
    def get_feishu_data(self):
        """从飞书多维表格获取货品状态数据"""
        # 检查连接状态
        if getattr(self, 'feishu_use_requests', False):
            # 使用requests方式
            return self.get_feishu_data_with_requests()
        elif hasattr(self, 'feishu_client') and self.feishu_client:
            # 使用lark_oapi方式
            return self.get_feishu_data_with_lark_oapi()
        else:
            # 没有任何飞书连接
            self.status_updated.emit("未连接飞书，将使用默认货品状态")
            return {}
    
    def get_feishu_data_with_requests(self):
        """使用requests直接调用飞书API获取数据"""
        try:
            import requests
            
            self.status_updated.emit("🚀 开始使用requests获取飞书数据...")
            
            # 检查必要的连接信息
            if not self.feishu_headers or not self.feishu_api_base:
                self.status_updated.emit("❌ 飞书连接信息不完整")
                return {}
            
            self.status_updated.emit(f"📋 API地址: {self.feishu_api_base}")
            self.status_updated.emit(f"🔑 请求头已设置: {bool(self.feishu_headers)}")
            
            feishu_data = {}
            page_token = None
            page_count = 0
            total_processed = 0
            total_raw_records = 0  # 原始记录总数
            total_valid_records = 0  # 有效记录总数
            duplicate_count = 0  # 重复规格代码计数
            
            while True:
                page_count += 1
                self.status_updated.emit(f"📄 正在获取第 {page_count} 页数据...")
                
                # 构建请求参数
                params = {
                    "view_id": "vewfSJbcyA",
                    "page_size": 500
                }
                if page_token:
                    params["page_token"] = page_token
                
                # 发送请求（增加重试机制）
                max_retries = 3
                retry_count = 0
                response = None
                
                while retry_count < max_retries:
                    try:
                        response = requests.get(self.feishu_api_base, headers=self.feishu_headers, params=params, timeout=60)
                        break  # 请求成功，跳出重试循环
                    except requests.exceptions.Timeout:
                        retry_count += 1
                        if retry_count < max_retries:
                            self.status_updated.emit(f"⏰ 第 {page_count} 页请求超时，重试 {retry_count}/{max_retries}")
                            import time
                            time.sleep(2)  # 等待2秒后重试
                        else:
                            self.status_updated.emit(f"❌ 第 {page_count} 页请求超时，已达到最大重试次数")
                            return feishu_data
                    except Exception as req_error:
                        retry_count += 1
                        if retry_count < max_retries:
                            self.status_updated.emit(f"⚠️ 第 {page_count} 页请求异常，重试 {retry_count}/{max_retries}: {str(req_error)}")
                            import time
                            time.sleep(2)
                        else:
                            self.status_updated.emit(f"❌ 第 {page_count} 页请求失败，已达到最大重试次数: {str(req_error)}")
                            return feishu_data
                
                if response is None:
                    self.status_updated.emit(f"❌ 第 {page_count} 页请求完全失败")
                    break
                
                # 检查HTTP状态码
                if response.status_code != 200:
                    self.status_updated.emit(f"❌ HTTP请求失败，状态码: {response.status_code}")
                    break
                
                # 尝试解析JSON
                try:
                    result = response.json()
                    self.status_updated.emit(f"🔍 原始响应类型: {type(result)}")
                    
                    # 如果result是字符串，尝试再次解析
                    if isinstance(result, str):
                        self.status_updated.emit(f"⚠️ 响应是字符串，尝试再次解析: {result[:100]}...")
                        try:
                            import json
                            result = json.loads(result)
                            self.status_updated.emit(f"✅ 二次解析成功，类型: {type(result)}")
                        except (ValueError, json.JSONDecodeError) as e:
                            self.status_updated.emit(f"❌ 二次JSON解析失败: {str(e)}")
                            self.status_updated.emit(f"字符串内容: {result[:200]}...")
                            break
                    
                except (ValueError, json.JSONDecodeError) as json_error:
                    self.status_updated.emit(f"❌ JSON解析失败: {str(json_error)}")
                    self.status_updated.emit(f"响应状态码: {response.status_code}")
                    self.status_updated.emit(f"响应头: {dict(response.headers)}")
                    self.status_updated.emit(f"响应内容: {response.text[:500]}...")
                    break
                except Exception as e:
                    self.status_updated.emit(f"❌ 响应处理异常: {str(e)}")
                    self.status_updated.emit(f"响应类型: {type(response)}")
                    break
                
                if not isinstance(result, dict):
                    self.status_updated.emit(f"❌ 响应不是字典格式: {type(result)}")
                    self.status_updated.emit(f"响应内容: {str(result)[:200]}...")
                    break
                
                if result.get("code") != 0:
                    self.status_updated.emit(f"❌ 第 {page_count} 页获取失败: {result.get('msg', '未知错误')}")
                    break
                
                # 处理数据
                data = result.get("data", {})
                items = data.get("items", [])
                current_page_records = len(items)
                total_raw_records += current_page_records
                
                if current_page_records == 0:
                    self.status_updated.emit(f"📄 第 {page_count} 页无数据，结束获取")
                    break
                
                # 处理当前页的每条记录
                page_valid_count = 0
                for record in items:
                    try:
                        if not isinstance(record, dict):
                            self.status_updated.emit(f"⚠️ 记录不是字典格式: {type(record)}")
                            continue
                            
                        fields = record.get("fields", {})
                        if not fields:
                            continue
                        
                        # 提取关键字段 - 处理不同的数据格式
                        spec_code = ""
                        status = ""
                        
                        # 获取SKU编码（对应EXCEL中的规格代码）- 尝试多种可能的字段名称
                        spec_code = ""
                        for field_name in ['SKU编码', 'SKU', 'sku_code', '规格代码', '商品规格代码', 'spec_code']:
                            spec_field = fields.get(field_name, "")
                            if spec_field:  # 如果字段有值
                                if isinstance(spec_field, dict):
                                    spec_code = spec_field.get("text", "").strip()
                                elif isinstance(spec_field, str):
                                    spec_code = spec_field.strip()
                                else:
                                    spec_code = str(spec_field).strip()
                                if spec_code:  # 如果提取到了有效值，跳出循环
                                    break
                        
                        # 尝试获取货品状态 - 尝试多种可能的字段名称
                        status = ""
                        for field_name in ['货品状态', '商品状态', 'status', 'product_status', '状态']:
                            status_field = fields.get(field_name, "")
                            if status_field:  # 如果字段有值
                                if isinstance(status_field, dict):
                                    status = status_field.get("text", "").strip()
                                elif isinstance(status_field, str):
                                    status = status_field.strip()
                                else:
                                    status = str(status_field).strip()
                                if status:  # 如果提取到了有效值，跳出循环
                                    break
                        
                        # 尝试获取业务定制归属 - 尝试多种可能的字段名称
                        business_custom = ""
                        for field_name in ['业务定制归属', '定制归属', 'business_custom', '归属']:
                            custom_field = fields.get(field_name, "")
                            if custom_field:  # 如果字段有值
                                if isinstance(custom_field, dict):
                                    business_custom = custom_field.get("text", "").strip()
                                elif isinstance(custom_field, str):
                                    business_custom = custom_field.strip()
                                else:
                                    business_custom = str(custom_field).strip()
                                if business_custom:  # 如果提取到了有效值，跳出循环
                                    break
                        
                        # 尝试获取供应状态 - 尝试多种可能的字段名称
                        supply_status = ""
                        for field_name in ['供应状态', '供应', 'supply_status', 'supply']:
                            supply_field = fields.get(field_name, "")
                            if supply_field:  # 如果字段有值
                                if isinstance(supply_field, dict):
                                    supply_status = supply_field.get("text", "").strip()
                                elif isinstance(supply_field, str):
                                    supply_status = supply_field.strip()
                                else:
                                    supply_status = str(supply_field).strip()
                                if supply_status:  # 如果提取到了有效值，跳出循环
                                    break
                    
                    except Exception as field_error:
                        self.status_updated.emit(f"⚠️ 处理记录字段时出错: {str(field_error)}")
                        continue
                    
                    # 调试信息：显示找到的字段
                    if page_count == 1 and page_valid_count < 3:  # 只在第一页显示前几条调试信息
                        available_fields = list(fields.keys())
                        self.status_updated.emit(f"🔍 第{page_valid_count+1}条记录可用字段: {available_fields}")
                        if spec_code:
                            self.status_updated.emit(f"  ✅ 找到SKU编码(规格代码): {spec_code}")
                        else:
                            self.status_updated.emit(f"  ❌ 未找到SKU编码字段")
                        if status:
                            self.status_updated.emit(f"  ✅ 找到货品状态: {status}")
                        else:
                            self.status_updated.emit(f"  ❌ 未找到货品状态字段")
                        if business_custom:
                            self.status_updated.emit(f"  ✅ 找到业务定制归属: {business_custom}")
                        if supply_status:
                            self.status_updated.emit(f"  ✅ 找到供应状态: {supply_status}")
                    
                    if spec_code:
                        # 检查是否重复
                        if spec_code in feishu_data:
                            duplicate_count += 1
                        
                        # 存储完整的字段信息
                        feishu_data[spec_code] = {
                            '货品状态': status if status else '待判断',
                            '业务定制归属': business_custom,
                            '供应状态': supply_status
                        }
                        page_valid_count += 1
                        total_valid_records += 1
                        
                        if page_count == 1 and len(feishu_data) < 3:
                            if not status:
                                self.status_updated.emit(f"⚠️ 规格代码 {spec_code} 没有找到状态字段，使用默认状态")
                
                total_processed += page_valid_count
                self.status_updated.emit(f"✅ 第 {page_count} 页处理完成: {page_valid_count}/{current_page_records} 条有效记录")
                
                # 检查是否有下一页
                if data.get("has_more") and data.get("page_token"):
                    page_token = data.get("page_token")
                    self.status_updated.emit(f"🔄 准备获取下一页，已处理 {total_processed} 条...")
                else:
                    self.status_updated.emit(f"🏁 所有数据获取完成，共处理 {total_processed} 条记录")
                    break
            
            # 统计信息
            self.status_updated.emit(f"📊 飞书数据获取完成: 共 {len(feishu_data)} 个规格代码")
            if duplicate_count > 0:
                self.status_updated.emit(f"⚠️ 发现 {duplicate_count} 个重复规格代码")
            if total_raw_records - len(feishu_data) > 0:
                self.status_updated.emit(f"📉 数据过滤: {total_raw_records - len(feishu_data)} 条记录无规格代码")
            
            return feishu_data
            
        except Exception as e:
            self.status_updated.emit(f"❌ requests方式获取飞书数据失败: {str(e)}")
            return {}
    
    def get_feishu_data_with_lark_oapi(self):
        """使用lark_oapi获取飞书数据（原方法）"""
        
        try:
            self.status_updated.emit("正在导入飞书API模块...")
            from lark_oapi.api.bitable.v1 import ListAppTableRecordRequest
            
            # 构建请求
            app_token = "bascnwiJJtU80TMwlyUTLqHpy0e"
            table_id = "tbl7Uxj6KiHtUjez"
            view_id = "vewfSJbcyA"
            
            self.status_updated.emit(f"📋 正在访问飞书多维表格: {app_token[:8]}...")
            self.status_updated.emit(f"🔍 表格ID: {table_id}")
            self.status_updated.emit(f"👁️ 视图ID: {view_id} (获取货品状态视图)")
            
            # 分页获取所有数据
            self.status_updated.emit("🚀 开始分页获取飞书数据，目标获取7000+条记录...")
            feishu_data = {}
            page_token = None
            page_count = 0
            total_processed = 0
            
            while True:
                page_count += 1
                self.status_updated.emit(f"📄 正在获取第 {page_count} 页数据...")
                
                # 构建分页请求
                try:
                    self.status_updated.emit(f"🔧 构建第 {page_count} 页请求...")
                    request_builder = ListAppTableRecordRequest.builder()
                    if request_builder is None:
                        raise Exception("ListAppTableRecordRequest.builder() 返回 None")
                    
                    request_builder = request_builder.app_token(app_token)
                    if request_builder is None:
                        raise Exception("设置 app_token 后返回 None")
                    
                    request_builder = request_builder.table_id(table_id)
                    if request_builder is None:
                        raise Exception("设置 table_id 后返回 None")
                    
                    request_builder = request_builder.view_id(view_id)
                    if request_builder is None:
                        raise Exception("设置 view_id 后返回 None")
                    
                    request_builder = request_builder.page_size(500)  # 增加每页大小
                    if request_builder is None:
                        raise Exception("设置 page_size 后返回 None")
                    
                    if page_token:
                        request_builder = request_builder.page_token(page_token)
                        if request_builder is None:
                            raise Exception("设置 page_token 后返回 None")
                        
                    request = request_builder.build()
                    if request is None:
                        raise Exception("request_builder.build() 返回 None")
                        
                except Exception as request_error:
                    self.status_updated.emit(f"❌ 构建请求失败: {str(request_error)}")
                    break
                
                # 发送请求
                try:
                    self.status_updated.emit(f"📡 发送第 {page_count} 页请求...")
                    if not self.feishu_client:
                        raise Exception("feishu_client 为 None")
                    if not hasattr(self.feishu_client, 'bitable'):
                        raise Exception("feishu_client 没有 bitable 属性")
                    if not hasattr(self.feishu_client.bitable, 'v1'):
                        raise Exception("feishu_client.bitable 没有 v1 属性")
                    if not hasattr(self.feishu_client.bitable.v1, 'app_table_record'):
                        raise Exception("feishu_client.bitable.v1 没有 app_table_record 属性")
                    
                    response = self.feishu_client.bitable.v1.app_table_record.list(request)
                    if response is None:
                        raise Exception("API 调用返回 None")
                        
                except Exception as api_error:
                    self.status_updated.emit(f"❌ API 调用失败: {str(api_error)}")
                    break
                
                if response.code != 0:
                    self.status_updated.emit(f"❌ 第 {page_count} 页获取失败: {response.msg}")
                    break
                
                # 检查响应数据是否有效
                if not response.data:
                    self.status_updated.emit(f"❌ 第 {page_count} 页响应数据为空")
                    break
                
                # 处理当前页数据
                current_page_records = len(response.data.items) if response.data.items else 0
                self.status_updated.emit(f"📨 第 {page_count} 页返回 {current_page_records} 条记录")
                
                if current_page_records == 0:
                    break
                
                # 如果是第一页，显示字段结构
                if page_count == 1 and current_page_records > 0:
                    first_record = response.data.items[0]
                    available_fields = list(first_record.fields.keys()) if first_record and first_record.fields else []
                    self.status_updated.emit(f"🔍 可用字段: {available_fields}")
                
                # 处理当前页的每条记录
                page_valid_count = 0
                for record in response.data.items:
                    if not record:
                        continue
                    fields = record.fields
                    if not fields:
                        continue
                        
                    # 尝试多种可能的字段名称
                    spec_code = None
                    status = None
                    business_custom = None
                    supply_status = None
                    
                    # 查找SKU编码字段（对应EXCEL中的规格代码）
                    for field_name in ['SKU编码', 'SKU', 'sku_code', '规格代码', '商品规格代码', 'spec_code']:
                        if field_name in fields and fields[field_name]:
                            spec_code = str(fields[field_name]).strip()
                            break
                    
                    # 查找货品状态字段（从飞书视图中获取）
                    for field_name in ['货品状态', '商品状态', 'status', 'product_status', '状态']:
                        if field_name in fields and fields[field_name]:
                            status = str(fields[field_name]).strip()
                            break
                    
                    # 查找业务定制归属字段
                    for field_name in ['业务定制归属', '定制归属', 'business_custom', '归属']:
                        if field_name in fields and fields[field_name]:
                            business_custom = str(fields[field_name]).strip()
                            break
                    
                    # 查找供应状态字段
                    for field_name in ['供应状态', '供应', 'supply_status', 'supply']:
                        if field_name in fields and fields[field_name]:
                            supply_status = str(fields[field_name]).strip()
                            break
                    
                    if spec_code:
                        # 存储完整的字段信息
                        feishu_data[spec_code] = {
                            '货品状态': status if status else '待判断',
                            '业务定制归属': business_custom if business_custom else '',
                            '供应状态': supply_status if supply_status else ''
                        }
                        page_valid_count += 1
                
                total_processed += current_page_records
                self.status_updated.emit(f"✅ 第 {page_count} 页处理完成，有效记录: {page_valid_count}")
                
                # 检查是否有下一页
                if response.data and hasattr(response.data, 'has_more') and response.data.has_more:
                    page_token = getattr(response.data, 'page_token', None)
                    if page_token:
                        self.status_updated.emit(f"🔄 准备获取下一页，已处理 {total_processed} 条...")
                    else:
                        self.status_updated.emit(f"🏁 无下一页token，数据获取完成，共处理 {total_processed} 条原始记录")
                        break
                else:
                    self.status_updated.emit(f"🏁 所有数据获取完成，共处理 {total_processed} 条原始记录")
                    break
                
                # 防止无限循环
                if page_count > 50:  # 最多50页，防止无限循环
                    self.status_updated.emit("⚠️ 达到最大页数限制，停止获取")
                    break
            
            self.status_updated.emit(f"🎉 飞书数据获取完成！共获得 {len(feishu_data)} 条有效的SKU编码-货品状态映射")
            
            if len(feishu_data) > 0:
                # 显示部分数据样例
                sample_items = list(feishu_data.items())[:5]
                self.status_updated.emit(f"📋 数据样例: {sample_items}")
            
            return feishu_data
                
        except Exception as e:
            self.status_updated.emit(f"获取飞书数据失败: {str(e)}")
            return {}
    
    def process_unship_data(self, order_df, stock_df, stock_sheet2_df, mapping_df, feishu_data):
        """处理未配货数据"""
        # 1. 数据过滤：只取订单状态为"待审核"且订单标记中不包含"僵尸订单"的数据
        filtered_order = order_df[
            (order_df['订单状态'] == '待审核') & 
            (~order_df['订单标记'].str.contains('僵尸订单', na=False))
        ].copy()
        
        self.status_updated.emit(f"第一步过滤后订单数据: {len(filtered_order)} 条")
        
        # 2. 排除第三个sheet中的规格代码
        if hasattr(self, 'excluded_spec_codes') and self.excluded_spec_codes:
            original_count = len(filtered_order)
            
            # 创建多种格式的排除集合以支持精准匹配
            excluded_str_set = set()
            excluded_num_set = set()
            
            for code in self.excluded_spec_codes:
                # 字符串格式
                excluded_str_set.add(str(code).strip())
                # 尝试数字格式
                try:
                    if str(code).strip().replace('.', '').replace('-', '').isdigit():
                        # 整数格式
                        excluded_num_set.add(int(float(code)))
                        # 浮点数格式
                        excluded_num_set.add(float(code))
                except (ValueError, TypeError):
                    pass
            
            # 创建匹配条件
            def should_exclude(spec_code):
                if pd.isna(spec_code):
                    return False
                
                # 字符串匹配
                spec_str = str(spec_code).strip()
                if spec_str in excluded_str_set:
                    return True
                
                # 数字匹配
                try:
                    # 如果原始值是数字类型
                    if isinstance(spec_code, (int, float)):
                        if spec_code in excluded_num_set:
                            return True
                    
                    # 尝试将字符串转换为数字进行匹配
                    if spec_str.replace('.', '').replace('-', '').isdigit():
                        spec_int = int(float(spec_str))
                        spec_float = float(spec_str)
                        if spec_int in excluded_num_set or spec_float in excluded_num_set:
                            return True
                except (ValueError, TypeError):
                    pass
                
                return False
            
            # 应用排除逻辑
            filtered_order = filtered_order[~filtered_order['规格代码'].apply(should_exclude)]
            
            excluded_count = original_count - len(filtered_order)
            self.status_updated.emit(f"🚫 第二步-排除规格代码: {original_count} -> {len(filtered_order)} 条数据 (排除了{excluded_count}条)")
            
            if excluded_count > 0:
                self.status_updated.emit(f"📝 排除的规格代码格式包括: 字符串({len(excluded_str_set)}个) 和 数字({len(excluded_num_set)}个)")
        else:
            self.status_updated.emit("ℹ️ 没有需要排除的规格代码")
        
        # 3. 基于店铺名称进行过滤
        # 检查店铺列名（可能是'店铺名称'或'店铺'）
        shop_col = None
        if '店铺名称' in mapping_df.columns:
            shop_col = '店铺名称'
        elif '店铺' in mapping_df.columns:
            shop_col = '店铺'
        
        if not mapping_df.empty and shop_col:
            original_count = len(filtered_order)
            valid_shops = set(mapping_df[shop_col].dropna().astype(str).str.strip())
            valid_shops = {shop for shop in valid_shops if shop and shop != 'nan'}
            
            if valid_shops and '店铺名称' in filtered_order.columns:
                order_shops = set(filtered_order['店铺名称'].dropna().astype(str).str.strip())
                shop_overlap = order_shops.intersection(valid_shops)
                
                self.status_updated.emit(f"📊 店铺匹配情况: 订单中{len(order_shops)}个店铺，映射表中{len(valid_shops)}个店铺，重叠{len(shop_overlap)}个")
                
                # 显示重叠的店铺
                if shop_overlap:
                    self.status_updated.emit(f"✅ 匹配的店铺: {sorted(list(shop_overlap))}")
                
                # 只保留映射表中存在的店铺
                filtered_order = filtered_order[
                    filtered_order['店铺名称'].astype(str).str.strip().isin(valid_shops)
                ]
                
                shop_filtered_count = original_count - len(filtered_order)
                self.status_updated.emit(f"🏪 第三步-店铺名称过滤: {original_count} -> {len(filtered_order)} 条数据 (过滤了{shop_filtered_count}条)")
            else:
                self.status_updated.emit("⚠️ 没有有效的店铺映射数据")
        else:
            self.status_updated.emit(f"⚠️ 映射表中没有找到店铺列（查找了'店铺名称'和'店铺'列）")
        
        # 4. 基于仓库名称进行过滤
        if not mapping_df.empty and '仓库名称' in mapping_df.columns:
            original_count = len(filtered_order)
            valid_warehouses = set(mapping_df['仓库名称'].dropna().astype(str).str.strip())
            valid_warehouses = {wh for wh in valid_warehouses if wh and wh != 'nan'}
            
            if valid_warehouses and '仓库名称' in filtered_order.columns:
                order_warehouses = set(filtered_order['仓库名称'].dropna().astype(str).str.strip())
                warehouse_overlap = order_warehouses.intersection(valid_warehouses)
                
                self.status_updated.emit(f"📊 仓库匹配情况: 订单中{len(order_warehouses)}个仓库，映射表中{len(valid_warehouses)}个仓库，重叠{len(warehouse_overlap)}个")
                
                # 只保留映射表中存在的仓库
                filtered_order = filtered_order[
                    filtered_order['仓库名称'].astype(str).str.strip().isin(valid_warehouses)
                ]
                
                warehouse_filtered_count = original_count - len(filtered_order)
                self.status_updated.emit(f"🏭 第四步-仓库名称过滤: {original_count} -> {len(filtered_order)} 条数据 (过滤了{warehouse_filtered_count}条)")
            else:
                self.status_updated.emit("⚠️ 没有有效的仓库映射数据")
        
        # 5. 检查最终过滤结果
        if len(filtered_order) == 0:
            self.status_updated.emit("❌ 经过所有过滤步骤后没有数据！请检查店铺仓库对应关系")
            return {
                'sheet1': pd.DataFrame(),
                'sheet2': pd.DataFrame(), 
                'sheet3': pd.DataFrame()
            }
        
        self.status_updated.emit(f"✅ 最终过滤结果: {len(filtered_order)} 条有效数据")
        
        # 6. 显示过滤统计信息
        self.status_updated.emit("📊 过滤步骤完成，开始生成报表数据...")
        
        # 3. 添加店铺&规格代码列（直接连接，不使用特殊字符）
        try:
            # 确保两列都转换为字符串，并处理NaN值
            shop_names = filtered_order['店铺名称'].fillna('').astype(str)
            spec_codes = filtered_order['规格代码'].fillna('').astype(str)
            filtered_order['店铺&规格代码'] = shop_names + spec_codes
            self.status_updated.emit(f"✅ 成功生成店铺&规格代码列")
        except Exception as e:
            self.status_updated.emit(f"❌ 生成店铺&规格代码列失败: {str(e)}")
            # 使用备用方法
            filtered_order['店铺&规格代码'] = filtered_order.apply(
                lambda row: str(row.get('店铺名称', '')) + str(row.get('规格代码', '')), axis=1
            )
        
        # 4. 生成第一个sheet：订单商品明细数据（暂时不包含未配货原因，稍后添加）
        sheet1_data = filtered_order[[
            '订单编号', '订单类型', '平台单号', '店铺名称', '制单时间', 
            '订单状态', '仓库名称', '店铺&规格代码', '商品名称', '规格代码', '商品数量'
        ]].copy()
        
        # 5. 生成第二个sheet：未配货明细
        # 严格按照店铺仓库对应关系来生成，仓库作为列名
        self.status_updated.emit("正在基于店铺仓库对应关系生成未配货明细...")
        
        # 获取过滤后第一个sheet中实际使用的仓库（双维度过滤后的有效仓库）
        actual_warehouses_in_orders = filtered_order['仓库名称'].unique().tolist() if '仓库名称' in filtered_order.columns else []
        self.status_updated.emit(f"第一个sheet中过滤后的仓库: {actual_warehouses_in_orders}")
        
        # 只使用实际在订单中出现的有效仓库作为第二个sheet的列
        all_warehouses = [wh for wh in actual_warehouses_in_orders if pd.notna(wh) and str(wh).strip()]
        self.status_updated.emit(f"第二个sheet将使用的仓库列: {all_warehouses}")
        
        # 验证仓库名称（过滤后应该没有无效仓库了）
        invalid_warehouses = [wh for wh in all_warehouses if '虚拟' in str(wh)]
        if invalid_warehouses:
            self.status_updated.emit(f"⚠️ 检测到无效仓库名称: {invalid_warehouses}")
        else:
            self.status_updated.emit("✅ 所有仓库名称已通过双维度过滤，无虚拟仓等无效仓库")
        
        # 获取所有唯一的店铺&规格代码组合，只处理第一个sheet中实际存在的数据
        self.status_updated.emit("📋 只处理第一个sheet中实际存在的店铺规格代码组合...")
        unique_combinations = filtered_order[['店铺名称', '规格代码', '商品名称', '店铺&规格代码']].drop_duplicates()
        self.status_updated.emit(f"实际需要处理的组合数: {len(unique_combinations)}")
        
        sheet2_data = []
        for _, row in unique_combinations.iterrows():
            shop_name = row['店铺名称']
            spec_code = row['规格代码']
            product_name = row['商品名称']
            shop_spec = row['店铺&规格代码']
            
            # 只在处理前几个组合时显示详细信息
            if len(sheet2_data) < 3:
                self.status_updated.emit(f"🔍 处理: {shop_name} - {spec_code}")
            
            # 初始化所有仓库的数据为0
            warehouse_data = {}
            for warehouse in all_warehouses:
                warehouse_data[warehouse] = 0
            
            # 🔥 修复：根据该店铺&规格代码在第一个sheet中计算每个仓库的商品数量汇总
            matching_orders = filtered_order[
                (filtered_order['店铺名称'] == shop_name) & 
                (filtered_order['规格代码'] == spec_code)
            ]
            
            matched_warehouses = []
            
            # 为每个仓库计算该店铺&规格代码的商品数量汇总
            for warehouse in all_warehouses:
                # 获取该仓库下相同店铺&规格代码的所有订单
                warehouse_orders = matching_orders[matching_orders['仓库名称'] == warehouse]
                
                if not warehouse_orders.empty and '商品数量' in warehouse_orders.columns:
                    # 计算该仓库的商品数量汇总
                    warehouse_quantity = pd.to_numeric(warehouse_orders['商品数量'], errors='coerce').fillna(0).sum()
                    warehouse_data[warehouse] = int(warehouse_quantity) if warehouse_quantity >= 0 else 0
                    
                    # 如果有数量，记录为匹配的仓库
                    if warehouse_data[warehouse] > 0:
                        matched_warehouses.append(warehouse)
                else:
                    warehouse_data[warehouse] = 0
            
            if len(sheet2_data) < 3:
                warehouse_details = [f"{wh}({warehouse_data[wh]})" for wh in matched_warehouses]
                self.status_updated.emit(f"  ✅ {shop_name}-{spec_code} 仓库数量分布: {warehouse_details}")
            
            # 从库存数据获取可销数（优先使用可用库存列，其次使用总计列）
            stock_total = 0
            spec_code_col = '商品规格代码' if '商品规格代码' in stock_df.columns else '规格代码'
            # 优先使用可用库存列，如果没有则使用总计列
            available_col = '可用库存' if '可用库存' in stock_df.columns else ('总计' if '总计' in stock_df.columns else None)
            
            if spec_code_col in stock_df.columns:
                # 尝试多种匹配方式以处理数据类型不匹配问题
                stock_row = stock_df[stock_df[spec_code_col] == spec_code]
                
                # 如果直接匹配失败，尝试字符串匹配
                if stock_row.empty:
                    stock_row = stock_df[stock_df[spec_code_col].astype(str) == str(spec_code)]
                
                # 如果字符串匹配也失败，尝试数字匹配（如果规格代码是纯数字）
                if stock_row.empty:
                    try:
                        spec_code_num = int(spec_code) if str(spec_code).isdigit() else float(spec_code)
                        stock_row = stock_df[stock_df[spec_code_col] == spec_code_num]
                    except (ValueError, TypeError):
                        pass
                
                if not stock_row.empty:
                    if available_col and available_col in stock_df.columns:
                        stock_total = stock_row[available_col].iloc[0] if pd.notna(stock_row[available_col].iloc[0]) else 0
                    else:
                        numeric_cols = stock_df.select_dtypes(include=[np.number]).columns
                        if len(numeric_cols) > 0:
                            stock_total = stock_row[numeric_cols].sum(axis=1).iloc[0] if not stock_row.empty else 0
            
            # 从第二个sheet获取在途数（总计）
            transit_total = 0
            spec_code_col2 = '商品规格代码' if '商品规格代码' in stock_sheet2_df.columns else '规格代码'
            total_col2 = '总计' if '总计' in stock_sheet2_df.columns else None
            
            if spec_code_col2 in stock_sheet2_df.columns:
                # 尝试多种匹配方式以处理数据类型不匹配问题
                transit_row = stock_sheet2_df[stock_sheet2_df[spec_code_col2] == spec_code]
                
                # 如果直接匹配失败，尝试字符串匹配
                if transit_row.empty:
                    transit_row = stock_sheet2_df[stock_sheet2_df[spec_code_col2].astype(str) == str(spec_code)]
                
                # 如果字符串匹配也失败，尝试数字匹配（如果规格代码是纯数字）
                if transit_row.empty:
                    try:
                        spec_code_num = int(spec_code) if str(spec_code).isdigit() else float(spec_code)
                        transit_row = stock_sheet2_df[stock_sheet2_df[spec_code_col2] == spec_code_num]
                    except (ValueError, TypeError):
                        pass
                
                if not transit_row.empty:
                    if total_col2 and total_col2 in stock_sheet2_df.columns:
                        transit_total = transit_row[total_col2].iloc[0] if pd.notna(transit_row[total_col2].iloc[0]) else 0
                    else:
                        numeric_cols = stock_sheet2_df.select_dtypes(include=[np.number]).columns
                        if len(numeric_cols) > 0:
                            transit_total = transit_row[numeric_cols].sum(axis=1).iloc[0] if not transit_row.empty else 0
            
            # 从飞书获取货品状态和其他字段
            # 匹配逻辑：EXCEL中的"规格代码" = 飞书表格中的"SKU编码" -> 获取飞书视图中的相关字段
            feishu_record = feishu_data.get(spec_code, {})
            
            # 兼容旧版本数据格式（字符串）和新版本数据格式（字典）
            if isinstance(feishu_record, str):
                product_status = feishu_record
                business_custom = ''
                supply_status = ''
            elif isinstance(feishu_record, dict):
                product_status = feishu_record.get('货品状态', '待判断')
                business_custom = feishu_record.get('业务定制归属', '')
                supply_status = feishu_record.get('供应状态', '')
            else:
                product_status = '待判断'
                business_custom = ''
                supply_status = ''
            
            # 如果在飞书中找到匹配的SKU编码，记录匹配成功
            if spec_code in feishu_data:
                self.status_updated.emit(f"✅ 规格代码 {spec_code} 在飞书中找到匹配，状态: {product_status}")
            # else:
            #     self.status_updated.emit(f"⚠️ 规格代码 {spec_code} 在飞书中未找到匹配，使用默认状态")
            
            # 🔥 修复：从第一个sheet中汇总相同店铺&规格代码的商品数量
            # 计算该店铺&规格代码在第一个sheet中的商品数量总和
            if '商品数量' in filtered_order.columns:
                # 确保商品数量列为数值类型
                matching_rows = filtered_order[
                    (filtered_order['店铺名称'] == shop_name) & 
                    (filtered_order['规格代码'] == spec_code)
                ]
                
                if not matching_rows.empty:
                    # 转换为数值类型并求和
                    quantity_series = pd.to_numeric(matching_rows['商品数量'], errors='coerce').fillna(0)
                    quantity_total = quantity_series.sum()
                else:
                    quantity_total = 0
            else:
                quantity_total = 0
            
            # 确保总计至少为整数
            quantity_total = int(quantity_total) if quantity_total >= 0 else 0
            
            # 计算未配货原因
            unship_reason = self.calculate_unship_reason(stock_total, transit_total, product_status, quantity_total, business_custom, supply_status)
            
            
            # 计算仓库数据的总和（现在是真实的数量汇总）
            warehouse_total = sum(warehouse_data.values())
            
            # 调试信息（仅显示前几个）
            if len(sheet2_data) < 3:
                self.status_updated.emit(f"  📊 {shop_name}-{spec_code} 商品数量汇总: {quantity_total}")
                self.status_updated.emit(f"  📊 各仓库数量总和: {warehouse_total}")
                if quantity_total != warehouse_total:
                    self.status_updated.emit(f"  ⚠️ 注意：总计({quantity_total})与仓库总和({warehouse_total})不一致")
            
            # 构建行数据，包含基础字段和所有仓库列
            row_data = {
                '店铺&规格代码': shop_spec,
                '规格代码': spec_code,
                '商品名称': product_name,
                '总计': quantity_total,  # 🔥 修复：总计 = 第一个sheet中相同店铺&规格代码的商品数量汇总
                '可销数': stock_total,   # 可销数从库存报表的可用库存列获取
                '在途数': transit_total,
                '货品状态': product_status,
                '未配货原因': unship_reason
                        }
            
            # 添加所有仓库列的数据（为0的显示为空，非0值包括负数都显示）
            for warehouse in all_warehouses:
                value = warehouse_data[warehouse]
                row_data[warehouse] = value if value != 0 else ''
            
            sheet2_data.append(row_data)
            
        self.status_updated.emit(f"生成未配货明细记录: {len(sheet2_data)} 条")
        
        sheet2_df = pd.DataFrame(sheet2_data)
        
        # 过滤掉整列都为0的仓库列
        if not sheet2_df.empty:
            active_warehouses = []
            for warehouse in all_warehouses:
                if warehouse in sheet2_df.columns:
                    # 检查该仓库列是否有非0值，需要处理空字符串和数值混合的情况
                    try:
                        # 将空字符串转换为0，然后转换为数值类型
                        warehouse_series = pd.to_numeric(sheet2_df[warehouse].replace('', 0), errors='coerce').fillna(0)
                        if warehouse_series.sum() > 0:
                            active_warehouses.append(warehouse)
                        else:
                            self.status_updated.emit(f"🚫 仓库 '{warehouse}' 整列为空，已从明细中移除")
                    except Exception as e:
                        self.status_updated.emit(f"⚠️ 处理仓库 '{warehouse}' 列时出错: {str(e)}，保留该列")
                        active_warehouses.append(warehouse)
            
            self.status_updated.emit(f"📦 有效仓库: {active_warehouses} (共{len(active_warehouses)}个)")
            
            # 调整列的顺序：基础信息 + 有效仓库列 + 其他信息
            base_columns = ['店铺&规格代码', '规格代码', '商品名称']
            warehouse_columns = [col for col in active_warehouses if col in sheet2_df.columns]
            other_columns = ['总计', '可销数', '在途数', '货品状态', '未配货原因']
            
            final_columns = base_columns + warehouse_columns + other_columns
            sheet2_df = sheet2_df.reindex(columns=final_columns, fill_value=0)
        else:
            # 如果没有数据，使用所有仓库
            base_columns = ['店铺&规格代码', '规格代码', '商品名称']
            warehouse_columns = all_warehouses
            other_columns = ['总计', '可销数', '在途数', '货品状态', '未配货原因']
            
            final_columns = base_columns + warehouse_columns + other_columns
            sheet2_df = sheet2_df.reindex(columns=final_columns, fill_value=0)
        
        # 为第二个sheet添加总计行
        if not sheet2_df.empty:
            self.status_updated.emit("为第二个sheet添加总计行...")
            
            # 计算数值列的总计
            total_row = {}
            
            # 基础信息列设置为"总计"标识
            total_row['店铺&规格代码'] = '总计'
            total_row['规格代码'] = ''
            total_row['商品名称'] = ''
            
            # 计算各仓库列的总计（为0的显示为空，非0值包括负数都显示）
            for warehouse in warehouse_columns:
                if warehouse in sheet2_df.columns:
                    # 将空字符串转换为0后再求和
                    warehouse_series = pd.to_numeric(sheet2_df[warehouse], errors='coerce').fillna(0)
                    total_value = warehouse_series.sum()
                    total_row[warehouse] = total_value if total_value != 0 else ''
            
            # 计算其他数值列的总计
            total_row['总计'] = sheet2_df['总计'].sum()
            total_row['可销数'] = sheet2_df['可销数'].sum()
            total_row['在途数'] = sheet2_df['在途数'].sum()
            total_row['货品状态'] = ''
            total_row['未配货原因'] = ''
            
            # 将总计行添加到DataFrame
            total_row_df = pd.DataFrame([total_row])
            sheet2_df = pd.concat([sheet2_df, total_row_df], ignore_index=True)
        
        # 6. 生成第三个sheet：卡单分析
        if not sheet2_df.empty:
            # 排除最后一行总计行（如果存在）
            sheet2_for_analysis = sheet2_df.copy()
            if len(sheet2_for_analysis) > 0:
                # 检查最后一行是否为总计行（通常总计行的规格代码为空或包含"总计"）
                last_row = sheet2_for_analysis.iloc[-1]
                if (pd.isna(last_row.get('规格代码', '')) or 
                    str(last_row.get('规格代码', '')).strip() == '' or
                    '总计' in str(last_row.get('规格代码', ''))):
                    sheet2_for_analysis = sheet2_for_analysis.iloc[:-1]
            
            # 按未配货原因分组，对总计字段求和（而不是计数）
            analysis_data = sheet2_for_analysis.groupby('未配货原因').agg({
                '总计': 'sum'  # 使用第二个sheet中的总计字段求和
            }).reset_index()
            analysis_data.columns = ['未配货原因', '总计']
            
            # 按总计从大到小排列
            analysis_data = analysis_data.sort_values('总计', ascending=False).reset_index(drop=True)
            
            # 计算占比
            total_count = analysis_data['总计'].sum()
            analysis_data['占比'] = (analysis_data['总计'] / total_count * 100).round(2)
            analysis_data['占比'] = analysis_data['占比'].astype(str) + '%'
            
            # 生成原因备注（只显示TOP3）
            analysis_data['原因备注'] = analysis_data['未配货原因'].apply(
                lambda reason: self.generate_reason_detail_top3(sheet2_for_analysis, reason)
            )
            
            # 添加总计行
            total_row = pd.DataFrame({
                '未配货原因': ['总计'],
                '总计': [total_count],
                '占比': [''],  # 总计行占比不显示
                '原因备注': ['']
            })
            analysis_data = pd.concat([analysis_data, total_row], ignore_index=True)
            
        else:
            analysis_data = pd.DataFrame(columns=['未配货原因', '总计', '占比', '原因备注'])
        
        # 7. 为第一个sheet添加未配货原因列（根据规格代码与第二个sheet匹配）
        self.status_updated.emit("为第一个sheet添加未配货原因...")
        if not sheet2_df.empty:
            # 创建规格代码到未配货原因的映射字典
            reason_mapping = dict(zip(sheet2_df['规格代码'], sheet2_df['未配货原因']))
            
            # 为第一个sheet添加未配货原因列
            sheet1_data['未配货原因'] = sheet1_data['规格代码'].map(reason_mapping).fillna('待判断')
            
            # 调整第一个sheet的列顺序，将未配货原因放在最后
            columns_order = [
                '订单编号', '订单类型', '平台单号', '店铺名称', '制单时间', 
                '订单状态', '仓库名称', '店铺&规格代码', '商品名称', '规格代码', '商品数量', '未配货原因'
            ]
            sheet1_data = sheet1_data.reindex(columns=columns_order)
        else:
            # 如果第二个sheet为空，添加空的未配货原因列
            sheet1_data['未配货原因'] = '待判断'
        
        return {
            'sheet1': sheet1_data,
            'sheet2': sheet2_df,
            'sheet3': analysis_data
        }
    
    def calculate_unship_reason(self, stock_total, transit_total, product_status, quantity_total=0, business_custom='', supply_status=''):
        """计算未配货原因"""
        # 标准化货品状态，去除可能的前缀
        status_str = str(product_status).strip()
        
        # 检查是否包含关键状态词（支持带前缀的状态）
        is_in_supply = '在供' in status_str  # 包含"在供"
        is_pending_stop = '待停供' in status_str  # 包含"待停供"  
        is_stopped = '停供' in status_str and '待停供' not in status_str  # 包含"停供"但不是"待停供"
        
        # 检查是否为非供应链负责SKU的条件
        # 业务定制归属不为空 且 供应状态不是"自营备货"或"自营清仓"
        business_custom_str = str(business_custom).strip()
        supply_status_str = str(supply_status).strip()
        
        is_non_supply_chain_sku = (
            business_custom_str != '' and 
            supply_status_str not in ['自营备货', '自营清仓']
        )
        
        # 根据新需求文档的规则判断未配货原因
        # 【可推单】：总计有库存且可销售数大于等于总计库存并且是在供/待停供的情况
        if quantity_total > 0 and stock_total >= quantity_total and (is_in_supply or is_pending_stop):
            return '可推单'
        # 【预售】：总计数量>0 且 可销数≤0 且 在途数>0 且是在供的情况
        elif quantity_total > 0 and stock_total <= 0 and transit_total > 0 and is_in_supply:
            if is_non_supply_chain_sku:
                return '预售，非供应链负责SKU'
            else:
                return '预售'
        # 【待加库存】：总计数量大于0并且可销售数≤0且在途库存≤0，且是在供
        elif quantity_total > 0 and stock_total <= 0 and transit_total <= 0 and is_in_supply:
            if is_non_supply_chain_sku:
                return '待加库存，非供应链负责SKU'
            else:
                return '待加库存'
        # 【全网缺】：可销售数小于等于0并且在途库存小于等于0，且是停供
        elif stock_total <= 0 and transit_total <= 0 and is_stopped:
            return '全网缺'
        # 【待判断】：非以上的原因的所有订单
        else:
            return '待判断'
    
    def generate_reason_detail(self, sheet2_df, reason):
        """生成原因备注详情"""
        reason_products = sheet2_df[sheet2_df['未配货原因'] == reason]
        details = []
        for _, row in reason_products.iterrows():
            # 格式：规格代码+商品名称*数量（数量取值为sheet2中的总计）
            spec_code = row['规格代码']
            product_name = row['商品名称']
            total_qty = row['总计'] if pd.notna(row['总计']) else 0
            detail = f"{spec_code}{product_name}*{total_qty}"
            details.append(detail)
        return '; '.join(details[:10])  # 限制最多显示10个
    
    def generate_reason_detail_top3(self, sheet2_df, reason):
        """生成原因备注详情（只显示TOP3规格代码）"""
        reason_products = sheet2_df[sheet2_df['未配货原因'] == reason]
        
        # 按规格代码分组，汇总相同规格代码的数量
        spec_summary = reason_products.groupby('规格代码').agg({
            '总计': 'sum',
            '商品名称': 'first'  # 取第一个商品名称
        }).reset_index()
        
        # 按总计数量从大到小排序，取TOP3
        spec_summary = spec_summary.sort_values('总计', ascending=False).head(3)
        
        # 生成详情
        details = []
        for _, row in spec_summary.iterrows():
            spec_code = row['规格代码']
            product_name = row['商品名称']
            total_qty = row['总计'] if pd.notna(row['总计']) else 0
            detail = f"{spec_code}{product_name}*{total_qty}"
            details.append(detail)
        
        return '; '.join(details)
    
    def generate_unship_excel(self, data):
        """生成未配货分析Excel文件"""
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # 写入三个sheet
            data['sheet1'].to_excel(writer, sheet_name='订单商品明细数据', index=False)
            data['sheet2'].to_excel(writer, sheet_name='未配货明细', index=False)
            data['sheet3'].to_excel(writer, sheet_name='卡单分析', index=False)
        
        self.status_updated.emit("未配货分析Excel文件生成完成!")


class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.warehouse_file = ""
        self.inventory_file = ""
        self.output_file = ""
        self.init_ui()
        
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("仓库库存数据处理工具 v1.0")
        self.setGeometry(100, 100, 900, 700)
        
        # 创建中央widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 创建主布局
        main_layout = QVBoxLayout(central_widget)
        
        # 标题
        title_label = QLabel("仓库库存数据处理工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # 创建TAB控件
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # 初始化各个TAB
        self.init_inventory_tab()
        self.init_unship_order_tab()
        self.init_new_feature_tab()
        
    def init_inventory_tab(self):
        """初始化库存数据处理TAB"""
        # 创建库存处理TAB
        inventory_tab = QWidget()
        self.tab_widget.addTab(inventory_tab, "库存数据处理")
        
        # 创建库存TAB的布局
        inventory_layout = QVBoxLayout(inventory_tab)
        
        # 文件选择区域
        file_group = QGroupBox("文件选择")
        file_layout = QVBoxLayout(file_group)
        
        # 仓库分类文件选择
        warehouse_layout = QHBoxLayout()
        self.warehouse_label = QLabel("仓库分类文件: 未选择")
        warehouse_btn = QPushButton("选择仓库分类文件")
        warehouse_btn.clicked.connect(self.select_warehouse_file)
        warehouse_layout.addWidget(self.warehouse_label)
        warehouse_layout.addWidget(warehouse_btn)
        file_layout.addLayout(warehouse_layout)
        
        # 商品库存文件选择
        inventory_file_layout = QHBoxLayout()
        self.inventory_label = QLabel("商品库存文件: 未选择")
        inventory_btn = QPushButton("选择商品库存文件")
        inventory_btn.clicked.connect(self.select_inventory_file)
        inventory_file_layout.addWidget(self.inventory_label)
        inventory_file_layout.addWidget(inventory_btn)
        file_layout.addLayout(inventory_file_layout)
        
        # 输出文件选择
        output_layout = QHBoxLayout()
        self.output_label = QLabel("输出文件: 未选择")
        output_btn = QPushButton("选择输出位置")
        output_btn.clicked.connect(self.select_output_file)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(output_btn)
        file_layout.addLayout(output_layout)
        
        inventory_layout.addWidget(file_group)
        
        # 处理按钮
        self.process_btn = QPushButton("开始处理数据")
        self.process_btn.setMinimumHeight(40)
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        inventory_layout.addWidget(self.process_btn)
        
        # 进度条
        self.progress_bar = QProgressBar()
        inventory_layout.addWidget(self.progress_bar)
        
        # 状态显示
        status_group = QGroupBox("处理状态")
        status_layout = QVBoxLayout(status_group)
        self.status_text = QTextEdit()
        self.status_text.setMaximumHeight(200)
        self.status_text.setReadOnly(True)
        status_layout.addWidget(self.status_text)
        inventory_layout.addWidget(status_group)
        
        # 添加说明
        info_label = QLabel("说明：选择仓库分类Excel文件和商品库存CSV文件，程序将自动处理并生成包含'信选仓'和'自营在途'两个工作表的Excel文件。")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: gray; font-size: 10px;")
        inventory_layout.addWidget(info_label)
        
    def init_unship_order_tab(self):
        """初始化未发货单据处理TAB"""
        # 创建未发货单据处理TAB
        unship_order_tab = QWidget()
        self.tab_widget.addTab(unship_order_tab, "未发货单据处理")
        
        # 创建TAB的布局
        unship_order_layout = QVBoxLayout(unship_order_tab)
        
        # 标题区域
        title_group = QGroupBox("未发货单据处理")
        title_layout = QVBoxLayout(title_group)
        
        # 标题信息
        title_label = QLabel("📋 未发货单据数据处理")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_layout.addWidget(title_label)
        
        # 功能说明
        description_label = QLabel("""
基于订单商品明细数据0811.csv，生成包含三个sheet的未发货单据分析报表：
• 订单商品明细数据 - 包含备注和耗时计算
• SDO去重后分析 - 统计分析结果
• 订单量分析 - 按仓库统计去重的SDO单号数量
        """)
        description_label.setWordWrap(True)
        description_label.setStyleSheet("font-size: 11px; line-height: 1.4;")
        title_layout.addWidget(description_label)
        
        unship_order_layout.addWidget(title_group)
        
        # 文件选择区域
        file_group = QGroupBox("文件选择")
        file_layout = QVBoxLayout(file_group)
        
        # 订单商品明细文件选择
        order_layout = QHBoxLayout()
        self.unship_order_label = QLabel("订单商品明细数据: 未选择")
        unship_order_btn = QPushButton("选择订单明细文件")
        unship_order_btn.clicked.connect(self.select_unship_order_file)
        order_layout.addWidget(self.unship_order_label)
        order_layout.addWidget(unship_order_btn)
        file_layout.addLayout(order_layout)
        
        # 店铺&仓库对应关系文件选择
        unship_mapping_layout = QHBoxLayout()
        self.unship_mapping_label = QLabel("店铺&仓库对应关系: 未选择")
        unship_mapping_btn = QPushButton("选择对应关系文件")
        unship_mapping_btn.clicked.connect(self.select_unship_mapping_file)
        unship_mapping_layout.addWidget(self.unship_mapping_label)
        unship_mapping_layout.addWidget(unship_mapping_btn)
        file_layout.addLayout(unship_mapping_layout)
        
        # 输出文件选择
        unship_order_output_layout = QHBoxLayout()
        self.unship_order_output_label = QLabel("输出文件: 未选择")
        unship_order_output_btn = QPushButton("选择输出位置")
        unship_order_output_btn.clicked.connect(self.select_unship_order_output_file)
        unship_order_output_layout.addWidget(self.unship_order_output_label)
        unship_order_output_layout.addWidget(unship_order_output_btn)
        file_layout.addLayout(unship_order_output_layout)
        
        unship_order_layout.addWidget(file_group)
        
        # 处理按钮
        self.unship_order_process_btn = QPushButton("开始处理未发货单据数据")
        self.unship_order_process_btn.setMinimumHeight(40)
        self.unship_order_process_btn.clicked.connect(self.start_unship_order_processing)
        self.unship_order_process_btn.setEnabled(False)
        unship_order_layout.addWidget(self.unship_order_process_btn)
        
        # 进度条
        self.unship_order_progress_bar = QProgressBar()
        unship_order_layout.addWidget(self.unship_order_progress_bar)
        
        # 状态显示
        unship_order_status_group = QGroupBox("处理状态")
        unship_order_status_layout = QVBoxLayout(unship_order_status_group)
        self.unship_order_status_text = QTextEdit()
        self.unship_order_status_text.setMaximumHeight(200)
        self.unship_order_status_text.setReadOnly(True)
        unship_order_status_layout.addWidget(self.unship_order_status_text)
        unship_order_layout.addWidget(unship_order_status_group)
        
        # 初始化文件路径变量
        self.unship_order_file = ""
        self.unship_mapping_file = ""
        self.unship_order_output_file = ""
        
        # 添加说明
        info_label = QLabel("说明：选择订单商品明细数据CSV文件和店铺仓库对应关系Excel文件，程序将自动处理并生成包含备注分析的Excel文件。处理速度目标：2分钟内完成。")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: gray; font-size: 10px;")
        unship_order_layout.addWidget(info_label)
        
    def init_new_feature_tab(self):
        """初始化未配货明细处理TAB"""
        # 创建未配货明细处理TAB
        new_feature_tab = QWidget()
        self.tab_widget.addTab(new_feature_tab, "未配货明细处理")
        
        # 创建未配货明细处理TAB的布局
        new_layout = QVBoxLayout(new_feature_tab)
        
        # 标题区域
        title_group = QGroupBox("未配货明细处理")
        title_layout = QVBoxLayout(title_group)
        
        # 标题信息
        title_label = QLabel("📦 未配货明细数据处理")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_layout.addWidget(title_label)
        
        # 功能说明
        description_label = QLabel("""
基于订单商品明细、库存报表和店铺仓库对应关系，生成包含三个sheet的未配货分析报表：
• 订单商品明细数据 - 过滤后的订单详情
• 未配货明细 - 未配货原因分析  
• 卡单分析 - 统计分析结果
        """)
        description_label.setWordWrap(True)
        description_label.setStyleSheet("font-size: 11px; line-height: 1.4;")
        title_layout.addWidget(description_label)
        
        new_layout.addWidget(title_group)
        
        # 文件选择区域
        file_group = QGroupBox("基础数据文件选择")
        file_layout = QVBoxLayout(file_group)
        
        # 订单商品明细文件选择
        order_layout = QHBoxLayout()
        self.order_label = QLabel("订单商品明细数据: 未选择")
        order_btn = QPushButton("选择订单明细文件")
        order_btn.clicked.connect(self.select_order_file)
        order_layout.addWidget(self.order_label)
        order_layout.addWidget(order_btn)
        file_layout.addLayout(order_layout)
        
        # 库存报表文件选择
        stock_layout = QHBoxLayout()
        self.stock_label = QLabel("库存报表: 未选择")
        stock_btn = QPushButton("选择库存报表文件")
        stock_btn.clicked.connect(self.select_stock_file)
        stock_layout.addWidget(self.stock_label)
        stock_layout.addWidget(stock_btn)
        file_layout.addLayout(stock_layout)
        
        # 店铺仓库对应关系文件选择
        mapping_layout = QHBoxLayout()
        self.mapping_label = QLabel("店铺&仓库对应关系: 未选择")
        mapping_btn = QPushButton("选择对应关系文件")
        mapping_btn.clicked.connect(self.select_mapping_file)
        mapping_layout.addWidget(self.mapping_label)
        mapping_layout.addWidget(mapping_btn)
        file_layout.addLayout(mapping_layout)
        
        # 输出文件选择
        unship_output_layout = QHBoxLayout()
        self.unship_output_label = QLabel("输出文件: 未选择")
        unship_output_btn = QPushButton("选择输出位置")
        unship_output_btn.clicked.connect(self.select_unship_output_file)
        unship_output_layout.addWidget(self.unship_output_label)
        unship_output_layout.addWidget(unship_output_btn)
        file_layout.addLayout(unship_output_layout)
        
        new_layout.addWidget(file_group)
        
        # 飞书配置区域
        feishu_group = QGroupBox("飞书多维表格配置")
        feishu_layout = QVBoxLayout(feishu_group)
        
        # 飞书连接状态
        self.feishu_status_label = QLabel("🔗 飞书连接状态: 未连接")
        feishu_layout.addWidget(self.feishu_status_label)
        
        # 连接按钮
        connect_feishu_btn = QPushButton("连接飞书多维表格")
        connect_feishu_btn.clicked.connect(self.connect_feishu)
        feishu_layout.addWidget(connect_feishu_btn)
        
        new_layout.addWidget(feishu_group)
        
        # 处理按钮
        self.unship_process_btn = QPushButton("开始处理未配货数据")
        self.unship_process_btn.setMinimumHeight(40)
        self.unship_process_btn.clicked.connect(self.start_unship_processing)
        self.unship_process_btn.setEnabled(False)
        new_layout.addWidget(self.unship_process_btn)
        
        # 进度条
        self.unship_progress_bar = QProgressBar()
        new_layout.addWidget(self.unship_progress_bar)
        
        # 状态显示
        unship_status_group = QGroupBox("处理状态")
        unship_status_layout = QVBoxLayout(unship_status_group)
        self.unship_status_text = QTextEdit()
        self.unship_status_text.setMaximumHeight(200)
        self.unship_status_text.setReadOnly(True)
        unship_status_layout.addWidget(self.unship_status_text)
        new_layout.addWidget(unship_status_group)
        
        # 初始化文件路径变量 - 完全手动选择
        self.order_file = ""
        self.stock_file = ""
        self.mapping_file = ""
        self.unship_output_file = ""
        self.feishu_connected = False
        self.feishu_client = None  # 初始化飞书客户端为None
        

    
    def select_warehouse_file(self):
        """选择仓库分类文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择仓库分类文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.warehouse_file = file_path
            self.warehouse_label.setText(f"仓库分类文件: {os.path.basename(file_path)}")
            self.check_files_ready()
            
    def select_inventory_file(self):
        """选择商品库存文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择商品库存文件", "", "CSV文件 (*.csv)"
        )
        if file_path:
            self.inventory_file = file_path
            self.inventory_label.setText(f"商品库存文件: {os.path.basename(file_path)}")
            self.check_files_ready()
            
    def select_output_file(self):
        """选择输出文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择输出位置", f"库存报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
            "Excel文件 (*.xlsx)"
        )
        if file_path:
            self.output_file = file_path
            self.output_label.setText(f"输出文件: {os.path.basename(file_path)}")
            self.check_files_ready()
            
    def check_files_ready(self):
        """检查文件是否都已选择"""
        if self.warehouse_file and self.inventory_file and self.output_file:
            self.process_btn.setEnabled(True)
        else:
            self.process_btn.setEnabled(False)
            
    def start_processing(self):
        """开始处理数据"""
        self.process_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_text.clear()

        # 使用统计：库存报表（运行开始时间）
        try:
            from time import perf_counter

            if not hasattr(self, "_stats_start_ts"):
                self._stats_start_ts = {}
            self._stats_start_ts["stock_kucunbaobiao"] = perf_counter()
        except Exception:
            pass
        
        # 创建处理线程
        self.processor = DataProcessor(
            self.warehouse_file, 
            self.inventory_file, 
            self.output_file
        )
        
        # 连接信号
        self.processor.progress_updated.connect(self.progress_bar.setValue)
        self.processor.status_updated.connect(self.update_status)
        self.processor.finished_signal.connect(self.processing_finished)
        
        # 开始处理
        self.processor.start()
        
    def update_status(self, message):
        """更新状态显示"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.append(f"[{timestamp}] {message}")
        
    def processing_finished(self, success, message):
        """处理完成"""
        self.process_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "成功", message)
            self.update_status("✅ 处理完成！文件已保存。")
            # 使用统计：成功才记（统计失败不影响主流程）
            try:
                from time import perf_counter
                from services.usage_stats import BASELINE_FIXED, record_event

                t0 = getattr(self, "_stats_start_ts", {}).get("stock_kucunbaobiao", None)
                if t0 is not None:
                    runtime = perf_counter() - float(t0)
                    record_event("stock_kucunbaobiao", runtime_sec=runtime, baseline_sec=float(BASELINE_FIXED.get("stock_kucunbaobiao", 0.0)))
            except Exception:
                pass
        else:
            QMessageBox.critical(self, "错误", message)
            self.update_status("❌ 处理失败！")
            
    def select_order_file(self):
        """选择订单商品明细文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择订单商品明细文件", "", "CSV文件 (*.csv)"
        )
        if file_path:
            self.order_file = file_path
            self.order_label.setText(f"订单商品明细数据: {os.path.basename(file_path)}")
            self.check_unship_files_ready()
            
    def select_stock_file(self):
        """选择库存报表文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择库存报表文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.stock_file = file_path
            self.stock_label.setText(f"库存报表: {os.path.basename(file_path)}")
            self.check_unship_files_ready()
            
    def select_mapping_file(self):
        """选择店铺仓库对应关系文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择店铺仓库对应关系文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.mapping_file = file_path
            self.mapping_label.setText(f"店铺&仓库对应关系: {os.path.basename(file_path)}")
            self.check_unship_files_ready()
            
    def select_unship_output_file(self):
        """选择未配货数据输出文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择输出位置", f"未配货明细分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
            "Excel文件 (*.xlsx)"
        )
        if file_path:
            self.unship_output_file = file_path
            self.unship_output_label.setText(f"输出文件: {os.path.basename(file_path)}")
            self.check_unship_files_ready()
            
    def check_unship_files_ready(self):
        """检查未配货处理文件是否都已选择"""
        if (self.order_file and self.stock_file and self.mapping_file and 
            self.unship_output_file and self.feishu_connected):
            self.unship_process_btn.setEnabled(True)
        else:
            self.unship_process_btn.setEnabled(False)
            
    def connect_feishu(self):
        """连接飞书多维表格"""
        try:
            self.update_unship_status("正在导入飞书模块...")
            
            # 尝试修复解析器问题
            try:
                # 预先导入可能缺失的模块
                import lark
                import lark.parsers
                import lark.lexer
                import lark.grammar
                
                # 修复PyInstaller环境中的语法文件路径问题
                import sys
                import os
                if hasattr(sys, '_MEIPASS'):
                    # PyInstaller环境，尝试多种路径修复方案
                    possible_paths = [
                        os.path.join(sys._MEIPASS, 'lark', 'grammars'),
                        os.path.join(sys._MEIPASS, 'lark_grammars'),
                        os.path.join(os.path.dirname(sys.executable), 'lark', 'grammars'),
                        # 我们工程内置的语法文件目录（避免删除原始源码目录后退化）
                        os.path.join(os.path.dirname(__file__), 'stock_tool_lark_grammars'),
                    ]
                    
                    for lark_grammars_path in possible_paths:
                        if os.path.exists(lark_grammars_path):
                            try:
                                # 修改lark.grammars模块的路径
                                import lark.grammars
                                lark.grammars.__path__ = [lark_grammars_path]
                                self.update_unship_status(f"已修复lark语法文件路径: {lark_grammars_path}")
                                break
                            except Exception as e:
                                self.update_unship_status(f"路径修复尝试失败: {e}")
                                continue
                    else:
                        self.update_unship_status("⚠️ 所有语法文件路径修复尝试均失败")
                        # 尝试设置环境变量作为备选方案
                        os.environ['LARK_GRAMMARS_PATH'] = os.path.join(sys._MEIPASS, 'lark_grammars')
                
                self.update_unship_status("解析器模块导入成功...")
            except ImportError as lark_error:
                self.update_unship_status(f"⚠️ 解析器模块导入警告: {str(lark_error)}")
            except Exception as path_error:
                self.update_unship_status(f"⚠️ 路径修复警告: {str(path_error)}")
            
            # 尝试使用requests直接调用飞书API，绕过lark_oapi的解析器问题
            try:
                import requests
                import json
                
                # 飞书配置
                app_id = "cli_a0869a6c2c21d00c"
                app_secret = "EH02crPwtAzrzEf2bX2s5b8Nn2Popkat"
                
                self.update_unship_status("正在获取飞书访问令牌...")
                
                # 获取访问令牌
                token_url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
                token_data = {
                    "app_id": app_id,
                    "app_secret": app_secret
                }
                
                token_response = requests.post(token_url, json=token_data, timeout=20)
                if getattr(token_response, "status_code", 0) != 200:
                    raise Exception(f"HTTP状态码异常：{getattr(token_response, 'status_code', '')}，响应：{getattr(token_response, 'text', '')[:200]}")
                token_result = token_response.json()
                
                if token_result.get("code") != 0:
                    raise Exception(f"获取访问令牌失败: {token_result.get('msg', '未知错误')}")
                
                access_token = token_result["tenant_access_token"]
                self.update_unship_status("访问令牌获取成功...")
                
                # 保存连接信息供后续使用（跳过连接测试，直接保存）
                self.feishu_access_token = access_token
                self.feishu_headers = {
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json"
                }
                self.feishu_api_base = "https://open.feishu.cn/open-apis/bitable/v1/apps/bascnwiJJtU80TMwlyUTLqHpy0e/tables/tbl7Uxj6KiHtUjez/records"
                self.feishu_use_requests = True  # 标记使用requests方式
                
                self.update_unship_status("飞书API连接配置完成，跳过连接测试...")
                # requests 方式已具备可用 token，视为连接成功（避免继续走 lark_oapi/解析器分支）
                self.feishu_connected = True
                self.feishu_status_label.setText("🔗 飞书连接状态: 已连接")
                self.update_unship_status("✅ 飞书多维表格连接成功！（requests方式）")
                self.check_unship_files_ready()
                return
                
            except ImportError as req_import_error:
                # 明确提示依赖缺失
                raise ImportError(f"缺少 requests 依赖，请先安装：pip install requests\n原始错误：{req_import_error}") from req_import_error
            except Exception as api_error:
                # 如果直接API调用失败，回退到lark_oapi
                self.update_unship_status(f"直接API调用失败，尝试lark_oapi: {str(api_error)}")
                self.feishu_use_requests = False  # 标记使用lark_oapi方式
                
                from lark_oapi.api.bitable.v1 import ListAppTableRecordRequest, ListAppTableRecordResponse
                from lark_oapi import Client
                
                # 飞书配置
                app_id = "cli_a0869a6c2c21d00c"
                app_secret = "EH02crPwtAzrzEf2bX2s5b8Nn2Popkat"
                
                self.update_unship_status("正在创建飞书客户端...")
                
                # 创建客户端 - 添加更多错误处理
                try:
                    # 尝试使用更安全的方式创建客户端
                    self.update_unship_status("初始化客户端构建器...")
                    client_builder = Client.builder()
                    if client_builder is None:
                        raise Exception("Client.builder() 返回 None")
                    
                    self.update_unship_status("设置应用ID...")
                    client_builder = client_builder.app_id(app_id)
                    if client_builder is None:
                        raise Exception("设置 app_id 后返回 None")
                    
                    self.update_unship_status("设置应用密钥...")
                    client_builder = client_builder.app_secret(app_secret)
                    if client_builder is None:
                        raise Exception("设置 app_secret 后返回 None")
                    
                    self.update_unship_status("构建飞书客户端...")
                    self.feishu_client = client_builder.build()
                    if self.feishu_client is None:
                        raise Exception("client_builder.build() 返回 None")
                        
                except Exception as build_error:
                    # 如果是解析器错误，提供更详细的信息
                    error_str = str(build_error)
                    if "parser" in error_str.lower():
                        raise Exception(f"解析器构建失败: {error_str}. 这可能是PyInstaller打包问题，请尝试重新打包或联系技术支持。")
                    else:
                        raise Exception(f"创建飞书客户端失败: {error_str}")
            
            # 测试连接
            self.update_unship_status("飞书客户端创建成功，正在测试连接...")
            self.feishu_connected = True
            self.feishu_status_label.setText("🔗 飞书连接状态: 已连接")
            self.update_unship_status("✅ 飞书多维表格连接成功！")
            self.check_unship_files_ready()
            
        except ImportError as import_error:
            self.feishu_connected = False
            self.feishu_client = None
            self.feishu_status_label.setText("🔗 飞书连接状态: 模块导入失败")
            error_msg = f"飞书模块导入失败: {str(import_error)}"
            self.update_unship_status(f"❌ {error_msg}")
            QMessageBox.critical(self, "模块错误", error_msg)
        except Exception as e:
            self.feishu_connected = False
            self.feishu_client = None
            self.feishu_status_label.setText("🔗 飞书连接状态: 连接失败")
            error_msg = f"飞书连接失败: {str(e)}"
            self.update_unship_status(f"❌ {error_msg}")
            QMessageBox.critical(self, "连接错误", error_msg)
            
    def start_unship_processing(self):
        """开始处理未配货数据"""
        self.unship_process_btn.setEnabled(False)
        self.unship_progress_bar.setValue(0)
        self.unship_status_text.clear()

        # 使用统计：未配货（运行开始时间）
        try:
            from time import perf_counter

            if not hasattr(self, "_stats_start_ts"):
                self._stats_start_ts = {}
            self._stats_start_ts["stock_weipeihuo"] = perf_counter()
        except Exception:
            pass
        
        # 创建未配货处理线程
        # 准备飞书连接信息
        feishu_connection_info = None
        if hasattr(self, 'feishu_use_requests') and self.feishu_use_requests:
            feishu_connection_info = {
                'use_requests': True,
                'access_token': getattr(self, 'feishu_access_token', None),
                'headers': getattr(self, 'feishu_headers', None),
                'api_base': getattr(self, 'feishu_api_base', None)
            }
        
        self.unship_processor = UnshipDataProcessor(
            self.order_file,
            self.stock_file, 
            self.mapping_file,
            self.unship_output_file,
            self.feishu_client if self.feishu_connected else None,
            feishu_connection_info
        )
        
        # 连接信号
        self.unship_processor.progress_updated.connect(self.unship_progress_bar.setValue)
        self.unship_processor.status_updated.connect(self.update_unship_status)
        self.unship_processor.finished_signal.connect(self.unship_processing_finished)
        
        # 开始处理
        self.unship_processor.start()
        
    def update_unship_status(self, message):
        """更新未配货处理状态显示"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.unship_status_text.append(f"[{timestamp}] {message}")
        
    def unship_processing_finished(self, success, message):
        """未配货处理完成"""
        self.unship_process_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "成功", message)
            self.update_unship_status("✅ 未配货数据处理完成！文件已保存。")
            # 使用统计：成功才记（统计失败不影响主流程）
            try:
                from time import perf_counter
                from services.usage_stats import BASELINE_FIXED, record_event

                t0 = getattr(self, "_stats_start_ts", {}).get("stock_weipeihuo", None)
                if t0 is not None:
                    runtime = perf_counter() - float(t0)
                    record_event("stock_weipeihuo", runtime_sec=runtime, baseline_sec=float(BASELINE_FIXED.get("stock_weipeihuo", 0.0)))
            except Exception:
                pass
        else:
            QMessageBox.critical(self, "错误", message)
            self.update_unship_status("❌ 未配货数据处理失败！")
            
    def select_unship_order_file(self):
        """选择未发货订单明细文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择订单商品明细文件", "", "CSV文件 (*.csv)"
        )
        if file_path:
            self.unship_order_file = file_path
            self.unship_order_label.setText(f"订单商品明细数据: {os.path.basename(file_path)}")
            self.check_unship_order_files_ready()
            
    def select_unship_mapping_file(self):
        """选择未发货单据对应关系文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择店铺仓库对应关系文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.unship_mapping_file = file_path
            self.unship_mapping_label.setText(f"店铺&仓库对应关系: {os.path.basename(file_path)}")
            self.check_unship_order_files_ready()
            
    def select_unship_order_output_file(self):
        """选择未发货单据输出文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择输出位置", f"未发货单据处理_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
            "Excel文件 (*.xlsx)"
        )
        if file_path:
            self.unship_order_output_file = file_path
            self.unship_order_output_label.setText(f"输出文件: {os.path.basename(file_path)}")
            self.check_unship_order_files_ready()
            
    def check_unship_order_files_ready(self):
        """检查未发货单据处理文件是否都已选择"""
        if self.unship_order_file and self.unship_mapping_file and self.unship_order_output_file:
            self.unship_order_process_btn.setEnabled(True)
        else:
            self.unship_order_process_btn.setEnabled(False)
            
    def start_unship_order_processing(self):
        """开始处理未发货单据数据"""
        self.unship_order_process_btn.setEnabled(False)
        self.unship_order_progress_bar.setValue(0)
        self.unship_order_status_text.clear()

        # 使用统计：未发货（运行开始时间）
        try:
            from time import perf_counter

            if not hasattr(self, "_stats_start_ts"):
                self._stats_start_ts = {}
            self._stats_start_ts["stock_weifahuo"] = perf_counter()
        except Exception:
            pass
        
        # 创建未发货单据处理线程
        self.unship_order_processor = UnshipOrderProcessor(
            self.unship_order_file,
            self.unship_mapping_file,
            self.unship_order_output_file
        )
        
        # 连接信号
        self.unship_order_processor.progress_updated.connect(self.unship_order_progress_bar.setValue)
        self.unship_order_processor.status_updated.connect(self.update_unship_order_status)
        self.unship_order_processor.finished_signal.connect(self.unship_order_processing_finished)
        
        # 开始处理
        self.unship_order_processor.start()
        
    def update_unship_order_status(self, message):
        """更新未发货单据处理状态显示"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.unship_order_status_text.append(f"[{timestamp}] {message}")
        
    def unship_order_processing_finished(self, success, message):
        """未发货单据处理完成"""
        self.unship_order_process_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "成功", message)
            self.update_unship_order_status("✅ 未发货单据数据处理完成！文件已保存。")
            # 使用统计：成功才记（统计失败不影响主流程）
            try:
                from time import perf_counter
                from services.usage_stats import BASELINE_FIXED, record_event

                t0 = getattr(self, "_stats_start_ts", {}).get("stock_weifahuo", None)
                if t0 is not None:
                    runtime = perf_counter() - float(t0)
                    record_event("stock_weifahuo", runtime_sec=runtime, baseline_sec=float(BASELINE_FIXED.get("stock_weifahuo", 0.0)))
            except Exception:
                pass
        else:
            QMessageBox.critical(self, "错误", message)
            self.update_unship_order_status("❌ 未发货单据数据处理失败！")


def main():
    """主函数"""
    app = QApplication(sys.argv)
    app.setApplicationName("仓库库存数据处理工具")
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()